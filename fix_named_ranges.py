"""
Replace all named range references with direct absolute cell references
for Google Sheets compatibility, then delete the named ranges.
"""
import os
import re
from openpyxl import load_workbook

os.chdir(os.path.join(os.path.dirname(__file__), '2026-pricing-strategy', 'docs', 'templates'))
wb = load_workbook('QC_Estimate_Template_2026_v2_FINAL.xlsx')

# ── Step 1: Record all defined names and their ranges ──
print('=== STEP 1: Defined Names ===')
# All names currently point to 'Client Setup - Blank' ranges.
# The cell ranges (without sheet) are the same for both Sample and Blank.
NAMED_RANGES = {}
for name, dn in wb.defined_names.items():
    for title, coord in dn.destinations:
        NAMED_RANGES[name] = coord  # e.g. '$A$23:$B$33'
        print(f'  {name} = {title}!{coord}')

# ── Step 2: Build replacement map per sheet ──
# Determine which Client Setup sheet each formula sheet should reference
def get_cs_sheet(sheet_title):
    if 'Sample' in sheet_title:
        return 'Client Setup - Sample'
    elif 'Blank' in sheet_title:
        return 'Client Setup - Blank'
    elif sheet_title == 'Client Setup - Sample':
        return 'Client Setup - Sample'
    elif sheet_title == 'Client Setup - Blank':
        return 'Client Setup - Blank'
    return None

# Sort names longest-first to avoid partial matches
sorted_names = sorted(NAMED_RANGES.keys(), key=len, reverse=True)
print(f'\nReplacement order (longest first): {sorted_names}')

def replace_named_ranges_in_formula(formula, sheet_title):
    """Replace named range references with direct absolute references."""
    cs = get_cs_sheet(sheet_title)
    if cs is None:
        return formula
    result = formula
    for name in sorted_names:
        if name in result:
            # Use word boundary check to avoid partial matches
            # Named ranges in formulas appear as bare words (not inside quotes)
            direct_ref = f"'{cs}'!{NAMED_RANGES[name]}"
            # Replace only whole-word occurrences (not inside other identifiers)
            result = re.sub(r'\b' + re.escape(name) + r'\b', direct_ref, result)
    return result

# ── Step 2: Replace in all formulas ──
print('\n=== STEP 2: Replacing named ranges in formulas ===')
formula_count = 0
for ws in wb.worksheets:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith('='):
                has_name = any(name in v for name in sorted_names)
                if has_name:
                    new_v = replace_named_ranges_in_formula(v, ws.title)
                    if new_v != v:
                        cell.value = new_v
                        formula_count += 1
print(f'  Replaced named ranges in {formula_count} formulas.')

# ── Step 3: Replace in data validations ──
print('\n=== STEP 3: Replacing named ranges in data validations ===')
dv_count = 0
for ws in wb.worksheets:
    cs = get_cs_sheet(ws.title)
    if cs is None:
        continue
    for dv in ws.data_validations.dataValidation:
        f1 = dv.formula1 or ''
        if any(name in f1 for name in sorted_names):
            for name in sorted_names:
                if name in f1:
                    direct_ref = f"'{cs}'!{NAMED_RANGES[name]}"
                    f1 = re.sub(r'\b' + re.escape(name) + r'\b', direct_ref, f1)
            dv.formula1 = f1
            dv_count += 1
            print(f'  {ws.title} sqref={dv.sqref} -> formula1={f1}')
print(f'  Updated {dv_count} data validations.')

# ── Step 4: Delete all named ranges ──
print('\n=== STEP 4: Deleting named ranges ===')
names_to_delete = list(wb.defined_names.keys())
for name in names_to_delete:
    del wb.defined_names[name]
    print(f'  Deleted: {name}')
print(f'  Remaining defined names: {len(wb.defined_names)}')

# ── Step 5: Save and verify ──
wb.save('QC_Estimate_Template_2026_v2_FINAL.xlsx')
print('\n=== STEP 5: Verification ===')

wb2 = load_workbook('QC_Estimate_Template_2026_v2_FINAL.xlsx')
all_names = list(NAMED_RANGES.keys())

# Check no formulas contain named range strings
print('\nChecking formulas for remaining named ranges...')
remaining_formulas = 0
for ws in wb2.worksheets:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith('='):
                for name in all_names:
                    if re.search(r'\b' + re.escape(name) + r'\b', v):
                        print(f'  REMAINING: {ws.title}!{cell.coordinate}: {v}')
                        remaining_formulas += 1
                        break
print(f'  Formulas with remaining named ranges: {remaining_formulas}')

# Check no data validations contain named range strings
print('\nChecking data validations for remaining named ranges...')
remaining_dvs = 0
for ws in wb2.worksheets:
    for dv in ws.data_validations.dataValidation:
        f1 = dv.formula1 or ''
        for name in all_names:
            if re.search(r'\b' + re.escape(name) + r'\b', f1):
                print(f'  REMAINING: {ws.title} sqref={dv.sqref} formula1={f1}')
                remaining_dvs += 1
                break
print(f'  Data validations with remaining named ranges: {remaining_dvs}')

# Check no defined names remain
print(f'\nDefined names remaining: {len(wb2.defined_names)}')

# Spot check with data_only
print('\nSpot-checking computed values (data_only=True)...')
wb3 = load_workbook('QC_Estimate_Template_2026_v2_FINAL.xlsx', data_only=True)

checks = [
    ('Venue Estimate - Sample', 'D43', 14000, 'F&B Subtotal'),
    ('Venue Estimate - Sample', 'D51', 2800, 'Service Charge'),
    ('Venue Estimate - Sample', 'D47', 102, 'Equip Tax'),
]
for sheet, ref, expected, label in checks:
    val = wb3[sheet][ref].value
    status = 'OK' if val == expected else f'GOT {val!r} (expected {expected})'
    print(f'  {sheet}!{ref} ({label}): {status}')

# Decor F104 nonzero
val = wb3['Decor Estimate - Sample']['F104'].value
print(f'  Decor Sample F104: {val!r} ({"OK nonzero" if val and val != 0 else "PROBLEM"})')

# Blank tabs should be 0 or None (no errors)
for sheet in ['Venue Estimate - Blank', 'Decor Estimate - Blank']:
    ws = wb3[sheet]
    total_ref = 'E57' if 'Venue' in sheet else 'F104'
    val = ws[total_ref].value
    print(f'  {sheet}!{total_ref}: {val!r} ({"OK" if val in (0, None) else "CHECK"})')

print('\nDone.')
