#!/usr/bin/env python3
"""
Phase 3: Highlight all dropdown cells with light blue fill.
Add color legend to each estimate tab.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import range_boundaries, coordinate_to_tuple, get_column_letter

SRC = '/Users/alex/2026-pricing-strategy/2026-pricing-strategy/docs/templates/QC_Estimate_Template_2026_v2.xlsx'
wb = openpyxl.load_workbook(SRC)

DROPDOWN_FILL = PatternFill('solid', fgColor='FFD6EAF8')  # Light blue
THIN = Side(style='thin', color='FFD4C5B0')
THIN_BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

FONT_LEGEND_HDR = Font(name='Playfair Display', size=9, bold=True, color='FF464543')
FONT_LEGEND = Font(name='Playfair Display', size=9, color='FF464543')
FILL_LEGEND_WHITE = PatternFill('solid', fgColor='FFFFFFFF')
AC = Alignment(horizontal='center', vertical='center')
AL = Alignment(horizontal='left', vertical='center')


def get_dropdown_cells(ws):
    """Return set of (row, col) tuples for all dropdown-validated cells."""
    cells = set()
    for dv in ws.data_validations.dataValidation:
        if dv.type != 'list':
            continue
        for sq_range in str(dv.sqref).split():
            if ':' in sq_range:
                min_col, min_row, max_col, max_row = range_boundaries(sq_range)
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        cells.add((r, c))
            else:
                r, c = coordinate_to_tuple(sq_range)
                cells.add((r, c))
    return cells


def apply_dropdown_fill(ws, sheet_name):
    """Apply light blue fill to all dropdown cells. Preserve font and border."""
    cells = get_dropdown_cells(ws)
    for r, c in cells:
        cell = ws.cell(row=r, column=c)
        cell.fill = DROPDOWN_FILL
    print(f"  {sheet_name}: {len(cells)} dropdown cells highlighted")
    return len(cells)


def is_merged(ws, row, col):
    """Check if a cell is part of a merged range."""
    for mr in ws.merged_cells.ranges:
        if (row, col) in [(r, c) for r in range(mr.min_row, mr.max_row + 1)
                          for c in range(mr.min_col, mr.max_col + 1)]:
            if not (row == mr.min_row and col == mr.min_col):
                return True
    return False


def add_legend(ws, row, col):
    """Add a color legend at the specified position. Skip merged cells."""
    # Find a safe starting row (avoid merged cells)
    while any(is_merged(ws, row + i, col + j) for i in range(3) for j in range(2)):
        row += 1
        if row > 20:
            print(f"    WARNING: Could not find safe legend spot, skipping")
            return

    # Header
    cell = ws.cell(row=row, column=col)
    cell.value = "Color"
    cell.font = FONT_LEGEND_HDR
    cell.alignment = AC
    cell.border = THIN_BORDER

    cell = ws.cell(row=row, column=col + 1)
    cell.value = "Meaning"
    cell.font = FONT_LEGEND_HDR
    cell.alignment = AL
    cell.border = THIN_BORDER

    # Row 1: Light blue = dropdown
    cell = ws.cell(row=row + 1, column=col)
    cell.fill = DROPDOWN_FILL
    cell.border = THIN_BORDER
    cell = ws.cell(row=row + 1, column=col + 1)
    cell.value = "Dropdown \u2014 click to select"
    cell.font = FONT_LEGEND
    cell.alignment = AL
    cell.border = THIN_BORDER

    # Row 2: White = manual/formula
    cell = ws.cell(row=row + 2, column=col)
    cell.fill = FILL_LEGEND_WHITE
    cell.border = THIN_BORDER
    cell = ws.cell(row=row + 2, column=col + 1)
    cell.value = "Manual entry or formula"
    cell.font = FONT_LEGEND
    cell.alignment = AL
    cell.border = THIN_BORDER

    print(f"    Legend placed at {get_column_letter(col)}{row}")


# ════════════════════════════════════════════════════════════════
# Apply dropdown highlights
# ════════════════════════════════════════════════════════════════
print("Applying dropdown fills...")
total = 0

for sn in wb.sheetnames:
    ws = wb[sn]
    total += apply_dropdown_fill(ws, sn)

# ════════════════════════════════════════════════════════════════
# Add legends to estimate tabs
# ════════════════════════════════════════════════════════════════
print("\nAdding legends...")

# Venue Estimate: put legend at H2 (right of venue details area)
add_legend(wb['Venue Estimate'], 2, 8)  # H2:I4
print("  Venue Estimate: legend at H2:I4")

# Decor Estimate: put legend at J2 (right of line item headers)
add_legend(wb['Decor Estimate'], 2, 10)  # J2:K4
print("  Decor Estimate: legend at J2:K4")

# SAMPLE Decor Estimate: same position
add_legend(wb['SAMPLE Decor Estimate'], 2, 10)  # J2:K4
print("  SAMPLE Decor Estimate: legend at J2:K4")

# Client Setup: put legend at H2
add_legend(wb['Client Setup'], 2, 8)  # H2:I4
print("  Client Setup: legend at H2:I4")

# ════════════════════════════════════════════════════════════════
# Save & Verify
# ════════════════════════════════════════════════════════════════
print(f"\nSaving... ({total} total dropdown cells)")
wb.save(SRC)

# Verify
print("\n" + "=" * 60)
print("VERIFICATION")
print("=" * 60)

wb2 = openpyxl.load_workbook(SRC)
for sn in wb2.sheetnames:
    ws = wb2[sn]
    dd_cells = get_dropdown_cells(ws)
    blue_count = 0
    non_dd_blue = 0
    for r, c in dd_cells:
        cell = ws.cell(row=r, column=c)
        if cell.fill.fgColor and 'D6EAF8' in str(cell.fill.fgColor.rgb):
            blue_count += 1
    # Check for false positives (non-dropdown cells with blue fill)
    for row in range(1, min(ws.max_row + 1, 200)):
        for col in range(1, min(ws.max_column + 1, 20)):
            cell = ws.cell(row=row, column=col)
            if cell.fill.fgColor and 'D6EAF8' in str(cell.fill.fgColor.rgb):
                if (row, col) not in dd_cells:
                    # Check if it's a legend cell (expected)
                    non_dd_blue += 1
    # Subtract legend cells (3 per tab that has a legend)
    legend_cells = 1  # The blue swatch in the legend
    non_dd_blue = max(0, non_dd_blue - legend_cells)

    print(f"  {sn}: {blue_count}/{len(dd_cells)} dropdown cells highlighted"
          f"{f', {non_dd_blue} unexpected blue cells' if non_dd_blue else ''}")

print(f"\n{'=' * 60}")
print("DROPDOWN INDICATORS COMPLETE")
print(f"{'=' * 60}")
