#!/usr/bin/env python3
"""
Audit script for QC_Estimate_Template_2026_v2.xlsx
- PART 1: Blue Fill Audit (D6EAF8)
- PART 2: Border Audit
- PART 3: Section Fill Patterns
"""

import openpyxl
from openpyxl.utils import get_column_letter
from collections import defaultdict, Counter

FILEPATH = "/Users/alex/2026-pricing-strategy/2026-pricing-strategy/docs/templates/QC_Estimate_Template_2026_v2.xlsx"

wb = openpyxl.load_workbook(FILEPATH, data_only=False)


def is_blue_fill(cell):
    """Check if cell has blue fill with D6EAF8 in the RGB."""
    fill = cell.fill
    if fill is None:
        return False
    # Check patternFill fgColor and bgColor
    if fill.patternType is not None:
        fg = fill.fgColor
        bg = fill.bgColor
        for color in [fg, bg]:
            if color and color.rgb and isinstance(color.rgb, str):
                if "D6EAF8" in color.rgb.upper():
                    return True
    return False


def is_dropdown_cell(ws, cell_coord):
    """Check if a cell falls within any data validation with type='list'."""
    for dv in ws.data_validations.dataValidation:
        if dv.type == "list":
            for rng in dv.sqref.ranges:
                if cell_coord in rng:
                    return True
            # Also check if sqref is a single cell string match
            if str(cell_coord) in str(dv.sqref):
                return True
    return False


def get_section_label(ws_title, row):
    """Return a rough section label based on tab and row number."""
    return f"Row {row}"


def get_fill_color_str(cell):
    """Return a string representation of the cell's fill color."""
    fill = cell.fill
    if fill is None or fill.patternType is None:
        return "none"
    if fill.patternType == "none":
        return "none"
    colors = []
    fg = fill.fgColor
    bg = fill.bgColor
    if fg and fg.rgb and isinstance(fg.rgb, str) and fg.rgb != "00000000":
        colors.append(f"fg={fg.rgb}")
    if bg and bg.rgb and isinstance(bg.rgb, str) and bg.rgb != "00000000":
        colors.append(f"bg={bg.rgb}")
    if fg and fg.theme is not None:
        colors.append(f"fg_theme={fg.theme}+tint={fg.tint}")
    if bg and bg.theme is not None:
        colors.append(f"bg_theme={bg.theme}+tint={bg.tint}")
    if fg and fg.indexed is not None:
        colors.append(f"fg_indexed={fg.indexed}")
    if bg and bg.indexed is not None:
        colors.append(f"bg_indexed={bg.indexed}")
    if not colors:
        return f"pattern={fill.patternType} (no explicit color)"
    return f"pattern={fill.patternType} " + ", ".join(colors)


def get_border_info(cell):
    """Return border details for a cell."""
    border = cell.border
    if border is None:
        return {"sides": 0, "styles": {}, "colors": {}}

    sides_with_style = 0
    styles = {}
    colors = {}

    for side_name in ["left", "right", "top", "bottom"]:
        side = getattr(border, side_name, None)
        if side and side.style:
            sides_with_style += 1
            styles[side_name] = side.style
            if side.color:
                if side.color.rgb and isinstance(side.color.rgb, str):
                    colors[side_name] = side.color.rgb
                elif side.color.theme is not None:
                    colors[side_name] = f"theme={side.color.theme}+tint={side.color.tint}"
                elif side.color.indexed is not None:
                    colors[side_name] = f"indexed={side.color.indexed}"

    return {"sides": sides_with_style, "styles": styles, "colors": colors}


# ============================================================================
# PART 1: Blue Fill Audit
# ============================================================================
print("=" * 100)
print("PART 1: BLUE FILL AUDIT (D6EAF8)")
print("=" * 100)

for ws_name in wb.sheetnames:
    ws = wb[ws_name]
    blue_cells = []

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if is_blue_fill(cell):
                coord = cell.coordinate
                dropdown = is_dropdown_cell(ws, coord)
                blue_cells.append({
                    "coord": coord,
                    "row": cell.row,
                    "col": cell.column,
                    "is_dropdown": dropdown,
                    "value": cell.value,
                })

    if blue_cells:
        print(f"\n--- Tab: {ws_name} --- ({len(blue_cells)} blue cells found)")
        print(f"  {'Cell':<8} {'Dropdown?':<12} {'Value':<50} {'Section'}")
        print(f"  {'----':<8} {'---------':<12} {'-----':<50} {'-------'}")
        for bc in blue_cells:
            val_str = str(bc["value"])[:48] if bc["value"] is not None else "(empty)"
            dropdown_str = "YES dropdown" if bc["is_dropdown"] else "NOT dropdown"
            section = get_section_label(ws_name, bc["row"])
            print(f"  {bc['coord']:<8} {dropdown_str:<12} {val_str:<50} {section}")
    else:
        print(f"\n--- Tab: {ws_name} --- (0 blue cells found)")


# ============================================================================
# PART 2: Border Audit
# ============================================================================
print("\n\n" + "=" * 100)
print("PART 2: BORDER AUDIT")
print("=" * 100)

for ws_name in wb.sheetnames:
    ws = wb[ws_name]
    print(f"\n{'=' * 80}")
    print(f"Tab: {ws_name}")
    print(f"{'=' * 80}")

    # Find all rows that have content
    rows_with_content = set()
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                rows_with_content.add(cell.row)

    # Also add rows near content rows (table regions)
    extended_rows = set()
    for r in rows_with_content:
        for delta in range(-1, 2):
            if r + delta >= 1:
                extended_rows.add(r + delta)
    all_rows = sorted(extended_rows)

    if not all_rows:
        print("  (no content)")
        continue

    # Group rows into contiguous sections
    sections = []
    current_section = [all_rows[0]]
    for i in range(1, len(all_rows)):
        if all_rows[i] - all_rows[i - 1] <= 2:
            current_section.append(all_rows[i])
        else:
            sections.append(current_section)
            current_section = [all_rows[i]]
    sections.append(current_section)

    total_full = 0
    total_partial = 0
    total_none = 0
    all_styles = Counter()
    all_border_colors = Counter()

    for section_rows in sections:
        sec_start = section_rows[0]
        sec_end = section_rows[-1]

        full_border = 0
        partial_border = 0
        no_border = 0
        sec_styles = Counter()
        sec_colors = Counter()
        cells_examined = 0

        for r in range(sec_start, sec_end + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                if cell.value is None:
                    # Check if it has any formatting (border/fill) — include if so
                    bi = get_border_info(cell)
                    fill_str = get_fill_color_str(cell)
                    if bi["sides"] == 0 and fill_str == "none":
                        continue

                cells_examined += 1
                bi = get_border_info(cell)

                if bi["sides"] == 4:
                    full_border += 1
                elif bi["sides"] > 0:
                    partial_border += 1
                else:
                    no_border += 1

                for side, style in bi["styles"].items():
                    sec_styles[style] += 1
                for side, color in bi["colors"].items():
                    sec_colors[color] += 1

        total_full += full_border
        total_partial += partial_border
        total_none += no_border
        all_styles.update(sec_styles)
        all_border_colors.update(sec_colors)

        print(f"\n  Section rows {sec_start}-{sec_end} ({cells_examined} cells with content/formatting):")
        print(f"    Full borders (4 sides):   {full_border}")
        print(f"    Partial borders:          {partial_border}")
        print(f"    No borders:               {no_border}")
        if sec_styles:
            print(f"    Border styles used:        {dict(sec_styles)}")
        if sec_colors:
            print(f"    Border colors used:        {dict(sec_colors)}")

    print(f"\n  --- TAB TOTALS for '{ws_name}' ---")
    print(f"    Full borders:   {total_full}")
    print(f"    Partial borders: {total_partial}")
    print(f"    No borders:      {total_none}")
    print(f"    All styles:      {dict(all_styles)}")
    print(f"    All colors:      {dict(all_border_colors)}")


# ============================================================================
# PART 3: Section Fill Patterns
# ============================================================================
print("\n\n" + "=" * 100)
print("PART 3: SECTION FILL PATTERNS")
print("=" * 100)

# Known named fills for reference
NAMED_FILLS = {
    "FFF8F0": "peach/input",
    "F5F0EB": "beige/calc",
    "ECDFCE": "header",
    "E8D9C8": "subheader",
    "D6EAF8": "blue/dropdown",
    "FFFFFF": "white",
    "000000": "black",
}


def classify_fill_color(color_str):
    """Try to match a fill color string to a named pattern."""
    for hex_code, name in NAMED_FILLS.items():
        if hex_code in color_str.upper():
            return f"{name} ({hex_code})"
    return color_str


for ws_name in wb.sheetnames:
    ws = wb[ws_name]
    print(f"\n{'=' * 80}")
    print(f"Tab: {ws_name}")
    print(f"{'=' * 80}")

    # Collect fill info per row
    row_fills = defaultdict(lambda: Counter())

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            fill_str = get_fill_color_str(cell)
            if fill_str != "none":
                row_fills[cell.row][fill_str] += 1

    if not row_fills:
        print("  (no fills found)")
        continue

    sorted_rows = sorted(row_fills.keys())

    # Group consecutive rows with same predominant fill
    def get_predominant(counter):
        if not counter:
            return "none"
        return counter.most_common(1)[0][0]

    groups = []
    current_start = sorted_rows[0]
    current_end = sorted_rows[0]
    current_fills = Counter()
    current_fills.update(row_fills[sorted_rows[0]])
    current_predominant = get_predominant(row_fills[sorted_rows[0]])

    for i in range(1, len(sorted_rows)):
        r = sorted_rows[i]
        r_predominant = get_predominant(row_fills[r])

        if r - current_end <= 1 and r_predominant == current_predominant:
            current_end = r
            current_fills.update(row_fills[r])
        else:
            groups.append((current_start, current_end, current_fills, current_predominant))
            current_start = r
            current_end = r
            current_fills = Counter()
            current_fills.update(row_fills[r])
            current_predominant = r_predominant

    groups.append((current_start, current_end, current_fills, current_predominant))

    # Print row-by-row detail first
    print(f"\n  Row-by-row fill detail:")
    print(f"  {'Row':<6} {'Fill Colors (count)'}")
    print(f"  {'---':<6} {'------------------'}")
    for r in sorted_rows:
        fills = row_fills[r]
        fill_parts = []
        for fill_str, count in fills.most_common():
            classified = classify_fill_color(fill_str)
            fill_parts.append(f"{classified} x{count}")
        print(f"  {r:<6} {'; '.join(fill_parts)}")

    # Print grouped summary
    print(f"\n  Grouped fill sections:")
    print(f"  {'Rows':<16} {'Predominant Fill':<60} {'All Fills'}")
    print(f"  {'----':<16} {'----------------':<60} {'---------'}")
    for start, end, fills, predom in groups:
        range_str = f"{start}-{end}" if start != end else str(start)
        classified_predom = classify_fill_color(predom)
        all_fills = "; ".join(
            f"{classify_fill_color(f)} x{c}" for f, c in fills.most_common(5)
        )
        print(f"  {range_str:<16} {classified_predom:<60} {all_fills}")


print("\n\n" + "=" * 100)
print("AUDIT COMPLETE")
print("=" * 100)
