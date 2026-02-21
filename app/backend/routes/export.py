"""Export routes — generate client-ready Excel/PDF from estimates."""

import io
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from ..database import get_db, Estimate, AuditLog
from ..pricing_engine import calculate_estimate, EstimateInput, LineItem, LineItemType
import json

router = APIRouter(prefix="/api/export", tags=["export"])


def _db_to_engine_input(estimate: Estimate) -> EstimateInput:
    line_items = []
    for li in estimate.line_items:
        line_items.append(LineItem(
            name=li.name,
            category_key=li.category_key,
            item_type=LineItemType(li.item_type),
            unit_price=li.unit_price,
            quantity=li.quantity,
            is_taxable=li.is_taxable,
            notes=li.notes,
            section=li.section,
        ))
    return EstimateInput(
        estimate_name=estimate.estimate_name,
        estimate_type=estimate.estimate_type,
        client_name=estimate.client_name,
        event_name=estimate.event_name,
        event_date=estimate.event_date,
        event_time=estimate.event_time,
        guest_count=estimate.guest_count,
        location=estimate.location,
        tax_rate=estimate.tax_rate,
        commission_scenario=estimate.commission_scenario,
        cc_processing_pct=estimate.cc_processing_pct,
        gratuity_pct=estimate.gratuity_pct,
        admin_fee_pct=estimate.admin_fee_pct,
        gdp_enabled=estimate.gdp_enabled,
        opex_hours=estimate.opex_hours,
        line_items=line_items,
    )


@router.get("/{estimate_id}/excel")
def export_excel(estimate_id: int, db: Session = Depends(get_db)):
    """Export a client-ready Excel file for the estimate."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    estimate = db.query(Estimate).filter(Estimate.id == estimate_id).first()
    if not estimate:
        raise HTTPException(status_code=404, detail="Estimate not found")

    engine_input = _db_to_engine_input(estimate)
    result = calculate_estimate(engine_input)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estimate"

    # Styles
    header_font = Font(name="Calibri", size=14, bold=True, color="FF2C3E50")
    subheader_font = Font(name="Calibri", size=11, bold=True, color="FF34495E")
    label_font = Font(name="Calibri", size=10, color="FF555555")
    value_font = Font(name="Calibri", size=10, bold=True)
    currency_fmt = '#,##0.00'
    pct_fmt = '0.0%'
    header_fill = PatternFill("solid", fgColor="FFF8F9FA")
    section_fill = PatternFill("solid", fgColor="FFEEF2F7")
    thin_border = Border(bottom=Side(style="thin", color="FFCCCCCC"))

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    row = 1

    # Header
    ws.cell(row=row, column=1, value="QUILL CREATIVE EVENT DESIGN").font = header_font
    row += 1
    ws.cell(row=row, column=1, value=result["estimate_name"]).font = subheader_font
    row += 2

    # Event details
    details = [
        ("Client", result["client_name"]),
        ("Event", result["event_name"]),
        ("Date", result["event_date"]),
        ("Location", result["location"]),
        ("Guests", result["guest_count"]),
    ]
    for label, value in details:
        if value:
            ws.cell(row=row, column=1, value=f"{label}:").font = label_font
            ws.cell(row=row, column=2, value=value).font = value_font
            row += 1
    row += 1

    # Client-ready summary (rounded)
    ws.cell(row=row, column=1, value="ESTIMATE SUMMARY").font = subheader_font
    for c in range(1, 6):
        ws.cell(row=row, column=c).fill = section_fill
    row += 1

    ws.cell(row=row, column=1, value="Category").font = label_font
    ws.cell(row=row, column=2, value="Amount").font = label_font
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2).border = thin_border
    row += 1

    for section_name, amount in result["client_ready"]["sections"].items():
        ws.cell(row=row, column=1, value=section_name).font = value_font
        cell = ws.cell(row=row, column=2, value=amount)
        cell.font = value_font
        cell.number_format = currency_fmt
        row += 1

    # Tax
    ws.cell(row=row, column=1, value="Tax").font = value_font
    cell = ws.cell(row=row, column=2, value=result["client_ready"]["tax"])
    cell.font = value_font
    cell.number_format = currency_fmt
    row += 1

    # Total
    ws.cell(row=row, column=1, value="TOTAL").font = Font(name="Calibri", size=12, bold=True, color="FF2C3E50")
    cell = ws.cell(row=row, column=2, value=result["client_ready"]["total"])
    cell.font = Font(name="Calibri", size=12, bold=True, color="FF2C3E50")
    cell.number_format = currency_fmt
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = Border(top=Side(style="double", color="FF2C3E50"))

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"QC_Estimate_{estimate.client_name or 'Draft'}_{estimate.event_date or 'undated'}.xlsx"
    filename = filename.replace(" ", "_")

    # Log export
    audit = AuditLog(estimate_id=estimate_id, action="exported", details=json.dumps({"format": "excel"}))
    db.add(audit)
    db.commit()

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
