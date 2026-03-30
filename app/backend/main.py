"""Quill Creative Pricing Engine — FastAPI application."""

from __future__ import annotations
from pathlib import Path
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from io import BytesIO

from . import pricing_engine, storage, config_loader
from .schemas import EstimateCreate, EstimateUpdate, CalculateRequest

app = FastAPI(
    title="Quill Creative Pricing Engine",
    version="1.0.0",
    description="Config-driven pricing engine for Quill Creative Event Design",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Serve built frontend static files
FRONTEND_DIST = Path(__file__).resolve().parent.parent / "frontend" / "dist"


# ── Health check ─────────────────────────────────────────────────────────

@app.get("/api/health")
def health_check():
    return {"status": "ok", "version": "1.0.0"}


# ── Config endpoints ─────────────────────────────────────────────────────

@app.get("/api/config")
def get_config():
    """Return all pricing configuration."""
    return config_loader.load_all_config()


@app.get("/api/config/categories")
def get_categories():
    """Return markup categories for dropdowns."""
    return pricing_engine.get_category_options()


@app.get("/api/config/templates")
def get_templates():
    """Return all estimate type templates."""
    return pricing_engine.get_estimate_templates()


# ── Calculate (stateless) ────────────────────────────────────────────────

@app.post("/api/calculate")
def calculate(req: CalculateRequest):
    """Run pricing calculation without saving. Used for live preview."""
    items = [item.model_dump() for item in req.line_items]
    try:
        result = pricing_engine.calculate_estimate(
            line_items=items,
            commission_scenario=req.commission_scenario,
            opex_hours=req.opex_hours,
            management_fee=req.management_fee,
            travel_expenses=req.travel_expenses,
            tax_rate_pct=req.tax_rate_pct,
        )
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


# ── Estimate CRUD ────────────────────────────────────────────────────────

@app.get("/api/estimates")
def list_estimates():
    """List all saved estimates with calculated summaries."""
    estimates = storage.get_all_estimates()
    results = []
    for est in estimates:
        # Recalculate P&L from stored line items
        try:
            calc = pricing_engine.calculate_estimate(
                line_items=est.get("line_items", []),
                commission_scenario=est.get("commission_scenario", "direct_client"),
                opex_hours=est.get("opex_hours", 0),
                management_fee=est.get("management_fee", 0),
                travel_expenses=est.get("travel_expenses", 0),
                tax_rate_pct=est.get("tax_rate_pct", 0),
            )
            results.append({**est, "calculated": calc["summary"], "health": calc["health"]})
        except Exception:
            results.append({**est, "calculated": None, "health": None})
    return results


@app.post("/api/estimates", status_code=201)
def create_estimate(req: EstimateCreate):
    """Create and save a new estimate."""
    data = req.model_dump()
    data["line_items"] = [item.model_dump() for item in req.line_items]
    est = storage.create_estimate(data)

    # Calculate P&L
    calc = pricing_engine.calculate_estimate(
        line_items=est["line_items"],
        commission_scenario=est["commission_scenario"],
        opex_hours=est["opex_hours"],
        management_fee=est["management_fee"],
        travel_expenses=est["travel_expenses"],
        tax_rate_pct=est["tax_rate_pct"],
    )
    return {**est, "calculated": calc["summary"], "health": calc["health"]}


@app.get("/api/estimates/{estimate_id}")
def get_estimate(estimate_id: str):
    """Get a single estimate with full calculated breakdown."""
    est = storage.get_estimate(estimate_id)
    if not est:
        raise HTTPException(status_code=404, detail="Estimate not found")

    calc = pricing_engine.calculate_estimate(
        line_items=est.get("line_items", []),
        commission_scenario=est.get("commission_scenario", "direct_client"),
        opex_hours=est.get("opex_hours", 0),
        management_fee=est.get("management_fee", 0),
        travel_expenses=est.get("travel_expenses", 0),
        tax_rate_pct=est.get("tax_rate_pct", 0),
    )
    return {**est, "calculated": calc}


@app.put("/api/estimates/{estimate_id}")
def update_estimate(estimate_id: str, req: EstimateUpdate):
    """Update an existing estimate."""
    updates = req.model_dump(exclude_none=True)
    if req.line_items is not None:
        updates["line_items"] = [item.model_dump() for item in req.line_items]

    est = storage.update_estimate(estimate_id, updates)
    if not est:
        raise HTTPException(status_code=404, detail="Estimate not found")

    calc = pricing_engine.calculate_estimate(
        line_items=est.get("line_items", []),
        commission_scenario=est.get("commission_scenario", "direct_client"),
        opex_hours=est.get("opex_hours", 0),
        management_fee=est.get("management_fee", 0),
        travel_expenses=est.get("travel_expenses", 0),
        tax_rate_pct=est.get("tax_rate_pct", 0),
    )
    return {**est, "calculated": calc}


@app.delete("/api/estimates/{estimate_id}")
def delete_estimate(estimate_id: str):
    """Delete an estimate."""
    if not storage.delete_estimate(estimate_id):
        raise HTTPException(status_code=404, detail="Estimate not found")
    return {"deleted": True}


# ── Excel export ─────────────────────────────────────────────────────────

@app.get("/api/estimates/{estimate_id}/export/excel")
def export_estimate_excel(estimate_id: str):
    """Export an estimate as a client-ready Excel file."""
    est = storage.get_estimate(estimate_id)
    if not est:
        raise HTTPException(status_code=404, detail="Estimate not found")

    calc = pricing_engine.calculate_estimate(
        line_items=est.get("line_items", []),
        commission_scenario=est.get("commission_scenario", "direct_client"),
        opex_hours=est.get("opex_hours", 0),
        management_fee=est.get("management_fee", 0),
        travel_expenses=est.get("travel_expenses", 0),
        tax_rate_pct=est.get("tax_rate_pct", 0),
    )

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, numbers

    wb = Workbook()
    ws = wb.active
    ws.title = "Estimate"

    # Styles
    header_font = Font(name="Calibri", size=12, bold=True, color="4A3728")
    header_fill = PatternFill(start_color="F5E6D3", end_color="F5E6D3", fill_type="solid")
    money_fmt = '#,##0.00'

    # Title section
    ws.merge_cells("A1:F1")
    ws["A1"] = "QUILL CREATIVE EVENT DESIGN"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="4A3728")

    ws.merge_cells("A2:F2")
    ws["A2"] = est.get("name", "Estimate")
    ws["A2"].font = Font(name="Calibri", size=13, bold=True, color="7A6855")

    # Client info
    row = 4
    info_fields = [
        ("Client", est.get("client_name", "")),
        ("Event", est.get("event_name", "")),
        ("Date", est.get("event_date", "")),
        ("Guests", str(est.get("guest_count", ""))),
        ("Commission", calc["summary"]["commission_label"]),
    ]
    for label, value in info_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        row += 1

    # Line items header
    row += 1
    headers = ["Description", "Category", "Qty", "Vendor Cost", "Markup %", "Client Cost"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    row += 1

    # Line items
    for item in calc["line_items"]:
        ws.cell(row=row, column=1, value=item["description"])
        ws.cell(row=row, column=2, value=item["category_label"])
        ws.cell(row=row, column=3, value=item["quantity"])
        c = ws.cell(row=row, column=4, value=item["vendor_total"])
        c.number_format = money_fmt
        ws.cell(row=row, column=5, value=f"{item['markup_pct']}%")
        c = ws.cell(row=row, column=6, value=item["client_cost"])
        c.number_format = money_fmt
        row += 1

    # Summary section
    row += 1
    summary = calc["summary"]
    ws.cell(row=row, column=1, value="CLIENT SUMMARY").font = Font(bold=True, size=12, color="4A3728")
    row += 1

    summary_lines = [
        ("Subtotal", summary["client_invoice_subtotal"]),
        ("Tax", summary["total_tax"]),
        ("Client Invoice Total", summary["client_invoice_total"]),
    ]
    if summary["management_fee"] > 0:
        summary_lines.insert(1, ("Management Fee", summary["management_fee"]))

    for label, value in summary_lines:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        c = ws.cell(row=row, column=2, value=value)
        c.number_format = money_fmt
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 15

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"QC_Estimate_{est.get('name', 'export').replace(' ', '_')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Serve frontend ──────────────────────────────────────────────────────

if FRONTEND_DIST.exists():
    app.mount("/assets", StaticFiles(directory=FRONTEND_DIST / "assets"), name="static")

    @app.get("/{full_path:path}")
    async def serve_frontend(full_path: str):
        """Serve the React SPA for all non-API routes."""
        return FileResponse(FRONTEND_DIST / "index.html")
