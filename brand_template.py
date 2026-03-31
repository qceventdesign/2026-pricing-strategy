"""
Apply branding to QC_Estimate_Template_2026_v2_FINAL.xlsx.
Purely visual — NO cell values or formulas changed.
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, Protection, numbers
)
from openpyxl.formatting.rule import FormulaRule
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
import os

FILE = "2026-pricing-strategy/docs/templates/QC_Estimate_Template_2026_v2_FINAL.xlsx"
LOGO = "Quill_Creative_Branding-03.png"

# ── COLOR PALETTE ────────────────────────────────────────────────────────
CHARCOAL = "464543"
MOCHA    = "846E60"
CAMEL    = "C19C81"
CREAM    = "ECDFCE"
LINEN    = "FAF6F3"
SILVER   = "A9AEB4"
WHITE    = "FFFFFF"

# ── REUSABLE STYLES ──────────────────────────────────────────────────────
# Fills
fill_charcoal = PatternFill("solid", fgColor=CHARCOAL)
fill_mocha    = PatternFill("solid", fgColor=MOCHA)
fill_cream    = PatternFill("solid", fgColor=CREAM)
fill_linen    = PatternFill("solid", fgColor=LINEN)
fill_camel    = PatternFill("solid", fgColor=CAMEL)

# Fonts
font_logo_title   = Font(name="Garamond", size=14, bold=True, color=WHITE)
font_section_hdr   = Font(name="Garamond", size=12, bold=True, color=WHITE)
font_sub_hdr       = Font(name="Garamond", size=11, bold=True, color=WHITE)
font_col_hdr       = Font(name="Calibri", size=10, bold=True, color=CHARCOAL)
font_body          = Font(name="Calibri", size=10, color=CHARCOAL)
font_body_bold     = Font(name="Calibri", size=10, bold=True, color=CHARCOAL)
font_note          = Font(name="Calibri", size=9, italic=True, color=SILVER)
font_total         = Font(name="Calibri", size=12, bold=True, color=WHITE)
font_subtotal      = Font(name="Calibri", size=11, bold=True, color=CHARCOAL)
font_title_row     = Font(name="Garamond", size=14, bold=True, color=MOCHA)

# Client-ready fonts
font_cr_header     = Font(name="Garamond", size=11, bold=True, color=WHITE)
font_cr_body       = Font(name="Calibri", size=10, color=CHARCOAL)
font_cr_total      = Font(name="Calibri", size=11, bold=True, color=WHITE)

# Borders
thin_silver  = Side(style="thin", color=SILVER)
hair_silver  = Side(style="hair", color=SILVER)
medium_mocha = Side(style="medium", color=MOCHA)
medium_camel = Side(style="medium", color=CAMEL)

border_grid       = Border(left=hair_silver, right=hair_silver, top=hair_silver, bottom=hair_silver)
border_section_top = Border(top=medium_camel)
border_total       = Border(top=medium_mocha, bottom=medium_mocha)
border_bottom_mocha = Border(bottom=medium_mocha)

# Alignment
align_center = Alignment(horizontal="center", vertical="center")
align_left   = Alignment(horizontal="left", vertical="center")
align_right  = Alignment(horizontal="right", vertical="center")
align_wrap   = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ── HELPER FUNCTIONS ─────────────────────────────────────────────────────

def style_row(ws, row, max_col, font=None, fill=None, border=None, alignment=None, height=None):
    """Apply styles to an entire row up to max_col."""
    if height is not None:
        ws.row_dimensions[row].height = height
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        if font: cell.font = font
        if fill: cell.fill = fill
        if border: cell.border = border
        if alignment: cell.alignment = alignment


def style_range(ws, row_start, row_end, col_start, col_end, font=None, fill=None, border=None, alignment=None):
    """Apply styles to a rectangular range."""
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=c)
            if font: cell.font = font
            if fill: cell.fill = fill
            if border: cell.border = border
            if alignment: cell.alignment = alignment


def alt_rows(ws, row_start, row_end, col_start, col_end):
    """Apply alternating row fills (odd=linen, even=cream) + grid borders."""
    for r in range(row_start, row_end + 1):
        f = fill_linen if (r - row_start) % 2 == 0 else fill_cream
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = f
            cell.border = border_grid
            if not cell.font or cell.font.name == "Calibri":
                cell.font = font_body


def set_col_widths(ws, widths):
    """Set column widths from a dict {col_letter: width}."""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def insert_logo(ws, merge_range, logo_path):
    """Insert logo image into row 1 with charcoal background."""
    ws.row_dimensions[1].height = 55
    # Fill the merged area with charcoal
    for cell in ws[1]:
        cell.fill = fill_charcoal

    img = Image(logo_path)
    img.width = 320
    img.height = 48
    img.anchor = "A1"
    ws.add_image(img)


def setup_print(ws):
    """Configure print settings."""
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.oddHeader.center.text = "Quill Creative Event Design — Confidential"
    ws.oddHeader.center.font = "Calibri,Regular"
    ws.oddHeader.center.size = 8
    ws.oddHeader.center.color = SILVER
    ws.oddFooter.center.text = "Page &P of &N"
    ws.oddFooter.center.font = "Calibri,Regular"
    ws.oddFooter.center.size = 8


def add_margin_health_cf(ws, cell_ref):
    """Add conditional formatting for margin health cell."""
    green_fill  = PatternFill("solid", fgColor="D5F5D5")
    yellow_fill = PatternFill("solid", fgColor="FFF8DC")
    orange_fill = PatternFill("solid", fgColor="FFE8CC")
    red_fill    = PatternFill("solid", fgColor="FFD5D5")

    ws.conditional_formatting.add(cell_ref, FormulaRule(
        formula=[f'SEARCH("STRONG",{cell_ref})>0'], fill=green_fill))
    ws.conditional_formatting.add(cell_ref, FormulaRule(
        formula=[f'SEARCH("TARGET",{cell_ref})>0'], fill=yellow_fill))
    ws.conditional_formatting.add(cell_ref, FormulaRule(
        formula=[f'SEARCH("REVIEW",{cell_ref})>0'], fill=orange_fill))
    ws.conditional_formatting.add(cell_ref, FormulaRule(
        formula=[f'SEARCH("BELOW",{cell_ref})>0'], fill=red_fill))
    ws.conditional_formatting.add(cell_ref, FormulaRule(
        formula=[f'SEARCH("THIN",{cell_ref})>0'], fill=orange_fill))
    ws.conditional_formatting.add(cell_ref, FormulaRule(
        formula=[f'SEARCH("LOSING",{cell_ref})>0'], fill=red_fill))


# ── PREPARE LOGO ─────────────────────────────────────────────────────────
print("Preparing logo...")
try:
    pil_img = PILImage.open(LOGO).convert("RGBA")
    data = pil_img.getdata()
    new_data = []
    for item in data:
        # Make very dark pixels transparent (black/near-black background)
        if item[0] < 40 and item[1] < 40 and item[2] < 40:
            new_data.append((item[0], item[1], item[2], 0))
        else:
            new_data.append(item)
    pil_img.putdata(new_data)
    LOGO_PROCESSED = "logo_transparent.png"
    pil_img.save(LOGO_PROCESSED)
    print("  Logo background made transparent")
except Exception as e:
    LOGO_PROCESSED = LOGO
    print(f"  Pillow processing failed ({e}), using original")


# ── LOAD WORKBOOK ────────────────────────────────────────────────────────
print("Loading workbook...")
wb = openpyxl.load_workbook(FILE)
print(f"  Sheets: {wb.sheetnames}")


# ══════════════════════════════════════════════════════════════════════════
# CLIENT SETUP BRANDING
# ══════════════════════════════════════════════════════════════════════════

def brand_client_setup(ws):
    name = ws.title
    print(f"\n  Branding: {name}")
    max_col = 8  # A-H

    # Column widths
    set_col_widths(ws, {"A": 30, "B": 15, "C": 15, "D": 35, "E": 12, "F": 12, "G": 12, "H": 12})

    # Logo row 1
    insert_logo(ws, "A1:G1", LOGO_PROCESSED)

    # Title row 2
    ws.row_dimensions[2].height = 33
    style_row(ws, 2, max_col, font=font_title_row, fill=fill_linen, height=33)

    # Linen background for whole sheet
    for r in range(3, ws.max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if cell.fill.fgColor is None or cell.fill.fgColor.rgb in ("00000000", None):
                cell.fill = fill_linen

    # Section headers (charcoal bg, white text)
    section_rows = [4, 15, 21, 39, 44]
    for r in section_rows:
        style_row(ws, r, max_col, font=font_section_hdr, fill=fill_charcoal, height=27)

    # Column header row 22 (Category, Markup %, Note)
    style_row(ws, 22, max_col, font=font_col_hdr, fill=fill_cream, border=border_bottom_mocha, height=23)

    # Markup table rows 23-33 (alternating)
    alt_rows(ws, 23, 33, 1, 3)
    ws.row_dimensions[34].height = 18
    ws.cell(row=34, column=1).font = font_note

    # Input field rows: cream background
    for r in [5,6,7,8,9,10,11,12,13,16,17,18,19,35,36,37,40,41,42]:
        ws.row_dimensions[r].height = 20
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = font_body
        # Label column bold
        ws.cell(row=r, column=1).font = font_body_bold

    # Fee input rows with cream bg
    for r in [35,36,37,40,41,42]:
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = fill_cream
            ws.cell(row=r, column=c).border = border_grid

    # How-to-use rows (notes style)
    for r in range(45, 51):
        ws.row_dimensions[r].height = 18
        ws.cell(row=r, column=1).font = font_note

    # Data rows height
    for r in range(5, 14):
        ws.row_dimensions[r].height = 20
        for c in range(2, 4):
            ws.cell(row=r, column=c).fill = fill_cream
            ws.cell(row=r, column=c).border = border_grid


def brand_venue(ws):
    name = ws.title
    print(f"\n  Branding: {name}")
    max_col = 11  # A-K

    # Column widths
    set_col_widths(ws, {
        "A": 30, "B": 13, "C": 13, "D": 15, "E": 15, "F": 11, "G": 21,
        "H": 32, "I": 22, "J": 14, "K": 12
    })

    # Logo row 1
    insert_logo(ws, "A1:G1", LOGO_PROCESSED)

    # Title row 2
    ws.row_dimensions[2].height = 33
    style_row(ws, 2, max_col, font=font_title_row, fill=fill_linen, height=33)

    # Linen background base
    for r in range(3, ws.max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).fill = fill_linen

    # ── SECTION HEADERS (charcoal) ──
    for r in [4, 21, 42, 59, 65, 82]:
        style_row(ws, r, max_col, font=font_section_hdr, fill=fill_charcoal, height=27)

    # ── SUBHEADERS (mocha) ──
    for r in [23, 27, 32, 37]:
        style_row(ws, r, 7, font=font_sub_hdr, fill=fill_mocha, height=25)

    # ── FEE OVERRIDES (row 16-19) ──
    style_row(ws, 16, 7, font=font_note, fill=fill_linen, height=20)
    for r in [17, 18, 19]:
        ws.row_dimensions[r].height = 20
        for c in range(1, 8):
            ws.cell(row=r, column=c).fill = fill_cream
            ws.cell(row=r, column=c).border = border_grid
            ws.cell(row=r, column=c).font = font_body

    # ── COLUMN HEADERS (row 22) ──
    style_row(ws, 22, 7, font=font_col_hdr, fill=fill_cream, border=border_bottom_mocha, height=23)

    # ── DATA ROWS with alternating colors ──
    # F&B rows 24-26
    alt_rows(ws, 24, 26, 1, 7)
    for r in [24,25,26]: ws.row_dimensions[r].height = 20

    # Equipment rows 28-31
    alt_rows(ws, 28, 31, 1, 7)
    for r in [28,29,30,31]: ws.row_dimensions[r].height = 20

    # Venue fees 33-35
    alt_rows(ws, 33, 35, 1, 7)
    for r in [33,34,35]: ws.row_dimensions[r].height = 20

    # QC Staffing 38-40
    alt_rows(ws, 38, 40, 1, 7)
    for r in [38,39,40]: ws.row_dimensions[r].height = 20

    # ── SUMMARY SECTION (rows 43-57) ──
    for r in range(43, 58):
        ws.row_dimensions[r].height = 20
        for c in range(1, 8):
            ws.cell(row=r, column=c).fill = fill_linen
            ws.cell(row=r, column=c).font = font_body
            ws.cell(row=r, column=c).border = border_grid
        ws.cell(row=r, column=1).font = font_body_bold

    # SUBTOTAL row 54
    style_row(ws, 54, 7, font=font_subtotal, fill=fill_cream, border=border_total, height=22)

    # TOTAL row 57
    style_row(ws, 57, 7, font=font_total, fill=fill_charcoal, border=border_total, height=25)

    # ── CLIENT PRICING (59-62) ──
    for r in [60, 61, 62]:
        ws.row_dimensions[r].height = 20
        ws.cell(row=r, column=1).font = font_body_bold

    # ── TRAVEL SECTION (65-81) ──
    style_row(ws, 66, 4, font=font_col_hdr, fill=fill_cream, height=22)
    for r in range(67, 82):
        ws.row_dimensions[r].height = 20
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = fill_linen if (r % 2 == 1) else fill_cream
            ws.cell(row=r, column=c).border = border_grid
            ws.cell(row=r, column=c).font = font_body
        ws.cell(row=r, column=1).font = font_body_bold

    # Travel total row 81
    style_row(ws, 81, 4, font=font_subtotal, fill=fill_cream, border=border_total, height=22)

    # ── PROFIT & MARGIN (82-93) ──
    for r in range(83, 94):
        ws.row_dimensions[r].height = 20
        for c in range(1, 6):
            ws.cell(row=r, column=c).fill = fill_linen
            ws.cell(row=r, column=c).font = font_body
            ws.cell(row=r, column=c).border = border_grid
        ws.cell(row=r, column=1).font = font_body_bold

    # Margin health conditional formatting
    add_margin_health_cf(ws, "D88")
    add_margin_health_cf(ws, "D93")

    # ── CLIENT-READY TABLE (I-K) ──
    # Header row (I4 area)
    for c in [9, 10, 11]:
        ws.cell(row=4, column=c).font = font_section_hdr
        ws.cell(row=4, column=c).fill = fill_charcoal

    # Title row 7
    for c in [9, 10, 11]:
        ws.cell(row=7, column=c).fill = fill_mocha
        ws.cell(row=7, column=c).font = font_cr_header

    # Column headers row 11
    for c in [9, 10, 11]:
        ws.cell(row=11, column=c).fill = fill_mocha
        ws.cell(row=11, column=c).font = font_cr_header
        ws.cell(row=11, column=c).border = border_bottom_mocha

    # Data rows 12-22
    for r in range(12, 23):
        for c in [9, 10, 11]:
            ws.cell(row=r, column=c).fill = fill_linen
            ws.cell(row=r, column=c).font = font_cr_body
            ws.cell(row=r, column=c).border = border_grid

    # Total row 23
    for c in [9, 10, 11]:
        ws.cell(row=23, column=c).fill = fill_charcoal
        ws.cell(row=23, column=c).font = font_cr_total
        ws.cell(row=23, column=c).border = border_total

    # Guest count row 24
    for c in [9, 10]:
        ws.cell(row=24, column=c).fill = fill_linen
        ws.cell(row=24, column=c).font = font_body

    # Team notes area (I27+)
    for r in range(27, 36):
        for c in [9, 10, 11]:
            ws.cell(row=r, column=c).fill = fill_linen
            ws.cell(row=r, column=c).font = font_note


def brand_decor(ws):
    name = ws.title
    print(f"\n  Branding: {name}")
    max_col = 14  # A-N

    # Column widths
    set_col_widths(ws, {
        "A": 30, "B": 13, "C": 13, "D": 15, "E": 15, "F": 15, "G": 11,
        "H": 32, "I": 21, "J": 22, "K": 14, "L": 3, "M": 20, "N": 14
    })

    # Logo row 1
    insert_logo(ws, "A1:H1", LOGO_PROCESSED)

    # Title row 2
    ws.row_dimensions[2].height = 33
    style_row(ws, 2, max_col, font=font_title_row, fill=fill_linen, height=33)

    # Linen base
    for r in range(3, min(ws.max_row + 1, 140)):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).fill = fill_linen

    # ── SECTION HEADERS (charcoal) ──
    decor_section_rows = [4, 32, 77, 98, 110, 127]
    for r in decor_section_rows:
        if r <= ws.max_row:
            style_row(ws, r, 9, font=font_section_hdr, fill=fill_charcoal, height=27)

    # ── SUBHEADERS (mocha) ──
    decor_sub_rows = [13, 15, 26, 34, 43, 52, 61, 68, 79, 90]
    for r in decor_sub_rows:
        if r <= ws.max_row:
            style_row(ws, r, 9, font=font_sub_hdr, fill=fill_mocha, height=25)

    # ── COLUMN HEADERS ──
    for r in [14, 33, 78]:
        if r <= ws.max_row:
            style_row(ws, r, 9, font=font_col_hdr, fill=fill_cream, border=border_bottom_mocha, height=23)

    # ── DATA ROWS with alternating ──
    data_sections = [
        (16, 25),   # Floral
        (27, 29),   # Floral fees
        (35, 42),   # Seating
        (44, 51),   # Lounge
        (53, 60),   # Tables
        (62, 67),   # Decor & Accessories
        (69, 74),   # Rental fees
        (80, 89),   # AV equipment
        (91, 95),   # AV fees
    ]
    for start, end in data_sections:
        alt_rows(ws, start, end, 1, 9)
        for r in range(start, end + 1):
            ws.row_dimensions[r].height = 20

    # ── SUBTOTAL ROWS ──
    for r in [30, 75, 96]:
        if r <= ws.max_row:
            style_row(ws, r, 9, font=font_subtotal, fill=fill_cream, border=border_total, height=22)

    # ── SUMMARY (99-104) ──
    for r in range(99, 104):
        if r <= ws.max_row:
            ws.row_dimensions[r].height = 20
            for c in range(1, 10):
                ws.cell(row=r, column=c).fill = fill_linen
                ws.cell(row=r, column=c).font = font_body
                ws.cell(row=r, column=c).border = border_grid
            ws.cell(row=r, column=1).font = font_body_bold

    # TOTAL row 104
    if ws.max_row >= 104:
        style_row(ws, 104, 9, font=font_total, fill=fill_charcoal, border=border_total, height=25)

    # ── PROFIT & MARGIN (128-138) ──
    for r in range(128, 139):
        if r <= ws.max_row:
            ws.row_dimensions[r].height = 20
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = fill_linen
                ws.cell(row=r, column=c).font = font_body
                ws.cell(row=r, column=c).border = border_grid
            ws.cell(row=r, column=1).font = font_body_bold

    # Margin health CF
    if ws.max_row >= 133:
        add_margin_health_cf(ws, "D133")
    if ws.max_row >= 138:
        add_margin_health_cf(ws, "D138")

    # ── TRAVEL SECTION (110-126) ──
    if ws.max_row >= 111:
        style_row(ws, 111, 4, font=font_col_hdr, fill=fill_cream, height=22)
        for r in range(112, 127):
            if r <= ws.max_row:
                ws.row_dimensions[r].height = 20
                for c in range(1, 5):
                    ws.cell(row=r, column=c).fill = fill_linen if (r % 2 == 0) else fill_cream
                    ws.cell(row=r, column=c).border = border_grid
                    ws.cell(row=r, column=c).font = font_body
                ws.cell(row=r, column=1).font = font_body_bold

    # ── DETAILED CLIENT VIEW (J-K) ──
    # Section header row 4 (J4, K4)
    for c in [10, 11]:
        ws.cell(row=4, column=c).fill = fill_charcoal
        ws.cell(row=4, column=c).font = font_section_hdr

    # Title row 7
    for c in [10, 11]:
        ws.cell(row=7, column=c).fill = fill_mocha
        ws.cell(row=7, column=c).font = font_cr_header

    # Column headers row 14
    for c in [10, 11]:
        ws.cell(row=14, column=c).fill = fill_mocha
        ws.cell(row=14, column=c).font = font_cr_header

    # Data rows
    for r in range(15, 90):
        if r <= ws.max_row:
            for c in [10, 11]:
                cell = ws.cell(row=r, column=c)
                cell.fill = fill_linen
                cell.border = border_grid
                cell.font = font_cr_body

    # Subtotal rows in client view
    for r in [29, 68, 86]:
        for c in [10, 11]:
            ws.cell(row=r, column=c).fill = fill_cream
            ws.cell(row=r, column=c).font = font_subtotal
            ws.cell(row=r, column=c).border = border_total

    # Total row 90
    for c in [10, 11]:
        ws.cell(row=90, column=c).fill = fill_charcoal
        ws.cell(row=90, column=c).font = font_cr_total
        ws.cell(row=90, column=c).border = border_total

    # Tax row 88
    for c in [10, 11]:
        ws.cell(row=88, column=c).fill = fill_linen
        ws.cell(row=88, column=c).font = font_body

    # ── SUMMARY VIEW (M-N) ──
    for c in [13, 14]:
        ws.cell(row=4, column=c).fill = fill_charcoal
        ws.cell(row=4, column=c).font = font_section_hdr

    # Title row 7
    for c in [13, 14]:
        ws.cell(row=7, column=c).fill = fill_mocha
        ws.cell(row=7, column=c).font = font_cr_header

    # Column headers 14
    for c in [13, 14]:
        ws.cell(row=14, column=c).fill = fill_mocha
        ws.cell(row=14, column=c).font = font_cr_header

    # Data rows 15-22
    for r in range(15, 23):
        for c in [13, 14]:
            ws.cell(row=r, column=c).fill = fill_linen
            ws.cell(row=r, column=c).border = border_grid
            ws.cell(row=r, column=c).font = font_cr_body

    # Total row 24
    for c in [13, 14]:
        ws.cell(row=24, column=c).fill = fill_charcoal
        ws.cell(row=24, column=c).font = font_cr_total
        ws.cell(row=24, column=c).border = border_total

    # Guest count 25
    for c in [13, 14]:
        ws.cell(row=25, column=c).fill = fill_linen
        ws.cell(row=25, column=c).font = font_body

    # Team notes area (M28+)
    for r in range(28, 41):
        for c in [13, 14]:
            ws.cell(row=r, column=c).fill = fill_linen
            ws.cell(row=r, column=c).font = font_note


# ══════════════════════════════════════════════════════════════════════════
# APPLY BRANDING TO ALL 6 TABS
# ══════════════════════════════════════════════════════════════════════════

print("\nApplying branding...")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    if "Client Setup" in sheet_name:
        brand_client_setup(ws)
    elif "Venue" in sheet_name:
        brand_venue(ws)
    elif "Decor" in sheet_name:
        brand_decor(ws)

    # Print setup for all sheets
    setup_print(ws)


# ══════════════════════════════════════════════════════════════════════════
# VERIFICATION
# ══════════════════════════════════════════════════════════════════════════

print("\n\nVerification...")

# Check formulas on Sample tabs
vs = wb["Venue Estimate - Sample"]
checks = {
    "D43": "=SUM(D24:D26)",
    "D47": "=IFERROR(D28*F28+D29*F29+D30*F30+D31*F31,0)",
    "J23": "=SUM(J15:J21)",
}
for cell, expected in checks.items():
    actual = vs[cell].value
    status = "PASS" if actual == expected else "FAIL"
    print(f"  Venue Sample {cell}: {status}")
    if status == "FAIL":
        print(f"    Expected: {expected}")
        print(f"    Got:      {actual}")

# Check D51 robust formula
assert "SUBSTITUTE" in vs["D51"].value, "D51 formula corrupted!"
print("  Venue Sample D51: PASS (robust formula intact)")

# Check B62
assert "D43" in vs["B62"].value, "B62 formula corrupted!"
print("  Venue Sample B62: PASS (D43 ref intact)")

# Check cross-sheet refs still correct
assert "'Client Setup - Sample'" in vs["B9"].value, "Cross-ref broken!"
print("  Cross-sheet refs: PASS")

# Check blank tab formulas
vb = wb["Venue Estimate - Blank"]
assert vb["D43"].value == "=SUM(D24:D26)", "Blank D43 formula lost!"
assert "'Client Setup - Blank'" in vb["B24"].value, "Blank cross-ref broken!"
print("  Venue Blank formulas: PASS")

# Check protection still active
for name in wb.sheetnames:
    assert wb[name].protection.sheet, f"{name} protection lost!"
print("  Protection: PASS (all 6 tabs)")

print(f"\n  Tab order: {wb.sheetnames}")


# ══════════════════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════════════════

wb.save(FILE)
print(f"\nSaved: {FILE}")

# Cleanup
if os.path.exists("logo_transparent.png"):
    os.remove("logo_transparent.png")

print("Done! Branding applied to all 6 tabs.")
