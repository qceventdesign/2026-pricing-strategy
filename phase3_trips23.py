#!/usr/bin/env python3
"""
Phase 3: Add Trip 2 & Trip 3 formulas on Venue Estimate tab.
Copy Trip 1 formulas from column B to columns C (Trip 2) and D (Trip 3).
Update Total Travel row. Run tests 3-6.
"""

import openpyxl

SRC = '/Users/alex/2026-pricing-strategy/2026-pricing-strategy/docs/templates/QC_Estimate_Template_2026_v2.xlsx'
wb = openpyxl.load_workbook(SRC)
ve = wb['Venue Estimate']

# Row references (from Phase 2)
R_DEST = 68
R_TTYPE = 69
R_FTYPE = 70
R_LASTMIN = 71
R_STAFF = 72
R_NIGHTS = 73
R_HOTELBUD = 74
R_VTYPE = 75
R_VSERVICE = 76
R_VHOURS = 77
R_VCUSTOM = 78
R_TRAVEL = 80
R_HOTEL = 81
R_PERDIEM = 82
R_VEHICLE = 83
R_TRIPTOTAL = 84
R_TOTAL = 86


# ════════════════════════════════════════════════════════════════
# STEP 1: Write Trip 2 (C) and Trip 3 (D) formulas
# ════════════════════════════════════════════════════════════════
print("Step 1: Writing Trip 2 & Trip 3 formulas...")

for col_idx, CL in [(3, 'C'), (4, 'D')]:
    origin = f'{CL}{R_DEST - 1}'  # Origin row = 67
    dest = f'{CL}{R_DEST}'
    ttype = f'{CL}{R_TTYPE}'
    ftype = f'{CL}{R_FTYPE}'
    lastmin = f'{CL}{R_LASTMIN}'
    staff = f'{CL}{R_STAFF}'
    nights = f'{CL}{R_NIGHTS}'
    hotelbud = f'{CL}{R_HOTELBUD}'
    vtype = f'{CL}{R_VTYPE}'
    vservice = f'{CL}{R_VSERVICE}'
    vhours = f'{CL}{R_VHOURS}'
    vcustom = f'{CL}{R_VCUSTOM}'

    # Travel Cost
    ve.cell(row=R_TRAVEL, column=col_idx).value = (
        f'=IF(OR({dest}="",{staff}=""),0,'
        f'IF({ttype}="Drive",'
        f'IFERROR(VLOOKUP({origin}&" to "&{dest},DriveRoutes,2,FALSE),0),'
        f'IF({ttype}="Train",'
        f'IFERROR(VLOOKUP({origin}&" to "&{dest},TrainRoutes,2,FALSE),0)*{staff},'
        f'IF({ttype}="Flight",'
        f'IFERROR(VLOOKUP({ftype},FlightTypes,2,FALSE),0)*{staff}'
        f'+IF({lastmin}="Yes",150*{staff},0),'
        f'0))))'
    )

    # Hotel Cost
    ve.cell(row=R_HOTEL, column=col_idx).value = (
        f'=IF(OR({dest}="",{staff}="",{nights}=""),0,'
        f'{nights}*IFERROR(IF({hotelbud}="High",'
        f'VLOOKUP({dest},HotelRates,3,FALSE),'
        f'VLOOKUP({dest},HotelRates,2,FALSE)),0)*{staff})'
    )

    # Per Diem Cost
    ve.cell(row=R_PERDIEM, column=col_idx).value = (
        f'=IF(OR({dest}="",{staff}=""),0,'
        f'IF({dest}="NYC",'
        f'IF({nights}=0,\'Client Setup\'!C95,'
        f'IF({nights}=1,\'Client Setup\'!C95+\'Client Setup\'!B95,'
        f'2*\'Client Setup\'!C95+({nights}-1)*\'Client Setup\'!B95)),'
        f'IF({nights}=0,\'Client Setup\'!C94,'
        f'IF({nights}=1,\'Client Setup\'!C94+\'Client Setup\'!B94,'
        f'2*\'Client Setup\'!C94+({nights}-1)*\'Client Setup\'!B94)))'
        f'*{staff})'
    )

    # Vehicle Cost
    ve.cell(row=R_VEHICLE, column=col_idx).value = (
        f'=IF(AND(ISNUMBER({vcustom}),{vcustom}>0),{vcustom},'
        f'IF(OR({vtype}="",{vtype}="None"),0,'
        f'IF({vservice}="Airport Transfer",'
        f'IFERROR(VLOOKUP({dest},VehicleRates,'
        f'IF({vtype}="Sedan",3,IF({vtype}="SUV",5,7)),FALSE),0),'
        f'IF({vservice}="Hourly",'
        f'IFERROR(VLOOKUP({dest},VehicleRates,'
        f'IF({vtype}="Sedan",2,IF({vtype}="SUV",4,6)),FALSE),0)*{vhours},'
        f'0))))'
    )

    # Trip Total
    ve.cell(row=R_TRIPTOTAL, column=col_idx).value = (
        f'={CL}{R_TRAVEL}+{CL}{R_HOTEL}+{CL}{R_PERDIEM}+{CL}{R_VEHICLE}'
    )

    print(f"  Trip {'2' if CL == 'C' else '3'} (column {CL}): 5 formulas written")


# ════════════════════════════════════════════════════════════════
# STEP 2: Update Total Travel to sum all 3 trips
# ════════════════════════════════════════════════════════════════
print("Step 2: Updating Total Travel formula...")
ve.cell(row=R_TOTAL, column=4).value = f'=B{R_TRIPTOTAL}+C{R_TRIPTOTAL}+D{R_TRIPTOTAL}'
print(f"  D{R_TOTAL} = =B{R_TRIPTOTAL}+C{R_TRIPTOTAL}+D{R_TRIPTOTAL}")


# ════════════════════════════════════════════════════════════════
# STEP 3: Save
# ════════════════════════════════════════════════════════════════
print("\nSaving...")
wb.save(SRC)
print(f"Saved to {SRC}")


# ════════════════════════════════════════════════════════════════
# STEP 4: Test simulation (Tests 3-6)
# ════════════════════════════════════════════════════════════════
print("\n" + "=" * 60)
print("TEST SIMULATION (Tests 3-6)")
print("=" * 60)

# Rate data
DRIVE = {
    "DC to Richmond VA": 205, "Charlotte to Raleigh NC": 230,
    "Charlotte to Asheville NC": 190, "Charlotte to Charleston SC": 300,
    "Charlotte to Atlanta": 360, "DC to Virginia Secondary": 215,
    "DC to Philadelphia": 190, "DC to NYC": 325, "Charlotte to DC": 270,
}
TRAIN = {"DC to NYC": (180, 300), "Philadelphia to NYC": (120, 220), "DC to Philadelphia": (120, 200)}
FLIGHT = {"Short Haul": (300, 450), "Medium Haul": (400, 600), "Major Market to NYC": (450, 650)}
HOTEL = {
    "NYC": (450, 650), "DC": (350, 550), "Charleston SC": (350, 550),
    "Atlanta": (275, 450), "Charlotte": (250, 400), "Asheville NC": (300, 500),
    "Raleigh NC": (250, 400), "Philadelphia": (350, 550), "Virginia Secondary": (200, 350),
}
PERDIEM = {"Standard": (68, 34), "NYC": (92, 46)}
VEHICLE = {
    "NYC":      {"Sedan": (85, 250), "SUV": (125, 350), "Sprinter": (175, 500)},
    "DC":       {"Sedan": (75, 200), "SUV": (100, 275), "Sprinter": (150, 400)},
    "Philadelphia": {"Sedan": (70, 175), "SUV": (95, 250), "Sprinter": (140, 375)},
    "Charlotte":    {"Sedan": (60, 150), "SUV": (85, 200), "Sprinter": (125, 325)},
    "Raleigh NC":   {"Sedan": (60, 150), "SUV": (85, 200), "Sprinter": (125, 325)},
    "Asheville NC": {"Sedan": (65, 175), "SUV": (90, 225), "Sprinter": (130, 350)},
    "Charleston SC":{"Sedan": (60, 150), "SUV": (85, 200), "Sprinter": (125, 325)},
    "Atlanta":      {"Sedan": (65, 175), "SUV": (95, 250), "Sprinter": (140, 375)},
    "Virginia Secondary": {"Sedan": (55, 125), "SUV": (75, 175), "Sprinter": (110, 300)},
    "Other":        {"Sedan": (65, 175), "SUV": (90, 225), "Sprinter": (135, 350)},
}


def calc_travel(origin, dest, ttype, ftype, lastmin, staff):
    if ttype == "Drive":
        return DRIVE.get(f"{origin} to {dest}", 0)
    elif ttype == "Train":
        return TRAIN.get(f"{origin} to {dest}", (0, 0))[0] * staff
    elif ttype == "Flight":
        return FLIGHT.get(ftype, (0, 0))[0] * staff + (150 * staff if lastmin == "Yes" else 0)
    return 0

def calc_hotel(dest, nights, budget, staff):
    if dest not in HOTEL or nights == 0:
        return 0
    return nights * HOTEL[dest][1 if budget == "High" else 0] * staff

def calc_perdiem(dest, nights, staff):
    full, half = PERDIEM["NYC"] if dest == "NYC" else PERDIEM["Standard"]
    if nights == 0:
        return half * staff
    elif nights == 1:
        return (half + full) * staff
    else:
        return (2 * half + (nights - 1) * full) * staff

def calc_vehicle(dest, vtype, vservice, vhours, vcustom):
    if vcustom and vcustom > 0:
        return vcustom
    if not vtype or vtype == "None":
        return 0
    if dest not in VEHICLE or vtype not in VEHICLE.get(dest, {}):
        return 0
    hourly, airport = VEHICLE[dest][vtype]
    if vservice == "Airport Transfer":
        return airport
    elif vservice == "Hourly":
        return hourly * (vhours or 0)
    return 0


def run_test(name, origin, dest, ttype, ftype, lastmin, staff, nights,
             budget, vtype, vservice, vhours, vcustom,
             exp_travel, exp_hotel, exp_perdiem, exp_vehicle, exp_total):
    travel = calc_travel(origin, dest, ttype, ftype, lastmin, staff)
    hotel = calc_hotel(dest, nights, budget, staff)
    perdiem = calc_perdiem(dest, nights, staff)
    vehicle = calc_vehicle(dest, vtype, vservice, vhours, vcustom)
    total = travel + hotel + perdiem + vehicle

    results = [
        ("Travel",   travel,  exp_travel),
        ("Hotel",    hotel,   exp_hotel),
        ("Per Diem", perdiem, exp_perdiem),
        ("Vehicle",  vehicle, exp_vehicle),
        ("Total",    total,   exp_total),
    ]

    all_pass = True
    print(f"\n{name}:")
    for label, actual, expected in results:
        ok = actual == expected
        if not ok:
            all_pass = False
        print(f"  {'PASS' if ok else 'FAIL'}: {label} = ${actual:,} (expected ${expected:,})")
    return all_pass


t3 = run_test(
    "Test 3 — Flight Charlotte to Atlanta, 2 nights, last minute",
    origin="Charlotte", dest="Atlanta", ttype="Flight", ftype="Medium Haul",
    lastmin="Yes", staff=2, nights=2, budget="Low",
    vtype="None", vservice="", vhours=0, vcustom=None,
    exp_travel=1100, exp_hotel=1100, exp_perdiem=272, exp_vehicle=0, exp_total=2472
)

t4 = run_test(
    "Test 4 — Drive DC to Virginia Secondary, 1 night, custom vehicle",
    origin="DC", dest="Virginia Secondary", ttype="Drive", ftype="",
    lastmin="No", staff=1, nights=1, budget="Low",
    vtype="Sedan", vservice="Hourly", vhours=1, vcustom=2000,
    exp_travel=215, exp_hotel=200, exp_perdiem=102, exp_vehicle=2000, exp_total=2517
)

t5 = run_test(
    "Test 5 — Train Philadelphia to NYC, 2 nights",
    origin="Philadelphia", dest="NYC", ttype="Train", ftype="",
    lastmin="No", staff=1, nights=2, budget="Low",
    vtype="None", vservice="", vhours=0, vcustom=None,
    exp_travel=120, exp_hotel=900, exp_perdiem=184, exp_vehicle=0, exp_total=1204
)

t6 = run_test(
    "Test 6 — Drive Charlotte to Atlanta, 3 nights",
    origin="Charlotte", dest="Atlanta", ttype="Drive", ftype="",
    lastmin="No", staff=2, nights=3, budget="High",
    vtype="Sprinter", vservice="Hourly", vhours=8, vcustom=None,
    exp_travel=360, exp_hotel=2700, exp_perdiem=408, exp_vehicle=1120, exp_total=4588
)

# Formula structure check
print("\n" + "-" * 40)
print("Formula verification:")
wb2 = openpyxl.load_workbook(SRC)
ve2 = wb2['Venue Estimate']
for col_idx, trip in [(2, "Trip 1"), (3, "Trip 2"), (4, "Trip 3")]:
    f = ve2.cell(row=R_TRAVEL, column=col_idx).value
    has_formula = str(f).startswith("=")
    print(f"  {trip} Travel (row {R_TRAVEL}): {'PASS' if has_formula else 'FAIL'} — formula present")

total_f = ve2.cell(row=R_TOTAL, column=4).value
print(f"  Total Travel (D{R_TOTAL}): {total_f}")

all_pass = t3 and t4 and t5 and t6
print(f"\n{'=' * 60}")
print(f"OVERALL: {'ALL TESTS PASSED' if all_pass else 'FAILURES — review above'}")
print(f"{'=' * 60}")
