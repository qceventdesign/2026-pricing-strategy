#!/usr/bin/env python3
"""
Phase 1: Replace travel rate tables on Client Setup tab with comprehensive new system.

Tables added:
1. Transportation Rates (Drive/Train/Flight) — rows 54-78
2. Hotel Rates (9 markets, Low/High) — rows 80-90
3. Per Diem Rates (Standard/NYC + reference) — rows 92-101
4. Site Visit Vehicle Rates (10 markets × 3 vehicles × 2 services) — rows 103-114
5. Quick Budget Reference — rows 116-120
6. Dropdown helper lists in column I

NOTE: Estimate tab formulas referencing old named ranges will temporarily
show errors until Phase 2 rewrites them.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.workbook.defined_name import DefinedName

SRC = '/Users/alex/2026-pricing-strategy/2026-pricing-strategy/docs/templates/QC_Estimate_Template_2026_v2.xlsx'
wb = openpyxl.load_workbook(SRC)
cs = wb['Client Setup']

# ── Styles (matching existing spreadsheet) ──
FILL_INPUT   = PatternFill('solid', fgColor='FFFFF8F0')
FILL_CALC    = PatternFill('solid', fgColor='FFF5F0EB')
FILL_ROW_BG  = PatternFill('solid', fgColor='FFFAF6F3')
FILL_HEADER  = PatternFill('solid', fgColor='FFECDFCE')
FILL_SUBHEAD = PatternFill('solid', fgColor='FFE8D9C8')
FILL_NONE    = PatternFill()

FONT_SECTION = Font(name='Playfair Display', size=11, bold=True, color='FF846E60')
FONT_COLHEAD = Font(name='Playfair Display', size=10, bold=True, color='FFC19C81')
FONT_LABEL   = Font(name='Playfair Display', size=10, color='FF464543')
FONT_INPUT   = Font(name='Playfair Display', size=10, bold=True, color='FF846E60')
FONT_CALC    = Font(name='Playfair Display', size=10, color='FF846E60')
FONT_BOLD    = Font(name='Playfair Display', size=10, bold=True, color='FF464543')
FONT_NOTE    = Font(name='Playfair Display', size=9, italic=True, color='FF846E60')

AC = Alignment(horizontal='center', vertical='center')
AL = Alignment(horizontal='left', vertical='center')
AW = Alignment(horizontal='left', vertical='center', wrap_text=True)


# ── Helper functions ──
def section_hdr(row, text, cols=7):
    cs.cell(row=row, column=1).value = text
    for c in range(1, cols + 1):
        cs.cell(row=row, column=c).fill = FILL_HEADER
        cs.cell(row=row, column=c).font = FONT_SECTION


def sub_hdr(row, text, cols=3):
    cs.cell(row=row, column=1).value = text
    for c in range(1, cols + 1):
        cs.cell(row=row, column=c).fill = FILL_SUBHEAD
        cs.cell(row=row, column=c).font = FONT_COLHEAD
        cs.cell(row=row, column=c).alignment = AL


def col_hdrs(row, headers):
    for i, h in enumerate(headers):
        cell = cs.cell(row=row, column=1 + i)
        cell.value = h
        cell.fill = FILL_SUBHEAD
        cell.font = FONT_COLHEAD
        cell.alignment = AC


def write_row(row, values, dollar_cols=None, note_cols=None):
    dollar_cols = dollar_cols or []
    note_cols = note_cols or []
    for i, val in enumerate(values):
        cell = cs.cell(row=row, column=1 + i)
        cell.value = val
        if i == 0:
            cell.font = FONT_LABEL
            cell.fill = FILL_ROW_BG
            cell.alignment = AL
        elif i in note_cols:
            cell.font = FONT_NOTE
            cell.fill = FILL_ROW_BG
            cell.alignment = AW
        else:
            cell.font = FONT_INPUT
            cell.fill = FILL_INPUT
            cell.alignment = AC
            if i in dollar_cols:
                cell.number_format = '$#,##0'


# ════════════════════════════════════════════════════════════════
# STEP 1: Clear old travel/vehicle tables (rows 52-72)
# ════════════════════════════════════════════════════════════════
print("Step 1: Clearing old tables (rows 52-72)...")
for row in range(52, 73):
    for col in range(1, 10):
        cell = cs.cell(row=row, column=col)
        cell.value = None
        cell.fill = FILL_NONE
        cell.font = Font()
        cell.alignment = Alignment()
        cell.number_format = 'General'


# ════════════════════════════════════════════════════════════════
# STEP 2: Insert 49 rows at row 73 to create space
# Tax table shifts: row 74 → 123, rows 75-96 → 124-145
# ════════════════════════════════════════════════════════════════
print("Step 2: Inserting 49 rows at row 73...")
cs.insert_rows(73, 49)


# ════════════════════════════════════════════════════════════════
# STEP 3: Remove old named ranges, fix LocationList & tax formulas
# ════════════════════════════════════════════════════════════════
print("Step 3: Updating named ranges and tax formulas...")

for name in ['TravelRates', 'DestinationList', 'VehicleRates', 'VehicleList', 'LocationList']:
    if name in wb.defined_names:
        del wb.defined_names[name]

# LocationList now at rows 124-145
wb.defined_names.add(DefinedName('LocationList', attr_text="'Client Setup'!$A$124:$A$145"))

# Fix tax VLOOKUPs explicitly (insurance against openpyxl auto-update issues)
cs.cell(row=17, column=3).value = '=VLOOKUP(B16,$A$124:$D$145,2,FALSE)'
cs.cell(row=18, column=3).value = '=VLOOKUP(B16,$A$124:$D$145,3,FALSE)'
cs.cell(row=19, column=3).value = '=VLOOKUP(B16,$A$124:$D$145,4,FALSE)'


# ════════════════════════════════════════════════════════════════
# STEP 4: Write all new rate tables (rows 52-121)
# ════════════════════════════════════════════════════════════════
print("Step 4: Writing new rate tables...")

# ── Row 52: Main header ──
section_hdr(52, "QC TRAVEL RATE TABLES")

# ──────────────────────────────────────────────────────
# TABLE 1: TRANSPORTATION RATES (rows 54-78)
# ──────────────────────────────────────────────────────
section_hdr(54, "TRANSPORTATION RATES", 3)

# Drive Costs
sub_hdr(55, "Drive Costs (round trip per trip, based on $0.725/mile + parking/tolls)", 2)
col_hdrs(56, ["Route", "Cost"])

DRIVES = [
    ("DC to Richmond VA", 205),
    ("Charlotte to Raleigh", 230),
    ("Charlotte to Asheville", 190),
    ("Charlotte to Charleston", 300),
    ("Charlotte to Atlanta", 360),
    ("DC to Virginia Secondary Markets", 215),
    ("DC to Philadelphia", 190),
    ("DC to NYC", 325),
    ("Charlotte to DC", 270),
]
for i, (route, cost) in enumerate(DRIVES):
    write_row(57 + i, [route, cost], dollar_cols=[1])

# Train Costs
sub_hdr(67, "Train Costs (round trip, per person)", 3)
col_hdrs(68, ["Route", "Low", "High"])

TRAINS = [
    ("DC to NYC", 180, 300),
    ("Philadelphia to NYC", 120, 220),
    ("DC to Philadelphia", 120, 200),
]
for i, (route, lo, hi) in enumerate(TRAINS):
    write_row(69 + i, [route, lo, hi], dollar_cols=[1, 2])

# Flight Costs
sub_hdr(73, "Flight Costs (round trip, per person)", 3)
col_hdrs(74, ["Route Type", "Low", "High"])

FLIGHTS = [
    ("Short Haul (under 1.5 hrs)", 300, 450),
    ("Medium Haul (1.5\u20132.5 hrs)", 400, 600),
    ("Major Market to NYC", 450, 650),
]
for i, (ft, lo, hi) in enumerate(FLIGHTS):
    write_row(75 + i, [ft, lo, hi], dollar_cols=[1, 2])

# Last minute buffer note
cs.cell(row=78, column=1).value = "Last Minute Booking Buffer: +$150 per person"
cs.cell(row=78, column=1).font = FONT_NOTE
cs.cell(row=78, column=1).fill = FILL_ROW_BG
cs.cell(row=78, column=1).alignment = AL

# ──────────────────────────────────────────────────────
# TABLE 2: HOTEL RATES (rows 80-90)
# ──────────────────────────────────────────────────────
section_hdr(80, "HOTEL RATES (Per Night, Per Person)", 3)
col_hdrs(81, ["Market", "Low", "High"])

HOTELS = [
    ("NYC", 450, 650),
    ("DC", 350, 550),
    ("Charleston SC", 350, 550),
    ("Atlanta", 275, 450),
    ("Charlotte", 250, 400),
    ("Asheville NC", 300, 500),
    ("Raleigh NC", 250, 400),
    ("Philadelphia", 350, 550),
    ("Virginia Secondary Markets", 200, 350),
]
for i, (market, lo, hi) in enumerate(HOTELS):
    write_row(82 + i, [market, lo, hi], dollar_cols=[1, 2])

# ──────────────────────────────────────────────────────
# TABLE 3: PER DIEM RATES (rows 92-101)
# ──────────────────────────────────────────────────────
section_hdr(92, "PER DIEM RATES", 4)
col_hdrs(93, ["Market", "Full Day", "Half Day", "Notes"])

write_row(94, ["Standard (all except NYC)", 68, 34,
               "Full = breakfast, lunch, dinner. Half = dinner only"],
          dollar_cols=[1, 2], note_cols=[3])
write_row(95, ["NYC", 92, 46,
               "Full = breakfast, lunch, dinner. Half = dinner only"],
          dollar_cols=[1, 2], note_cols=[3])

# Per Diem Reference (auto-calculated)
sub_hdr(97, "Per Diem by Trip Length (auto-calculated reference)", 4)
col_hdrs(98, ["Trip Length", "Calculation", "Standard Total", "NYC Total"])

PD_REF = [
    (99,  "1 Night",  "Half + Full",          "=C94+B94",     "=C95+B95"),
    (100, "2 Nights", "Half + Full + Half",    "=2*C94+B94",   "=2*C95+B95"),
    (101, "3 Nights", "Half + 2 Full + Half",  "=2*C94+2*B94", "=2*C95+2*B95"),
]
for row, length, calc, std_f, nyc_f in PD_REF:
    cs.cell(row=row, column=1).value = length
    cs.cell(row=row, column=1).font = FONT_LABEL
    cs.cell(row=row, column=1).fill = FILL_ROW_BG
    cs.cell(row=row, column=1).alignment = AL

    cs.cell(row=row, column=2).value = calc
    cs.cell(row=row, column=2).font = FONT_NOTE
    cs.cell(row=row, column=2).fill = FILL_CALC
    cs.cell(row=row, column=2).alignment = AC

    for col, formula in [(3, std_f), (4, nyc_f)]:
        cs.cell(row=row, column=col).value = formula
        cs.cell(row=row, column=col).font = FONT_CALC
        cs.cell(row=row, column=col).fill = FILL_CALC
        cs.cell(row=row, column=col).alignment = AC
        cs.cell(row=row, column=col).number_format = '$#,##0'

# ──────────────────────────────────────────────────────
# TABLE 4: SITE VISIT VEHICLE RATES (rows 103-114)
# ──────────────────────────────────────────────────────
section_hdr(103, "SITE VISIT VEHICLE RATES")
col_hdrs(104, ["Market", "Sedan Hourly", "Sedan Airport Transfer",
               "SUV Hourly", "SUV Airport Transfer",
               "Sprinter Hourly", "Sprinter Airport Transfer"])

VEHICLES = [
    ("NYC",                85, 250, 125, 350, 175, 500),
    ("DC",                 75, 200, 100, 275, 150, 400),
    ("Philadelphia",       70, 175,  95, 250, 140, 375),
    ("Charlotte",          60, 150,  85, 200, 125, 325),
    ("Raleigh NC",         60, 150,  85, 200, 125, 325),
    ("Asheville NC",       65, 175,  90, 225, 130, 350),
    ("Charleston SC",      60, 150,  85, 200, 125, 325),
    ("Atlanta",            65, 175,  95, 250, 140, 375),
    ("Virginia Secondary", 55, 125,  75, 175, 110, 300),
    ("Other",              65, 175,  90, 225, 135, 350),
]
for i, (market, *rates) in enumerate(VEHICLES):
    write_row(105 + i, [market] + list(rates),
              dollar_cols=[1, 2, 3, 4, 5, 6])

# ──────────────────────────────────────────────────────
# QUICK BUDGET REFERENCE (rows 116-120)
# ──────────────────────────────────────────────────────
section_hdr(116, "QUICK BUDGET REFERENCE", 2)
col_hdrs(117, ["Travel Type", "Estimated Total Per Person (2-night event)"])

BUDGET = [
    ("Local Drive Market", "$400\u2013$750"),
    ("Regional Flight Market", "$1,200\u2013$1,800"),
    ("NYC Full Travel", "$2,000\u2013$2,800"),
]
for i, (ttype, est) in enumerate(BUDGET):
    row = 118 + i
    cs.cell(row=row, column=1).value = ttype
    cs.cell(row=row, column=1).font = FONT_LABEL
    cs.cell(row=row, column=1).fill = FILL_ROW_BG
    cs.cell(row=row, column=1).alignment = AL
    cs.cell(row=row, column=2).value = est
    cs.cell(row=row, column=2).font = FONT_INPUT
    cs.cell(row=row, column=2).fill = FILL_CALC
    cs.cell(row=row, column=2).alignment = AC


# ════════════════════════════════════════════════════════════════
# STEP 5: Dropdown helper lists (column I)
# ════════════════════════════════════════════════════════════════
print("Step 5: Writing dropdown helper lists in column I...")

cs.cell(row=52, column=9).value = "Origins"
cs.cell(row=52, column=9).font = FONT_COLHEAD
cs.cell(row=52, column=9).fill = FILL_SUBHEAD

ORIGINS = ["Charlotte", "DC", "Philadelphia", "NYC", "Atlanta", "Other"]
for i, o in enumerate(ORIGINS):
    cell = cs.cell(row=53 + i, column=9)
    cell.value = o
    cell.font = FONT_LABEL
    cell.fill = FILL_ROW_BG

cs.cell(row=60, column=9).value = "Destinations"
cs.cell(row=60, column=9).font = FONT_COLHEAD
cs.cell(row=60, column=9).fill = FILL_SUBHEAD

DESTS = ["NYC", "DC", "Philadelphia", "Charlotte", "Atlanta",
         "Richmond VA", "Charleston SC", "Raleigh NC", "Asheville NC",
         "Virginia Secondary", "Other"]
for i, d in enumerate(DESTS):
    cell = cs.cell(row=61 + i, column=9)
    cell.value = d
    cell.font = FONT_LABEL
    cell.fill = FILL_ROW_BG


# ════════════════════════════════════════════════════════════════
# STEP 6: Create named ranges for all new tables
# ════════════════════════════════════════════════════════════════
print("Step 6: Creating named ranges...")

RANGES = {
    'DriveRoutes':     "'Client Setup'!$A$57:$B$65",
    'TrainRoutes':     "'Client Setup'!$A$69:$C$71",
    'FlightTypes':     "'Client Setup'!$A$75:$C$77",
    'HotelRates':      "'Client Setup'!$A$82:$C$90",
    'PerDiemRates':    "'Client Setup'!$A$94:$C$95",
    'VehicleRates':    "'Client Setup'!$A$105:$G$114",
    'OriginList':      "'Client Setup'!$I$53:$I$58",
    'DestinationList': "'Client Setup'!$I$61:$I$71",
}
for name, ref in RANGES.items():
    wb.defined_names.add(DefinedName(name, attr_text=ref))


# ════════════════════════════════════════════════════════════════
# STEP 7: Save
# ════════════════════════════════════════════════════════════════
print("\nSaving...")
wb.save(SRC)
print(f"Saved to {SRC}")


# ════════════════════════════════════════════════════════════════
# STEP 8: Verification
# ════════════════════════════════════════════════════════════════
print("\n" + "=" * 60)
print("VERIFICATION")
print("=" * 60)

wb2 = openpyxl.load_workbook(SRC)
cs2 = wb2['Client Setup']

# Named ranges
print("\nNamed ranges:")
for dn in sorted(wb2.defined_names.values(), key=lambda x: x.name):
    print(f"  {dn.name} = {dn.attr_text}")

# Content spot-checks
checks = [
    (52, 1, "QC TRAVEL RATE TABLES", "Main header"),
    (54, 1, "TRANSPORTATION RATES", "Transport header"),
    (57, 1, "DC to Richmond VA", "First drive route"),
    (57, 2, 205, "DC-Richmond cost"),
    (65, 1, "Charlotte to DC", "Last drive route"),
    (65, 2, 270, "Charlotte-DC cost"),
    (69, 1, "DC to NYC", "First train route"),
    (69, 2, 180, "DC-NYC train low"),
    (69, 3, 300, "DC-NYC train high"),
    (75, 1, "Short Haul (under 1.5 hrs)", "First flight type"),
    (75, 2, 300, "Short haul low"),
    (75, 3, 450, "Short haul high"),
    (82, 1, "NYC", "First hotel market"),
    (82, 2, 450, "NYC hotel low"),
    (82, 3, 650, "NYC hotel high"),
    (90, 1, "Virginia Secondary Markets", "Last hotel market"),
    (90, 2, 200, "VA Secondary hotel low"),
    (90, 3, 350, "VA Secondary hotel high"),
    (94, 1, "Standard (all except NYC)", "Standard per diem"),
    (94, 2, 68, "Standard full day"),
    (94, 3, 34, "Standard half day"),
    (95, 1, "NYC", "NYC per diem"),
    (95, 2, 92, "NYC full day"),
    (95, 3, 46, "NYC half day"),
    (105, 1, "NYC", "First vehicle market"),
    (105, 2, 85, "NYC sedan hourly"),
    (105, 7, 500, "NYC sprinter airport"),
    (113, 1, "Virginia Secondary", "VA Secondary vehicles"),
    (113, 2, 55, "VA Secondary sedan hourly"),
    (114, 1, "Other", "Other vehicles"),
    (116, 1, "QUICK BUDGET REFERENCE", "Budget header"),
]

print("\nContent checks:")
all_pass = True
for row, col, expected, desc in checks:
    actual = cs2.cell(row=row, column=col).value
    ok = actual == expected
    if not ok:
        all_pass = False
        print(f"  FAIL: {desc} — Row {row}, Col {col}: expected '{expected}', got '{actual}'")
    else:
        print(f"  PASS: {desc}")

# Tax table integrity
print("\nTax table integrity:")
tax_hdr = cs2.cell(row=123, column=1).value
tax_first = cs2.cell(row=124, column=1).value
tax_last = cs2.cell(row=145, column=1).value
print(f"  Header row 123: {tax_hdr}")
print(f"  First location row 124: {tax_first}")
print(f"  Last location row 145: {tax_last}")

# VLOOKUP formulas
print("\nTax VLOOKUP formulas:")
for r in [17, 18, 19]:
    f = cs2.cell(row=r, column=3).value
    ok = '$A$124:$D$145' in str(f)
    print(f"  C{r}: {f} — {'PASS' if ok else 'FAIL'}")
    if not ok:
        all_pass = False

# Per diem reference formulas
print("\nPer Diem reference formulas:")
for r in [99, 100, 101]:
    std = cs2.cell(row=r, column=3).value
    nyc = cs2.cell(row=r, column=4).value
    print(f"  Row {r}: Standard={std}, NYC={nyc}")

# Origin/Destination lists
origins = [cs2.cell(row=r, column=9).value for r in range(53, 59)]
dests = [cs2.cell(row=r, column=9).value for r in range(61, 72)]
print(f"\nOrigin list (I53:I58): {origins}")
print(f"Destination list (I61:I71): {dests}")

# Table placement summary
print("\n" + "=" * 60)
print("TABLE PLACEMENT SUMMARY")
print("=" * 60)
print("  Row 52:       QC TRAVEL RATE TABLES (main header)")
print("  Rows 54-65:   Transportation — Drive Costs (9 routes)")
print("  Rows 67-71:   Transportation — Train Costs (3 routes)")
print("  Rows 73-78:   Transportation — Flight Costs (3 types + buffer)")
print("  Rows 80-90:   Hotel Rates (9 markets, Low/High)")
print("  Rows 92-95:   Per Diem Rates (Standard + NYC)")
print("  Rows 97-101:  Per Diem Reference (auto-calculated)")
print("  Rows 103-114: Vehicle Rates (10 markets × 6 rate columns)")
print("  Rows 116-120: Quick Budget Reference")
print("  Column I:     Origin list (I53:I58), Destination list (I61:I71)")
print("  Row 123+:     Location Tax Table (shifted from row 74)")

print(f"\n{'=' * 60}")
print(f"OVERALL: {'ALL CHECKS PASSED' if all_pass else 'SOME CHECKS FAILED — review above'}")
print(f"{'=' * 60}")
