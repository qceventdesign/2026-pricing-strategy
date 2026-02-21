#!/usr/bin/env python3
"""
Fix pass on QC_Estimate_Template_2026_v2.xlsx
1. Event Time → text, propagate to estimate tabs
2. CC Processing Fee dropdown
3. Client Hotel → remove data validation
4. Gratuity default → 20%, Admin Fee default → 5%
5. Venue B8 → "Yes"
6. Decor Column E: fix all broken formulas (off by 3 from row insert)
7. Decor summary, subtotals, profit section: fix all refs
8. SAMPLE Decor: same fixes
9. Full audit
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

DST = '/Users/alex/2026-pricing-strategy/2026-pricing-strategy/docs/templates/QC_Estimate_Template_2026_v2.xlsx'
wb = openpyxl.load_workbook(DST)

cs = wb['Client Setup']
ve = wb['Venue Estimate']

FILL_INPUT = PatternFill('solid', fgColor='FFFFF8F0')
FILL_CALC  = PatternFill('solid', fgColor='FFF5F0EB')
FILL_ROW_BG= PatternFill('solid', fgColor='FFFAF6F3')
FONT_INPUT = Font(name='Playfair Display', size=10, bold=True, color='FF846E60')
FONT_LABEL = Font(name='Playfair Display', size=10, color='FF464543')
AC = Alignment(horizontal='center', vertical='center')
AL = Alignment(horizontal='left', vertical='center')

changes = []

def fix(sheet, cell_ref, old_desc, new_val):
    """Log and apply a fix."""
    changes.append(f"  {sheet}!{cell_ref}: {old_desc} → {new_val[:60] if isinstance(new_val, str) else new_val}")


# ════════════════════════════════════════════════════════════
# FIX 1: Event Time → text format, propagate
# ════════════════════════════════════════════════════════════
print("Fix 1: Event Time...")
cell = cs.cell(row=10, column=2)
cell.number_format = '@'  # text format
cell.fill = FILL_INPUT
cell.font = FONT_INPUT
cell.alignment = AC
fix("Client Setup", "B10", "h:mm AM/PM format", "Text format (@)")

# Propagate Event Time to Venue Estimate (add after Guest Count row 10)
# Insert at row 11 area — but we already have Company at 12, Program at 13, Hotel at 14
# Let's add Event Time at row 15 (after Hotel)
# Actually, let's just add it in the existing info block. Row 11 = Location.
# Better: put it after Location (row 11), before Company (row 12)
# But inserting would shift everything again. Instead, let's use an empty area.
# Current Venue layout: rows 5-14 are info fields, row 15 is blank, row 16 is FEE OVERRIDES
# Let's put Event Time in the blank row 15.
ve.cell(row=15, column=1).value = "Event Time:"
ve.cell(row=15, column=1).fill = FILL_ROW_BG
ve.cell(row=15, column=1).font = FONT_LABEL
ve.cell(row=15, column=1).alignment = AL
ve.cell(row=15, column=2).value = "='Client Setup'!B10"
ve.cell(row=15, column=2).fill = FILL_CALC
ve.cell(row=15, column=2).font = FONT_LABEL
ve.cell(row=15, column=2).alignment = AC
ve.cell(row=15, column=2).number_format = '@'
for c in range(3, 8):
    ve.cell(row=15, column=c).fill = FILL_ROW_BG
fix("Venue Estimate", "A15:B15", "blank row", "Event Time: ='Client Setup'!B10")

# Propagate to Decor and SAMPLE Decor
# They have: row 9=Venue, row 10=Company, 11=Program, 12=Hotel, row 13=FLORALS header
# No blank row available before FLORALS. Use the space just before FLORALS — but that would
# require an insert. Instead, combine with the existing info.
# Actually let's put it right after Hotel (row 12), shifting FLORALS header.
# But that would break all our Decor fixes. Better approach: add Event Time to row 9
# alongside Venue, since Venue is a per-tab field anyway. Or just put it after row 12.
# Let's keep it simple and not insert — add Event Time display to the Client-Ready
# section header area or just note it. Actually the simplest: put it on row 9 col E-F area.
for sheet_name in ['Decor Estimate', 'SAMPLE Decor Estimate']:
    ws = wb[sheet_name]
    # Use column E-F on row 9 (currently empty in that area)
    ws.cell(row=9, column=5).value = "Event Time:"
    ws.cell(row=9, column=5).font = FONT_LABEL
    ws.cell(row=9, column=5).alignment = AL
    ws.cell(row=9, column=6).value = "='Client Setup'!B10"
    ws.cell(row=9, column=6).fill = FILL_CALC
    ws.cell(row=9, column=6).font = FONT_LABEL
    ws.cell(row=9, column=6).alignment = AC
    ws.cell(row=9, column=6).number_format = '@'
    fix(sheet_name, "E9:F9", "empty", "Event Time: ='Client Setup'!B10")


# ════════════════════════════════════════════════════════════
# FIX 2: CC Processing Fee dropdown
# ════════════════════════════════════════════════════════════
print("Fix 2: CC Processing Fee dropdown...")
dv_cc = DataValidation(type="list", formula1='"0,0.035,0.04,0.045,0.05"', allow_blank=False)
dv_cc.sqref = "C35"
cs.add_data_validation(dv_cc)
fix("Client Setup", "C35", "no dropdown", "Dropdown: 0%, 3.5%, 4%, 4.5%, 5%")


# ════════════════════════════════════════════════════════════
# FIX 3: Client Hotel — remove LocationList data validation from B13
# ════════════════════════════════════════════════════════════
print("Fix 3: Client Hotel — remove LocationList DV from B13...")
new_dvs = []
for dv in cs.data_validations.dataValidation:
    sqref = str(dv.sqref)
    # Remove the LocationList validation that's on B13 (should only be on B16)
    if 'B13' in sqref and dv.formula1 and 'LocationList' in str(dv.formula1):
        fix("Client Setup", "B13", "LocationList dropdown", "Plain text (no validation)")
        continue
    new_dvs.append(dv)
cs.data_validations.dataValidation = new_dvs


# ════════════════════════════════════════════════════════════
# FIX 4: Gratuity default → 20%, Admin Fee default → 5%
# ════════════════════════════════════════════════════════════
print("Fix 4: Gratuity/Admin defaults...")
cs.cell(row=41, column=3).value = "20%"
fix("Client Setup", "C41", "None", "20%")

cs.cell(row=42, column=3).value = "5%"
fix("Client Setup", "C42", "None", "5%")


# ════════════════════════════════════════════════════════════
# FIX 5: Venue Rental Taxable default → "Yes"
# ════════════════════════════════════════════════════════════
print("Fix 5: Venue B8 default...")
ve.cell(row=8, column=2).value = "Yes"
fix("Venue Estimate", "B8", "No", "Yes")


# ════════════════════════════════════════════════════════════
# FIXES 6-9: Decor Column E audit + full formula fix
# ════════════════════════════════════════════════════════════
print("Fix 6-9: Decor full formula audit...")

def fix_decor_sheet(ws, label):
    """Fix all broken formulas in a Decor-type sheet caused by 3-row insert."""
    count = 0

    # ── Column E: Our Cost ──
    # Pattern: qty×price rows use =C{row}*D{row}, flat fee rows use =D{row}

    # Floral product (rows 16-25)
    for r in range(16, 26):
        expected = f'=C{r}*D{r}'
        if str(ws.cell(row=r, column=5).value) != expected:
            ws.cell(row=r, column=5).value = expected
            fix(label, f"E{r}", f"was C{r-3}*D{r-3}", expected)
            count += 1

    # Floral fees (rows 27-29) — flat fee, no qty
    for r in range(27, 30):
        expected = f'=D{r}'
        if str(ws.cell(row=r, column=5).value) != expected:
            ws.cell(row=r, column=5).value = expected
            fix(label, f"E{r}", f"was D{r-3}", expected)
            count += 1

    # Floral subtotal (row 30)
    expected = '=SUM(E16:E25)+SUM(E27:E29)'
    ws.cell(row=30, column=5).value = expected
    ws.cell(row=30, column=6).value = '=SUM(F16:F25)+SUM(F27:F29)'
    fix(label, "E30:F30", "old range refs", expected)
    count += 1

    # Seating (rows 35-42)
    for r in range(35, 43):
        expected = f'=C{r}*D{r}'
        if str(ws.cell(row=r, column=5).value) != expected:
            ws.cell(row=r, column=5).value = expected
            fix(label, f"E{r}", f"was C{r-3}*D{r-3}", expected)
            count += 1

    # Lounge (rows 44-51)
    for r in range(44, 52):
        expected = f'=C{r}*D{r}'
        if str(ws.cell(row=r, column=5).value) != expected:
            ws.cell(row=r, column=5).value = expected
            fix(label, f"E{r}", f"was C{r-3}*D{r-3}", expected)
            count += 1

    # Tables (rows 53-60)
    for r in range(53, 61):
        expected = f'=C{r}*D{r}'
        if str(ws.cell(row=r, column=5).value) != expected:
            ws.cell(row=r, column=5).value = expected
            fix(label, f"E{r}", f"was C{r-3}*D{r-3}", expected)
            count += 1

    # Rugs/Accessories (rows 62-67)
    for r in range(62, 68):
        expected = f'=C{r}*D{r}'
        if str(ws.cell(row=r, column=5).value) != expected:
            ws.cell(row=r, column=5).value = expected
            fix(label, f"E{r}", f"was C{r-3}*D{r-3}", expected)
            count += 1

    # Non-taxable rental fees (rows 69-74) — flat fee
    for r in range(69, 75):
        expected = f'=D{r}'
        if str(ws.cell(row=r, column=5).value) != expected:
            ws.cell(row=r, column=5).value = expected
            fix(label, f"E{r}", f"was D{r-3}", expected)
            count += 1

    # Rentals & Lounge subtotal (row 75)
    ws.cell(row=75, column=5).value = '=SUM(E35:E42)+SUM(E44:E51)+SUM(E53:E60)+SUM(E62:E67)+SUM(E69:E74)'
    ws.cell(row=75, column=6).value = '=SUM(F35:F42)+SUM(F44:F51)+SUM(F53:F60)+SUM(F62:F67)+SUM(F69:F74)'
    fix(label, "E75:F75", "old range refs", "fixed subtotal")
    count += 1

    # AV equipment (rows 80-89)
    for r in range(80, 90):
        expected = f'=C{r}*D{r}'
        if str(ws.cell(row=r, column=5).value) != expected:
            ws.cell(row=r, column=5).value = expected
            fix(label, f"E{r}", f"was C{r-3}*D{r-3}", expected)
            count += 1

    # AV labor (rows 91-95) — flat fee
    for r in range(91, 96):
        expected = f'=D{r}'
        if str(ws.cell(row=r, column=5).value) != expected:
            ws.cell(row=r, column=5).value = expected
            fix(label, f"E{r}", f"was D{r-3}", expected)
            count += 1

    # AV subtotal (row 96)
    ws.cell(row=96, column=5).value = '=SUM(E80:E89)+SUM(E91:E95)'
    ws.cell(row=96, column=6).value = '=SUM(F80:F89)+SUM(F91:F95)'
    fix(label, "E96:F96", "old range refs", "fixed AV subtotal")
    count += 1

    # ── Column F: Client Cost — verify VLOOKUP pattern ──
    # These should already be correct (referencing own row), but verify
    line_item_rows = (list(range(16, 26)) + list(range(27, 30)) +
                      list(range(35, 43)) + list(range(44, 52)) +
                      list(range(53, 61)) + list(range(62, 68)) +
                      list(range(69, 75)) + list(range(80, 90)) +
                      list(range(91, 96)))
    for r in line_item_rows:
        f = str(ws.cell(row=r, column=6).value)
        expected_pattern = f'=IFERROR(E{r}*(1+VLOOKUP(I{r},CategoryTable,2,FALSE)),0)'
        if f != expected_pattern:
            ws.cell(row=r, column=6).value = expected_pattern
            fix(label, f"F{r}", f"wrong VLOOKUP ref", expected_pattern)
            count += 1

    # ── Column G: Tax Rate — verify refs ──
    # Taxable product rows should ref 'Client Setup'!B19
    taxable_rows = (list(range(16, 26)) + list(range(35, 43)) +
                    list(range(44, 52)) + list(range(53, 61)) +
                    list(range(62, 68)) + list(range(80, 90)))
    for r in taxable_rows:
        expected = "='Client Setup'!B19"
        if str(ws.cell(row=r, column=7).value) != expected:
            ws.cell(row=r, column=7).value = expected
            fix(label, f"G{r}", "wrong tax ref", expected)
            count += 1

    # Non-taxable rows should be 0
    nontax_rows = list(range(27, 30)) + list(range(69, 75)) + list(range(91, 96))
    for r in nontax_rows:
        if ws.cell(row=r, column=7).value != 0:
            ws.cell(row=r, column=7).value = 0
            fix(label, f"G{r}", "not 0", "0")
            count += 1

    # ── Estimate Summary section (rows 98-104) ──
    # Row 99: Taxable Product Subtotal
    ws.cell(row=99, column=5).value = '=SUM(E16:E25)+SUM(E35:E42)+SUM(E44:E51)+SUM(E53:E60)+SUM(E62:E67)+SUM(E80:E89)'
    ws.cell(row=99, column=6).value = '=SUM(F16:F25)+SUM(F35:F42)+SUM(F44:F51)+SUM(F53:F60)+SUM(F62:F67)+SUM(F80:F89)'
    fix(label, "E99:F99", "old SUM ranges", "fixed taxable subtotal")

    # Row 100: Sales Tax
    ws.cell(row=100, column=5).value = "=IFERROR(E99*'Client Setup'!B19,0)"
    ws.cell(row=100, column=6).value = "=IFERROR(F99*'Client Setup'!B19,0)"
    fix(label, "E100:F100", "old ref E96", "E99/F99 × tax rate")

    # Row 101: Non-Taxable Fees Subtotal
    ws.cell(row=101, column=5).value = '=SUM(E27:E29)+SUM(E69:E74)+SUM(E91:E95)'
    ws.cell(row=101, column=6).value = '=SUM(F27:F29)+SUM(F69:F74)+SUM(F91:F95)'
    fix(label, "E101:F101", "old SUM ranges", "fixed non-taxable subtotal")

    # Row 102: CC Processing Fee
    ws.cell(row=102, column=5).value = '=F102'
    ws.cell(row=102, column=6).value = "=IFERROR((F99+F100+F101)*'Client Setup'!C35,0)"
    fix(label, "E102:F102", "old refs", "fixed CC processing")

    # Row 103: Third-Party Commission
    ws.cell(row=103, column=5).value = '=F103'
    ws.cell(row=103, column=6).value = "=IFERROR((F99+F100+F101)*'Client Setup'!C36,0)"
    fix(label, "E103:F103", "old refs", "fixed commission")

    # Row 104: TOTAL DECOR ESTIMATE
    ws.cell(row=104, column=5).value = '=SUM(E99:E103)'
    ws.cell(row=104, column=6).value = '=SUM(F99:F103)'
    fix(label, "E104:F104", "old SUM", "=SUM(E99:E103) / =SUM(F99:F103)")

    # ── Client Pricing (row 107) ──
    ws.cell(row=107, column=2).value = '=F104'
    fix(label, "B107", "=F101", "=F104")

    # ── Profit & Margin (rows 109-114) ──
    # Row 110: Commission
    ws.cell(row=110, column=3).value = "='Client Setup'!C36"
    ws.cell(row=110, column=4).value = '=F104*C110'
    fix(label, "C110:D110", "old refs", "F104*C110")

    # Row 111: GDP
    ws.cell(row=111, column=3).value = "=IF('Client Setup'!C37=\"Yes\",0.065,0)"
    ws.cell(row=111, column=4).value = "=IF('Client Setup'!C37=\"Yes\",F104*0.065,0)"
    fix(label, "D111", "old ref", "F104*0.065")

    # Row 112: Total Vendor Costs
    ws.cell(row=112, column=4).value = '=E99+E100+E101+E102'
    fix(label, "D112", "E96+E97+E98+E99", "E99+E100+E101+E102")

    # Row 113: QC Revenue
    ws.cell(row=113, column=4).value = '=IFERROR(F104-D110-D111-D112,0)'
    fix(label, "D113", "F101-D107-D108-D109", "F104-D110-D111-D112")

    # Row 114: QC Margin %
    ws.cell(row=114, column=4).value = '=IFERROR(IF(F104>0,D113/F104,0),0)'
    fix(label, "D114", "F101 ref", "F104 ref")

    # Margin Health (row 115) — already refs D114, correct
    # OpEx (row 117) — already correct
    # True Net Profit (row 118)
    ws.cell(row=118, column=4).value = '=D113-D117'
    # True Net Margin (row 119)
    ws.cell(row=119, column=4).value = '=IFERROR(IF(F104>0,D118/F104,0),0)'

    # ── Client-Ready Table (columns J-K) — fix row refs ──
    # These reference old A/C/F column rows. Need +3 offset.
    # Row 15: J=A10 → should be A13 (FLORALS header)
    ws.cell(row=15, column=10).value = '=A13'
    # Floral items (rows 16-25 in J-K reference A/C/F 13-22 → should be 16-25)
    old_floral = list(range(13, 26))  # old rows referenced
    new_floral = list(range(16, 29))  # new rows to reference
    for i, jr in enumerate(range(16, 29)):
        # J column: =IF(A{old}="","",A{old}&" x"&C{old})
        nr = 16 + i if i < 10 else 27 + (i - 10)  # floral product then fees
        if i < 10:
            nr = 16 + i
            ws.cell(row=jr, column=10).value = f'=IF(A{nr}="","",A{nr}&" x"&C{nr})'
            ws.cell(row=jr, column=11).value = f'=IF(A{nr}="","",ROUNDUP(F{nr}/10,0)*10)'
        elif i < 13:
            nr = 27 + (i - 10)
            ws.cell(row=jr, column=10).value = f'=IF(A{nr}="","",A{nr})'
            ws.cell(row=jr, column=11).value = f'=IF(A{nr}="","",ROUNDUP(F{nr}/10,0)*10)'
    # Floral subtotal in J-K
    ws.cell(row=29, column=10).value = 'Floral Subtotal'
    ws.cell(row=29, column=11).value = '=SUM(K16:K28)'
    fix(label, "J15:K29", "old row refs", "fixed to shifted rows")

    # Rentals section in J-K (row 31 header, then item rows)
    ws.cell(row=31, column=10).value = '=A32'
    # Seating items J-K (rows 32 onwards)
    rental_jk_map = {
        32: 35, 33: 36, 34: 37, 35: 38, 36: 39, 37: 40, 38: 41, 39: 42,  # seating
        40: 44, 41: 45, 42: 46, 43: 47, 44: 48, 45: 49, 46: 50, 47: 51,  # lounge
        48: 53, 49: 54, 50: 55, 51: 56, 52: 57, 53: 58, 54: 59, 55: 60,  # tables
        56: 62, 57: 63, 58: 64, 59: 65, 60: 66, 61: 67,  # rugs
        62: 69, 63: 70, 64: 71, 65: 72, 66: 73, 67: 74,  # rental fees
    }
    for jk_row, data_row in rental_jk_map.items():
        if data_row >= 69:  # fee rows (no qty)
            ws.cell(row=jk_row, column=10).value = f'=IF(A{data_row}="","",A{data_row})'
        else:
            ws.cell(row=jk_row, column=10).value = f'=IF(A{data_row}="","",A{data_row}&" x"&C{data_row})'
        ws.cell(row=jk_row, column=11).value = f'=IF(A{data_row}="","",ROUNDUP(F{data_row}/10,0)*10)'

    ws.cell(row=68, column=10).value = 'Rentals & Lounge Subtotal'
    ws.cell(row=68, column=11).value = '=SUM(K32:K67)'
    fix(label, "J31:K68", "old row refs", "fixed rentals client-ready")

    # AV section in J-K
    ws.cell(row=70, column=10).value = '=A77'
    av_jk = {
        71: 80, 72: 81, 73: 82, 74: 83, 75: 84, 76: 85, 77: 86, 78: 87, 79: 88, 80: 89,  # equip
        81: 91, 82: 92, 83: 93, 84: 94, 85: 95,  # labor
    }
    for jk_row, data_row in av_jk.items():
        if data_row >= 91:  # labor (no qty)
            ws.cell(row=jk_row, column=10).value = f'=IF(A{data_row}="","",A{data_row})'
        else:
            ws.cell(row=jk_row, column=10).value = f'=IF(A{data_row}="","",A{data_row}&" x"&C{data_row})'
        ws.cell(row=jk_row, column=11).value = f'=IF(A{data_row}="","",ROUNDUP(F{data_row}/10,0)*10)'

    ws.cell(row=86, column=10).value = 'AV / Production Subtotal'
    ws.cell(row=86, column=11).value = '=SUM(K71:K85)'
    fix(label, "J70:K86", "old row refs", "fixed AV client-ready")

    # Tax and Total in J-K
    ws.cell(row=88, column=10).value = 'Tax'
    ws.cell(row=88, column=11).value = '=ROUNDUP(F100/10,0)*10'
    ws.cell(row=90, column=10).value = 'TOTAL'
    ws.cell(row=90, column=11).value = '=K29+K68+K86+K88'
    fix(label, "K88:K90", "old refs", "fixed tax/total")

    # ── Summary View (columns M-N) ──
    ws.cell(row=15, column=13).value = '=A13'
    ws.cell(row=15, column=14).value = '=ROUNDUP(F30/50,0)*50'
    ws.cell(row=16, column=14).value = '=ROUNDUP(SUM(F35:F42)/50,0)*50'
    ws.cell(row=17, column=14).value = '=ROUNDUP(SUM(F44:F51)/50,0)*50'
    ws.cell(row=18, column=14).value = '=ROUNDUP(SUM(F53:F60)/50,0)*50'
    ws.cell(row=19, column=14).value = '=ROUNDUP(SUM(F62:F67)/50,0)*50'
    ws.cell(row=20, column=14).value = '=ROUNDUP((SUM(F27:F29)+SUM(F69:F74)+SUM(F91:F95))/50,0)*50'
    ws.cell(row=21, column=13).value = '=A77'
    ws.cell(row=21, column=14).value = '=ROUNDUP(SUM(F80:F89)/50,0)*50'
    ws.cell(row=22, column=14).value = '=ROUNDUP(F100/50,0)*50'
    ws.cell(row=24, column=14).value = '=SUM(N15:N22)'
    fix(label, "M-N summary", "old refs", "fixed all summary view")

    print(f"  {label}: {count}+ formulas fixed")


fix_decor_sheet(wb['Decor Estimate'], 'Decor Estimate')
fix_decor_sheet(wb['SAMPLE Decor Estimate'], 'SAMPLE Decor Estimate')


# ════════════════════════════════════════════════════════════
# FINAL AUDIT: scan for errors
# ════════════════════════════════════════════════════════════
print("\nFinal audit...")
error_patterns = ["#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!"]
errors = []
stale_refs = []
for sn in wb.sheetnames:
    ws = wb[sn]
    for row in range(1, min(ws.max_row + 1, 200)):
        for col in range(1, min(ws.max_column + 1, 20)):
            v = str(ws.cell(row=row, column=col).value) if ws.cell(row=row, column=col).value else ''
            for ep in error_patterns:
                if ep in v:
                    errors.append(f"  {sn}!{ws.cell(row=row, column=col).coordinate}: {v[:50]}")
            if "'Client Setup'!C19" in v and "VLOOKUP" not in v:
                stale_refs.append(f"  {sn}!{ws.cell(row=row, column=col).coordinate}: stale C19")

if errors:
    print(f"  ERRORS FOUND: {len(errors)}")
    for e in errors:
        print(e)
else:
    print("  No formula error strings found.")

if stale_refs:
    print(f"  STALE REFS: {len(stale_refs)}")
    for s in stale_refs:
        print(s)
else:
    print("  No stale C19 references.")


# ════════════════════════════════════════════════════════════
# SAVE
# ════════════════════════════════════════════════════════════
wb.save(DST)
print(f"\nSaved to {DST}")

# ════════════════════════════════════════════════════════════
# CHANGE SUMMARY
# ════════════════════════════════════════════════════════════
print(f"\n{'=' * 60}")
print(f"CHANGE SUMMARY ({len(changes)} changes)")
print(f"{'=' * 60}")
for c in changes:
    print(c)
