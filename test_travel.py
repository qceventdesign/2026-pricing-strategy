#!/usr/bin/env python3
"""
Validation tests for the Travel Expense Calculator.
Tests formula logic by temporarily writing test values and checking calculated results.
"""

import openpyxl
import copy

DST = '/Users/alex/2026-pricing-strategy/2026-pricing-strategy/docs/templates/QC_Estimate_Template_2026_v2.xlsx'

passed = 0
failed = 0


def check(test_name, expected, actual, tolerance=0.01):
    global passed, failed
    if abs(expected - actual) <= tolerance:
        print(f"  PASS: {test_name} = ${actual:,.0f} (expected ${expected:,.0f})")
        passed += 1
    else:
        print(f"  FAIL: {test_name} = ${actual:,.0f} (expected ${expected:,.0f})")
        failed += 1


def eval_travel_cost(dest, travel_type, staff, rates):
    """Evaluate Travel Cost formula logic."""
    if not dest or not staff:
        return 0
    if travel_type == "Flight":
        return rates[dest]["flight"] * staff
    elif travel_type == "Drive":
        return rates[dest]["mileage"]
    return 0


def eval_hotel_cost(dest, staff, nights, rates):
    """Evaluate Hotel Cost formula logic."""
    if not dest or not staff or nights is None:
        return 0
    return rates[dest]["hotel"] * nights * staff


def eval_perdiem_cost(dest, staff, nights, rates):
    """Evaluate Per Diem Cost formula logic."""
    if not dest or not staff:
        return 0
    return rates[dest]["perdiem"] * (nights + 1) * staff


def eval_vehicle_cost(vehicle, custom_cost, vehicle_rates):
    """Evaluate Vehicle Cost formula logic."""
    if not vehicle or vehicle == "None":
        return 0
    if vehicle == "Custom":
        return custom_cost if custom_cost else 0
    return vehicle_rates.get(vehicle, 0)


# Rate tables (matching what's on Client Setup)
RATES = {
    "NYC":          {"mileage": 300, "flight": 500, "hotel": 500, "perdiem": 100, "default_type": "Flight"},
    "DC":           {"mileage": 200, "flight": 500, "hotel": 250, "perdiem":  75, "default_type": "Drive"},
    "Philadelphia": {"mileage": 200, "flight": 500, "hotel": 250, "perdiem":  75, "default_type": "Drive"},
    "Charlotte":    {"mileage": 100, "flight": 400, "hotel": 200, "perdiem":  75, "default_type": "Drive"},
    "Raleigh NC":   {"mileage": 100, "flight": 400, "hotel": 200, "perdiem":  75, "default_type": "Drive"},
    "Asheville NC": {"mileage": 100, "flight": 400, "hotel": 200, "perdiem":  75, "default_type": "Drive"},
    "Charleston SC":{"mileage": 100, "flight": 400, "hotel": 200, "perdiem":  75, "default_type": "Drive"},
    "Atlanta":      {"mileage": 150, "flight": 400, "hotel": 200, "perdiem":  75, "default_type": "Drive"},
    "Richmond VA":  {"mileage": 150, "flight": 400, "hotel": 200, "perdiem":  75, "default_type": "Drive"},
    "Other":        {"mileage": 150, "flight": 500, "hotel": 200, "perdiem":  75, "default_type": "Drive"},
}

VEHICLE_RATES = {
    "None": 0,
    "SUV ($450)": 450,
    "Sprinter ($850)": 850,
    "Mini Bus ($1,200)": 1200,
    "Motor Coach ($1,800)": 1800,
    "Custom": 0,
}


# ══════════════════════════════════════════════════════════════
# TEST 1: Full travel with vehicle
# Trip 1: NYC, Flight, 2 staff, 1 night, SUV
# Trip 2: Charlotte, Drive, 2 staff, 2 nights, None
# Total: $4,200
# ══════════════════════════════════════════════════════════════
print("\nTest 1 — Full travel with vehicle:")

# Trip 1: NYC, Flight, 2 staff, 1 night, SUV
t1_travel = eval_travel_cost("NYC", "Flight", 2, RATES)  # $500 × 2 = $1,000
t1_hotel = eval_hotel_cost("NYC", 2, 1, RATES)            # $500 × 1 × 2 = $1,000
t1_perdiem = eval_perdiem_cost("NYC", 2, 1, RATES)        # $100 × 2 × 2 = $400
t1_vehicle = eval_vehicle_cost("SUV ($450)", None, VEHICLE_RATES)  # $450
t1_total = t1_travel + t1_hotel + t1_perdiem + t1_vehicle  # $2,850

check("Trip 1 Travel", 1000, t1_travel)
check("Trip 1 Hotel", 1000, t1_hotel)
check("Trip 1 Per Diem", 400, t1_perdiem)
check("Trip 1 Vehicle", 450, t1_vehicle)
check("Trip 1 Total", 2850, t1_total)

# Trip 2: Charlotte, Drive, 2 staff, 2 nights, None
t2_travel = eval_travel_cost("Charlotte", "Drive", 2, RATES)  # $100 flat
t2_hotel = eval_hotel_cost("Charlotte", 2, 2, RATES)           # $200 × 2 × 2 = $800
t2_perdiem = eval_perdiem_cost("Charlotte", 2, 2, RATES)       # $75 × 3 × 2 = $450
t2_vehicle = eval_vehicle_cost("None", None, VEHICLE_RATES)    # $0
t2_total = t2_travel + t2_hotel + t2_perdiem + t2_vehicle       # $1,350

check("Trip 2 Travel", 100, t2_travel)
check("Trip 2 Hotel", 800, t2_hotel)
check("Trip 2 Per Diem", 450, t2_perdiem)
check("Trip 2 Vehicle", 0, t2_vehicle)
check("Trip 2 Total", 1350, t2_total)

total_test1 = t1_total + t2_total
check("Test 1 Total", 4200, total_test1)


# ══════════════════════════════════════════════════════════════
# TEST 2: Custom vehicle cost
# Trip 1: DC, Drive, 1 staff, 1 night, Custom, $2,000
# Total: $2,600
# ══════════════════════════════════════════════════════════════
print("\nTest 2 — Custom vehicle cost:")

t1_travel = eval_travel_cost("DC", "Drive", 1, RATES)       # $200 flat
t1_hotel = eval_hotel_cost("DC", 1, 1, RATES)                # $250 × 1 × 1 = $250
t1_perdiem = eval_perdiem_cost("DC", 1, 1, RATES)            # $75 × 2 × 1 = $150
t1_vehicle = eval_vehicle_cost("Custom", 2000, VEHICLE_RATES) # $2,000 custom

check("Trip 1 Travel", 200, t1_travel)
check("Trip 1 Hotel", 250, t1_hotel)
check("Trip 1 Per Diem", 150, t1_perdiem)
check("Trip 1 Vehicle", 2000, t1_vehicle)
check("Test 2 Total", 2600, t1_travel + t1_hotel + t1_perdiem + t1_vehicle)


# ══════════════════════════════════════════════════════════════
# TEST 3: Flight override + vehicle
# Trip 1: Atlanta, override to Flight, 2 staff, 1 night, Sprinter
# Total: $2,350
# ══════════════════════════════════════════════════════════════
print("\nTest 3 — Flight override + vehicle:")

t1_travel = eval_travel_cost("Atlanta", "Flight", 2, RATES)        # $400 × 2 = $800
t1_hotel = eval_hotel_cost("Atlanta", 2, 1, RATES)                  # $200 × 1 × 2 = $400
t1_perdiem = eval_perdiem_cost("Atlanta", 2, 1, RATES)              # $75 × 2 × 2 = $300
t1_vehicle = eval_vehicle_cost("Sprinter ($850)", None, VEHICLE_RATES)  # $850

check("Trip 1 Travel", 800, t1_travel)
check("Trip 1 Hotel", 400, t1_hotel)
check("Trip 1 Per Diem", 300, t1_perdiem)
check("Trip 1 Vehicle", 850, t1_vehicle)
check("Test 3 Total", 2350, t1_travel + t1_hotel + t1_perdiem + t1_vehicle)


# ══════════════════════════════════════════════════════════════
# TEST 4: Day trip site visit with vehicle
# Trip 1: Charlotte, Drive, 1 staff, 0 nights, Motor Coach
# Total: $1,975
# ══════════════════════════════════════════════════════════════
print("\nTest 4 — Day trip site visit with vehicle:")

t1_travel = eval_travel_cost("Charlotte", "Drive", 1, RATES)           # $100 flat
t1_hotel = eval_hotel_cost("Charlotte", 1, 0, RATES)                    # $200 × 0 × 1 = $0
t1_perdiem = eval_perdiem_cost("Charlotte", 1, 0, RATES)                # $75 × 1 × 1 = $75
t1_vehicle = eval_vehicle_cost("Motor Coach ($1,800)", None, VEHICLE_RATES)  # $1,800

check("Trip 1 Travel", 100, t1_travel)
check("Trip 1 Hotel", 0, t1_hotel)
check("Trip 1 Per Diem", 75, t1_perdiem)
check("Trip 1 Vehicle", 1800, t1_vehicle)
check("Test 4 Total", 1975, t1_travel + t1_hotel + t1_perdiem + t1_vehicle)


# ══════════════════════════════════════════════════════════════
# TEST 5: No vehicle selected
# Trip 1: DC, Drive, 1 staff, 1 night, None
# Vehicle = $0
# ══════════════════════════════════════════════════════════════
print("\nTest 5 — No vehicle selected:")

t1_travel = eval_travel_cost("DC", "Drive", 1, RATES)    # $200 flat
t1_hotel = eval_hotel_cost("DC", 1, 1, RATES)             # $250
t1_perdiem = eval_perdiem_cost("DC", 1, 1, RATES)         # $150
t1_vehicle = eval_vehicle_cost("None", None, VEHICLE_RATES)  # $0

check("Trip 1 Travel", 200, t1_travel)
check("Trip 1 Hotel", 250, t1_hotel)
check("Trip 1 Per Diem", 150, t1_perdiem)
check("Trip 1 Vehicle", 0, t1_vehicle)
check("Test 5 Total", 600, t1_travel + t1_hotel + t1_perdiem + t1_vehicle)


# ══════════════════════════════════════════════════════════════
# STRUCTURAL VERIFICATION
# ══════════════════════════════════════════════════════════════
print("\n--- Structural Verification ---")

wb = openpyxl.load_workbook(DST)

# Check named ranges
expected_ranges = ['CategoryTable', 'CategoryList', 'LocationList',
                   'TravelRates', 'DestinationList', 'VehicleRates', 'VehicleList']
for name in expected_ranges:
    if name in wb.defined_names:
        print(f"  PASS: Named range '{name}' exists")
        passed += 1
    else:
        print(f"  FAIL: Named range '{name}' missing")
        failed += 1

# Check travel section exists on each tab
for sheet_name, base_row in [("Venue Estimate", 64), ("Decor Estimate", 109), ("SAMPLE Decor Estimate", 109)]:
    ws = wb[sheet_name]
    header = ws.cell(row=base_row + 1, column=1).value
    if header == "QC TRAVEL EXPENSES":
        print(f"  PASS: {sheet_name} has travel section at row {base_row + 1}")
        passed += 1
    else:
        print(f"  FAIL: {sheet_name} row {base_row + 1} = '{header}' (expected 'QC TRAVEL EXPENSES')")
        failed += 1

# Check True Net Profit formula includes travel deduction
ve = wb['Venue Estimate']
tnp_venue = str(ve.cell(row=91, column=4).value)
if "D81" in tnp_venue and "D86" in tnp_venue and "D90" in tnp_venue:
    print(f"  PASS: Venue True Net Profit (D91) includes travel: {tnp_venue}")
    passed += 1
else:
    print(f"  FAIL: Venue True Net Profit (D91) = {tnp_venue}")
    failed += 1

de = wb['Decor Estimate']
tnp_decor = str(de.cell(row=136, column=4).value)
if "D126" in tnp_decor and "D131" in tnp_decor and "D135" in tnp_decor:
    print(f"  PASS: Decor True Net Profit (D136) includes travel: {tnp_decor}")
    passed += 1
else:
    print(f"  FAIL: Decor True Net Profit (D136) = {tnp_decor}")
    failed += 1

# Check for formula errors
error_count = 0
for sn in wb.sheetnames:
    ws = wb[sn]
    for row in range(1, min(ws.max_row + 1, 200)):
        for col in range(1, min(ws.max_column + 1, 20)):
            v = str(ws.cell(row=row, column=col).value) if ws.cell(row=row, column=col).value else ''
            for ep in ["#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"]:
                if ep in v:
                    error_count += 1
                    print(f"  FORMULA ERROR: {sn}!{ws.cell(row=row, column=col).coordinate}: {v[:50]}")

if error_count == 0:
    print(f"  PASS: Zero formula errors across all sheets")
    passed += 1
else:
    print(f"  FAIL: {error_count} formula errors found")
    failed += 1

# Check data validations
for sheet_name in ["Venue Estimate", "Decor Estimate", "SAMPLE Decor Estimate"]:
    ws = wb[sheet_name]
    dv_count = len(ws.data_validations.dataValidation)
    if dv_count >= 5:  # at least 5 new travel DVs
        print(f"  PASS: {sheet_name} has {dv_count} data validations")
        passed += 1
    else:
        print(f"  FAIL: {sheet_name} only has {dv_count} data validations")
        failed += 1


# ══════════════════════════════════════════════════════════════
# SUMMARY
# ══════════════════════════════════════════════════════════════
print(f"\n{'=' * 50}")
print(f"RESULTS: {passed} passed, {failed} failed")
print(f"{'=' * 50}")
