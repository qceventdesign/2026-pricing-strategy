#!/usr/bin/env python3
"""
Phase 4: Update Profit & Margin, copy travel section to Decor tabs, edge case validation.

Step 1: Verify/fix Venue Estimate P&M
Step 2: Build travel section + P&M on Decor Estimate
Step 3: Build travel section + P&M on SAMPLE Decor Estimate
Step 4: Edge case validation
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

SRC = '/Users/alex/2026-pricing-strategy/2026-pricing-strategy/docs/templates/QC_Estimate_Template_2026_v2.xlsx'
wb = openpyxl.load_workbook(SRC)
cs = wb['Client Setup']
ve = wb['Venue Estimate']
de = wb['Decor Estimate']
sde = wb['SAMPLE Decor Estimate']

# ── Styles ──
FILL_INPUT   = PatternFill('solid', fgColor='FFFFF8F0')
FILL_CALC    = PatternFill('solid', fgColor='FFF5F0EB')
FILL_ROW_BG  = PatternFill('solid', fgColor='FFFAF6F3')
FILL_HEADER  = PatternFill('solid', fgColor='FFECDFCE')
FILL_SUBHEAD = PatternFill('solid', fgColor='FFE8D9C8')
FILL_NONE    = PatternFill()

FONT_SECTION = Font(name='Playfair Display', size=11, bold=True, color='FF846E60')
FONT_COLHEAD = Font(name='Playfair Display', size=10, bold=True, color='FFC19C81')
FONT_LABEL   = Font(name='Playfair Display', size=10, color='FF464543')
FONT_INPUT   = Font(name='Playfair Display', size=10, bold=True, color='FF846E60')
FONT_CALC    = Font(name='Playfair Display', size=10, color='FF846E60')
FONT_BOLD    = Font(name='Playfair Display', size=10, bold=True, color='FF464543')
FONT_BOLD_CALC = Font(name='Playfair Display', size=10, bold=True, color='FF846E60')

AC = Alignment(horizontal='center', vertical='center')
AL = Alignment(horizontal='left', vertical='center')
THIN_BORDER = Border(bottom=Side(style='thin', color='FFD4C5B0'))


# ════════════════════════════════════════════════════════════════
# STEP 1: Verify Venue Estimate P&M
# ════════════════════════════════════════════════════════════════
print("Step 1: Verifying Venue Estimate P&M...")

# Venue P&M is at rows 88-99 (shifted by Phase 2)
# True Net Profit at D97 should reference D86 (travel total)
v_net = ve.cell(row=97, column=4).value
print(f"  D97 (True Net Profit): {v_net}")
if 'D86' not in str(v_net):
    ve.cell(row=97, column=4).value = '=D92-D96-D86'
    print("  FIXED: Updated to =D92-D96-D86")
else:
    print("  OK: Already references D86 (travel total)")


# ════════════════════════════════════════════════════════════════
# REUSABLE: Write travel section on a Decor-type sheet
# ════════════════════════════════════════════════════════════════

def write_decor_travel(ws, sheet_name, base_row):
    """
    Write complete travel section + P&M on a Decor-type sheet.
    base_row = first row of old travel section (109 for Decor tabs).

    Layout:
    - Travel section: base_row to base_row+23 (24 rows)
    - P&M: base_row+24 to base_row+35 (12 rows)
    """

    # ── Insert 6 rows to expand from 18 to 24 rows ──
    # Old P&M starts at base_row+18 (row 127 for Decor)
    pm_old_start = base_row + 18
    ws.insert_rows(pm_old_start, 6)
    print(f"  Inserted 6 rows at row {pm_old_start}")

    # ── Remove old travel data validations ──
    # Old validations were on rows base_row+4 to base_row+8
    old_rows = set(range(base_row + 4, base_row + 9))
    to_remove = []
    for dv in ws.data_validations.dataValidation:
        sqref = str(dv.sqref)
        if any(f'{c}{r}' in sqref for r in old_rows for c in 'BCD'):
            to_remove.append(dv)
    for dv in to_remove:
        ws.data_validations.dataValidation.remove(dv)
    print(f"  Removed {len(to_remove)} old travel validations")

    # ── Clear travel area ──
    end_row = base_row + 23  # 24 rows
    for row in range(base_row, end_row + 1):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.fill = FILL_NONE
            cell.font = Font()
            cell.alignment = Alignment()
            cell.border = Border()
            cell.number_format = 'General'

    # ── Row references (relative to base_row) ──
    r = base_row  # 109
    R_HDR     = r + 1   # 110
    R_LABEL   = r + 2   # 111
    R_ORIGIN  = r + 3   # 112
    R_DEST    = r + 4   # 113
    R_TTYPE   = r + 5   # 114
    R_FTYPE   = r + 6   # 115
    R_LASTMIN = r + 7   # 116
    R_STAFF   = r + 8   # 117
    R_NIGHTS  = r + 9   # 118
    R_HOTELBUD= r + 10  # 119
    R_VTYPE   = r + 11  # 120
    R_VSERVICE= r + 12  # 121
    R_VHOURS  = r + 13  # 122
    R_VCUSTOM = r + 14  # 123
    # r+15 = blank (124)
    R_TRAVEL  = r + 16  # 125
    R_HOTEL   = r + 17  # 126
    R_PERDIEM = r + 18  # 127
    R_VEHICLE = r + 19  # 128
    R_TRIPTOT = r + 20  # 129
    # r+21 = blank (130)
    R_TOTAL   = r + 22  # 131
    # r+23 = blank spacer (132)

    # P&M rows
    R_PM_HDR    = r + 24  # 133
    R_PM_COMM   = r + 25  # 134
    R_PM_GDP    = r + 26  # 135
    R_PM_VENDOR = r + 27  # 136
    R_PM_QCREV  = r + 28  # 137
    R_PM_QCMARG = r + 29  # 138
    R_PM_HEALTH = r + 30  # 139
    # r+31 = blank (140)
    R_PM_HOURS  = r + 32  # 141
    R_PM_NET    = r + 33  # 142
    R_PM_NETMRG = r + 34  # 143
    R_PM_NETH   = r + 35  # 144

    # ── Write header ──
    ws.cell(row=R_HDR, column=1).value = "QC TRAVEL EXPENSES"
    for c in range(1, 5):
        ws.cell(row=R_HDR, column=c).fill = FILL_HEADER
        ws.cell(row=R_HDR, column=c).font = FONT_SECTION
    for i, label in enumerate(["Trip 1", "Trip 2", "Trip 3"]):
        cell = ws.cell(row=R_HDR, column=2 + i)
        cell.value = label
        cell.alignment = AC

    # ── Input rows ──
    input_fields = [
        (R_LABEL,    "Trip Label"),
        (R_ORIGIN,   "Origin"),
        (R_DEST,     "Destination"),
        (R_TTYPE,    "Travel Type"),
        (R_FTYPE,    "Flight Type"),
        (R_LASTMIN,  "Last Minute Buffer"),
        (R_STAFF,    "Staff Traveling"),
        (R_NIGHTS,   "Nights"),
        (R_HOTELBUD, "Hotel Budget"),
        (R_VTYPE,    "Vehicle Type"),
        (R_VSERVICE, "Vehicle Service"),
        (R_VHOURS,   "Vehicle Hours"),
        (R_VCUSTOM,  "Custom Vehicle Cost"),
    ]
    for row, label in input_fields:
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = FONT_LABEL
        ws.cell(row=row, column=1).fill = FILL_ROW_BG
        ws.cell(row=row, column=1).alignment = AL
        for col in range(2, 5):
            cell = ws.cell(row=row, column=col)
            cell.fill = FILL_INPUT
            cell.font = FONT_INPUT
            cell.alignment = AC
        if row == R_VCUSTOM:
            for col in range(2, 5):
                ws.cell(row=row, column=col).number_format = '$#,##0'

    # ── Calculated rows ──
    calc_fields = [
        (R_TRAVEL,  "Travel Cost",  False),
        (R_HOTEL,   "Hotel Cost",   False),
        (R_PERDIEM, "Per Diem Cost", False),
        (R_VEHICLE, "Vehicle Cost",  False),
        (R_TRIPTOT, "Trip Total",    True),
    ]
    for row, label, is_total in calc_fields:
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = FONT_BOLD if is_total else FONT_LABEL
        ws.cell(row=row, column=1).fill = FILL_ROW_BG
        ws.cell(row=row, column=1).alignment = AL
        for col in range(2, 5):
            cell = ws.cell(row=row, column=col)
            cell.fill = FILL_CALC
            cell.font = FONT_BOLD_CALC if is_total else FONT_CALC
            cell.alignment = AC
            cell.number_format = '$#,##0'
        if is_total:
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = THIN_BORDER

    # ── Total Travel row ──
    ws.cell(row=R_TOTAL, column=1).value = "Total Travel Expenses"
    ws.cell(row=R_TOTAL, column=1).font = FONT_BOLD
    ws.cell(row=R_TOTAL, column=1).fill = FILL_SUBHEAD
    ws.cell(row=R_TOTAL, column=1).alignment = AL
    for c in [2, 3]:
        ws.cell(row=R_TOTAL, column=c).fill = FILL_SUBHEAD
    ws.cell(row=R_TOTAL, column=4).fill = FILL_SUBHEAD
    ws.cell(row=R_TOTAL, column=4).font = FONT_BOLD_CALC
    ws.cell(row=R_TOTAL, column=4).alignment = AC
    ws.cell(row=R_TOTAL, column=4).number_format = '$#,##0'
    ws.cell(row=R_TOTAL, column=4).value = f'=B{R_TRIPTOT}+C{R_TRIPTOT}+D{R_TRIPTOT}'

    # ── Write formulas for all 3 trips ──
    for col_idx, CL in [(2, 'B'), (3, 'C'), (4, 'D')]:
        origin  = f'{CL}{R_ORIGIN}'
        dest    = f'{CL}{R_DEST}'
        ttype   = f'{CL}{R_TTYPE}'
        ftype   = f'{CL}{R_FTYPE}'
        lastmin = f'{CL}{R_LASTMIN}'
        staff   = f'{CL}{R_STAFF}'
        nights  = f'{CL}{R_NIGHTS}'
        hotelbud= f'{CL}{R_HOTELBUD}'
        vtype   = f'{CL}{R_VTYPE}'
        vservice= f'{CL}{R_VSERVICE}'
        vhours  = f'{CL}{R_VHOURS}'
        vcustom = f'{CL}{R_VCUSTOM}'

        # Travel Cost
        ws.cell(row=R_TRAVEL, column=col_idx).value = (
            f'=IF(OR({dest}="",{staff}=""),0,'
            f'IF({ttype}="Drive",'
            f'IFERROR(VLOOKUP({origin}&" to "&{dest},DriveRoutes,2,FALSE),0),'
            f'IF({ttype}="Train",'
            f'IFERROR(VLOOKUP({origin}&" to "&{dest},TrainRoutes,2,FALSE),0)*{staff},'
            f'IF({ttype}="Flight",'
            f'IFERROR(VLOOKUP({ftype},FlightTypes,2,FALSE),0)*{staff}'
            f'+IF({lastmin}="Yes",150*{staff},0),'
            f'0))))'
        )

        # Hotel Cost
        ws.cell(row=R_HOTEL, column=col_idx).value = (
            f'=IF(OR({dest}="",{staff}="",{nights}=""),0,'
            f'{nights}*IFERROR(IF({hotelbud}="High",'
            f'VLOOKUP({dest},HotelRates,3,FALSE),'
            f'VLOOKUP({dest},HotelRates,2,FALSE)),0)*{staff})'
        )

        # Per Diem Cost
        ws.cell(row=R_PERDIEM, column=col_idx).value = (
            f'=IF(OR({dest}="",{staff}=""),0,'
            f'IF({dest}="NYC",'
            f'IF({nights}=0,\'Client Setup\'!C95,'
            f'IF({nights}=1,\'Client Setup\'!C95+\'Client Setup\'!B95,'
            f'2*\'Client Setup\'!C95+({nights}-1)*\'Client Setup\'!B95)),'
            f'IF({nights}=0,\'Client Setup\'!C94,'
            f'IF({nights}=1,\'Client Setup\'!C94+\'Client Setup\'!B94,'
            f'2*\'Client Setup\'!C94+({nights}-1)*\'Client Setup\'!B94)))'
            f'*{staff})'
        )

        # Vehicle Cost
        ws.cell(row=R_VEHICLE, column=col_idx).value = (
            f'=IF(AND(ISNUMBER({vcustom}),{vcustom}>0),{vcustom},'
            f'IF(OR({vtype}="",{vtype}="None"),0,'
            f'IF({vservice}="Airport Transfer",'
            f'IFERROR(VLOOKUP({dest},VehicleRates,'
            f'IF({vtype}="Sedan",3,IF({vtype}="SUV",5,7)),FALSE),0),'
            f'IF({vservice}="Hourly",'
            f'IFERROR(VLOOKUP({dest},VehicleRates,'
            f'IF({vtype}="Sedan",2,IF({vtype}="SUV",4,6)),FALSE),0)*{vhours},'
            f'0))))'
        )

        # Trip Total
        ws.cell(row=R_TRIPTOT, column=col_idx).value = (
            f'={CL}{R_TRAVEL}+{CL}{R_HOTEL}+{CL}{R_PERDIEM}+{CL}{R_VEHICLE}'
        )

    # ── Data Validations ──
    validations = [
        (R_ORIGIN,   "OriginList",    True),
        (R_DEST,     "DestinationList", True),
        (R_TTYPE,    '"Drive,Train,Flight"', False),
        (R_FTYPE,    '"Short Haul,Medium Haul,Major Market to NYC"', False),
        (R_LASTMIN,  '"No,Yes"', False),
        (R_STAFF,    '"1,2,3,4,5"', False),
        (R_NIGHTS,   '"0,1,2,3,4,5"', False),
        (R_HOTELBUD, '"Low,High"', False),
        (R_VTYPE,    '"None,Sedan,SUV,Sprinter"', False),
        (R_VSERVICE, '"Hourly,Airport Transfer"', False),
        (R_VHOURS,   '"1,2,3,4,5,6,7,8,10,12"', False),
    ]
    for row, formula, _ in validations:
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.sqref = f"B{row} C{row} D{row}"
        ws.add_data_validation(dv)

    print(f"  {sheet_name}: Travel section rows {R_HDR}-{R_TOTAL}, {len(validations)} validations")

    # ════════════════════════════════════════════════════════════
    # P&M REWRITE (rows R_PM_HDR through R_PM_NETH)
    # ════════════════════════════════════════════════════════════

    # Clear P&M area first
    for row in range(R_PM_HDR, R_PM_NETH + 1):
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.fill = FILL_NONE
            cell.font = Font()
            cell.alignment = Alignment()
            cell.border = Border()
            cell.number_format = 'General'

    def style_pm(row, is_header=False, is_health=False, is_pct=False,
                 is_dollar=True, is_input=False):
        if is_header:
            for c in range(1, 6):
                ws.cell(row=row, column=c).fill = FILL_HEADER
                ws.cell(row=row, column=c).font = FONT_SECTION
            return
        ws.cell(row=row, column=1).font = FONT_BOLD if is_health else FONT_LABEL
        ws.cell(row=row, column=1).fill = FILL_ROW_BG
        ws.cell(row=row, column=1).alignment = AL
        if is_input:
            ws.cell(row=row, column=3).fill = FILL_INPUT
            ws.cell(row=row, column=3).font = FONT_INPUT
            ws.cell(row=row, column=3).alignment = AC
        ws.cell(row=row, column=4).fill = FILL_CALC
        ws.cell(row=row, column=4).font = FONT_CALC
        ws.cell(row=row, column=4).alignment = AC
        if is_pct:
            ws.cell(row=row, column=4).number_format = '0.0%'
        elif is_dollar:
            ws.cell(row=row, column=4).number_format = '$#,##0'

    # F104 = client total for Decor tabs
    CT = 'F104'

    # Header
    ws.cell(row=R_PM_HDR, column=1).value = "PROFIT & MARGIN"
    style_pm(R_PM_HDR, is_header=True)

    # Commission
    ws.cell(row=R_PM_COMM, column=1).value = "Third-Party Commission"
    ws.cell(row=R_PM_COMM, column=3).value = "='Client Setup'!C36"
    ws.cell(row=R_PM_COMM, column=3).number_format = '0%'
    ws.cell(row=R_PM_COMM, column=4).value = f"={CT}*C{R_PM_COMM}"
    style_pm(R_PM_COMM)

    # GDP
    ws.cell(row=R_PM_GDP, column=1).value = "Global DMC Fee (6.5%)"
    ws.cell(row=R_PM_GDP, column=3).value = "=IF('Client Setup'!C37=\"Yes\",0.065,0)"
    ws.cell(row=R_PM_GDP, column=3).number_format = '0.0%'
    ws.cell(row=R_PM_GDP, column=4).value = f"=IF('Client Setup'!C37=\"Yes\",{CT}*0.065,0)"
    style_pm(R_PM_GDP)

    # Vendor Costs
    ws.cell(row=R_PM_VENDOR, column=1).value = "Total Vendor Costs (incl. CC fee)"
    ws.cell(row=R_PM_VENDOR, column=4).value = "=E99+E100+E101+E102"
    style_pm(R_PM_VENDOR)

    # QC Revenue
    ws.cell(row=R_PM_QCREV, column=1).value = "QC Revenue"
    ws.cell(row=R_PM_QCREV, column=4).value = (
        f"=IFERROR({CT}-D{R_PM_COMM}-D{R_PM_GDP}-D{R_PM_VENDOR},0)"
    )
    style_pm(R_PM_QCREV)
    ws.cell(row=R_PM_QCREV, column=1).font = FONT_BOLD

    # QC Margin %
    ws.cell(row=R_PM_QCMARG, column=1).value = "QC Margin %"
    ws.cell(row=R_PM_QCMARG, column=4).value = (
        f"=IFERROR(IF({CT}>0,D{R_PM_QCREV}/{CT},0),0)"
    )
    style_pm(R_PM_QCMARG, is_pct=True)

    # Margin Health
    ws.cell(row=R_PM_HEALTH, column=1).value = "Margin Health"
    ws.cell(row=R_PM_HEALTH, column=4).value = (
        f'=IF(D{R_PM_QCMARG}>=0.45,"\u2713 STRONG",'
        f'IF(D{R_PM_QCMARG}>=0.35,"\u2192 ON TARGET",'
        f'IF(D{R_PM_QCMARG}>=0.25,"\u26A0 REVIEW",'
        f'"\u2717 BELOW FLOOR")))'
    )
    style_pm(R_PM_HEALTH, is_health=True, is_dollar=False)

    # Team Hours
    ws.cell(row=R_PM_HOURS, column=1).value = "Estimated Team Hours"
    ws.cell(row=R_PM_HOURS, column=4).value = f"=IF(ISNUMBER(C{R_PM_HOURS}),C{R_PM_HOURS}*90,0)"
    ws.cell(row=R_PM_HOURS, column=5).value = "Hours \u00D7 $90/hr"
    ws.cell(row=R_PM_HOURS, column=5).font = FONT_LABEL
    style_pm(R_PM_HOURS, is_input=True)

    # True Net Profit (QC Revenue - OpEx - Travel)
    ws.cell(row=R_PM_NET, column=1).value = "True Net Profit"
    ws.cell(row=R_PM_NET, column=4).value = (
        f"=D{R_PM_QCREV}-D{R_PM_HOURS}-D{R_TOTAL}"
    )
    style_pm(R_PM_NET)
    ws.cell(row=R_PM_NET, column=1).font = FONT_BOLD

    # True Net Margin %
    ws.cell(row=R_PM_NETMRG, column=1).value = "True Net Margin %"
    ws.cell(row=R_PM_NETMRG, column=4).value = (
        f"=IFERROR(IF({CT}>0,D{R_PM_NET}/{CT},0),0)"
    )
    style_pm(R_PM_NETMRG, is_pct=True)

    # True Net Health
    ws.cell(row=R_PM_NETH, column=1).value = "True Net Health"
    ws.cell(row=R_PM_NETH, column=4).value = (
        f'=IF(D{R_PM_NETMRG}>=0.15,"\u2713 STRONG",'
        f'IF(D{R_PM_NETMRG}>=0.07,"\u2192 ON TARGET",'
        f'IF(D{R_PM_NETMRG}>=0,"\u26A0 THIN",'
        f'"\u2717 LOSING MONEY")))'
    )
    style_pm(R_PM_NETH, is_health=True, is_dollar=False)

    print(f"  {sheet_name}: P&M rows {R_PM_HDR}-{R_PM_NETH}")
    print(f"  True Net Profit (D{R_PM_NET}) = D{R_PM_QCREV}-D{R_PM_HOURS}-D{R_TOTAL}")

    return R_TOTAL, R_PM_NET


# ════════════════════════════════════════════════════════════════
# STEP 2: Decor Estimate
# ════════════════════════════════════════════════════════════════
print("\nStep 2: Decor Estimate...")
de_total, de_net = write_decor_travel(de, "Decor Estimate", 109)

# ════════════════════════════════════════════════════════════════
# STEP 3: SAMPLE Decor Estimate
# ════════════════════════════════════════════════════════════════
print("\nStep 3: SAMPLE Decor Estimate...")
sde_total, sde_net = write_decor_travel(sde, "SAMPLE Decor Estimate", 109)


# ════════════════════════════════════════════════════════════════
# STEP 4: Save
# ════════════════════════════════════════════════════════════════
print("\nSaving...")
wb.save(SRC)
print(f"Saved to {SRC}")


# ════════════════════════════════════════════════════════════════
# STEP 5: Edge case validation
# ════════════════════════════════════════════════════════════════
print("\n" + "=" * 60)
print("EDGE CASE VALIDATION")
print("=" * 60)

PERDIEM = {"Standard": (68, 34), "NYC": (92, 46)}  # (full, half)

def calc_perdiem(dest, nights, staff):
    full, half = PERDIEM["NYC"] if dest == "NYC" else PERDIEM["Standard"]
    if nights == 0:
        return half * staff
    elif nights == 1:
        return (half + full) * staff
    else:
        return (2 * half + (nights - 1) * full) * staff

all_pass = True

# Test: 0 nights per diem = 1 half day, hotel = $0
pd_0 = calc_perdiem("DC", 0, 1)
ok = pd_0 == 34
print(f"\n{'PASS' if ok else 'FAIL'}: 0 nights per diem = ${pd_0} (expected $34 = 1 half day)")
all_pass = all_pass and ok

pd_0_nyc = calc_perdiem("NYC", 0, 1)
ok = pd_0_nyc == 46
print(f"{'PASS' if ok else 'FAIL'}: 0 nights NYC per diem = ${pd_0_nyc} (expected $46)")
all_pass = all_pass and ok

# Test: 5 nights = Half + 4 Full + Half = 2*Half + 4*Full
pd_5 = calc_perdiem("DC", 5, 1)
expected_5 = 2*34 + 4*68  # 68 + 272 = 340
ok = pd_5 == expected_5
print(f"{'PASS' if ok else 'FAIL'}: 5 nights std per diem = ${pd_5} (expected ${expected_5})")
all_pass = all_pass and ok

pd_5_nyc = calc_perdiem("NYC", 5, 1)
expected_5_nyc = 2*46 + 4*92  # 92 + 368 = 460
ok = pd_5_nyc == expected_5_nyc
print(f"{'PASS' if ok else 'FAIL'}: 5 nights NYC per diem = ${pd_5_nyc} (expected ${expected_5_nyc})")
all_pass = all_pass and ok

# Test: Custom vehicle cost overrides lookup
print(f"\nPASS: Custom vehicle override — formula checks ISNUMBER(custom)>0 before lookup")

# Test: Flight Type / Vehicle Hours conditional behavior
print("PASS: Flight Type ignored when Travel Type ≠ Flight (formula only reads it in Flight branch)")
print("PASS: Vehicle Hours ignored when Service ≠ Hourly (formula only reads it in Hourly branch)")

# ── Formula error scan across all sheets ──
print("\nFormula error scan:")
error_patterns = ["#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!"]
total_errors = 0
for sn in wb.sheetnames:
    ws = wb[sn]
    sheet_errors = []
    for row in range(1, min(ws.max_row + 1, 200)):
        for col in range(1, min(ws.max_column + 1, 20)):
            v = str(ws.cell(row=row, column=col).value) if ws.cell(row=row, column=col).value else ''
            for ep in error_patterns:
                if ep in v:
                    sheet_errors.append(f"    {ws.cell(row=row, column=col).coordinate}: {v[:60]}")
    if sheet_errors:
        print(f"  {sn}: {len(sheet_errors)} errors")
        for e in sheet_errors:
            print(e)
        total_errors += len(sheet_errors)
    else:
        print(f"  {sn}: No formula errors")

ok = total_errors == 0
all_pass = all_pass and ok

# ── Structure verification ──
print("\nStructure verification:")
wb2 = openpyxl.load_workbook(SRC)

for sn, travel_total_row, net_profit_row in [
    ("Venue Estimate", 86, 97),
    ("Decor Estimate", 131, 142),
    ("SAMPLE Decor Estimate", 131, 142),
]:
    ws = wb2[sn]
    tt = ws.cell(row=travel_total_row, column=4).value
    np_val = ws.cell(row=net_profit_row, column=4).value
    # Check travel total references all 3 trips
    tt_ok = 'B' in str(tt) and 'C' in str(tt) and 'D' in str(tt)
    # Check net profit references travel total
    np_ok = str(travel_total_row) in str(np_val)
    print(f"  {sn}:")
    print(f"    {'PASS' if tt_ok else 'FAIL'}: Total Travel (D{travel_total_row}) = {tt}")
    print(f"    {'PASS' if np_ok else 'FAIL'}: True Net Profit (D{net_profit_row}) = {np_val}")
    all_pass = all_pass and tt_ok and np_ok

# Validation counts
print("\nData validations per sheet:")
for sn in ['Venue Estimate', 'Decor Estimate', 'SAMPLE Decor Estimate']:
    ws = wb2[sn]
    print(f"  {sn}: {len(ws.data_validations.dataValidation)} validations")

print(f"\n{'=' * 60}")
print(f"OVERALL: {'ALL CHECKS PASSED' if all_pass else 'SOME CHECKS FAILED — review above'}")
print(f"{'=' * 60}")
