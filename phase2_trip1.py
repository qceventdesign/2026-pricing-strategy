#!/usr/bin/env python3
"""
Phase 2: Build Trip 1 structure on Venue Estimate tab.

Steps:
0. Fix Client Setup lookup names (drive routes, hotel market, flight types)
1. Insert 6 rows at row 82 to expand travel section (64-87)
2. Clear & rewrite travel section with 13 input fields + 5 calc fields
3. Add data validations for all dropdowns
4. Fix P&M True Net Profit reference
5. Simulate Test 1 & Test 2
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

SRC = '/Users/alex/2026-pricing-strategy/2026-pricing-strategy/docs/templates/QC_Estimate_Template_2026_v2.xlsx'
wb = openpyxl.load_workbook(SRC)
cs = wb['Client Setup']
ve = wb['Venue Estimate']

# ── Styles ──
FILL_INPUT   = PatternFill('solid', fgColor='FFFFF8F0')
FILL_CALC    = PatternFill('solid', fgColor='FFF5F0EB')
FILL_ROW_BG  = PatternFill('solid', fgColor='FFFAF6F3')
FILL_HEADER  = PatternFill('solid', fgColor='FFECDFCE')
FILL_SUBHEAD = PatternFill('solid', fgColor='FFE8D9C8')
FILL_NONE    = PatternFill()

FONT_SECTION = Font(name='Playfair Display', size=11, bold=True, color='FF846E60')
FONT_COLHEAD = Font(name='Playfair Display', size=10, bold=True, color='FFC19C81')
FONT_LABEL   = Font(name='Playfair Display', size=10, color='FF464543')
FONT_INPUT   = Font(name='Playfair Display', size=10, bold=True, color='FF846E60')
FONT_CALC    = Font(name='Playfair Display', size=10, color='FF846E60')
FONT_BOLD    = Font(name='Playfair Display', size=10, bold=True, color='FF464543')
FONT_BOLD_CALC = Font(name='Playfair Display', size=10, bold=True, color='FF846E60')

AC = Alignment(horizontal='center', vertical='center')
AL = Alignment(horizontal='left', vertical='center')
THIN_BORDER = Border(bottom=Side(style='thin', color='FFD4C5B0'))


# ════════════════════════════════════════════════════════════════
# STEP 0: Fix Client Setup lookup names
# ════════════════════════════════════════════════════════════════
print("Step 0: Fixing Client Setup lookup names...")

# Drive routes: match destination dropdown values
fixes = {
    58: "Charlotte to Raleigh NC",      # was "Charlotte to Raleigh"
    59: "Charlotte to Asheville NC",     # was "Charlotte to Asheville"
    60: "Charlotte to Charleston SC",    # was "Charlotte to Charleston"
    62: "DC to Virginia Secondary",      # was "DC to Virginia Secondary Markets"
}
for row, val in fixes.items():
    cs.cell(row=row, column=1).value = val

# Flight types: shorten for dropdown compatibility
cs.cell(row=75, column=1).value = "Short Haul"
cs.cell(row=76, column=1).value = "Medium Haul"
# Row 77 "Major Market to NYC" stays as-is

# Hotel market: match destination dropdown
cs.cell(row=90, column=1).value = "Virginia Secondary"  # was "Virginia Secondary Markets"

print("  Fixed 4 drive routes, 2 flight type labels, 1 hotel market name")


# ════════════════════════════════════════════════════════════════
# STEP 1: Insert 6 rows at row 82 on Venue Estimate
# ════════════════════════════════════════════════════════════════
print("Step 1: Inserting 6 rows at row 82...")
ve.insert_rows(82, 6)
# Travel section: rows 64-87 (24 rows)
# P&M: rows 88-99 (shifted from 82-93)


# ════════════════════════════════════════════════════════════════
# STEP 2: Clear travel section rows 64-87
# ════════════════════════════════════════════════════════════════
print("Step 2: Clearing rows 64-87...")
for row in range(64, 88):
    for col in range(1, 6):
        cell = ve.cell(row=row, column=col)
        cell.value = None
        cell.fill = FILL_NONE
        cell.font = Font()
        cell.alignment = Alignment()
        cell.border = Border()
        cell.number_format = 'General'


# ════════════════════════════════════════════════════════════════
# STEP 3: Remove old travel data validations
# ════════════════════════════════════════════════════════════════
print("Step 3: Removing old travel validations...")
old_travel_rows = set(range(68, 73))
to_remove = []
for dv in ve.data_validations.dataValidation:
    sqref = str(dv.sqref)
    if any(f'{c}{r}' in sqref for r in old_travel_rows for c in 'BCD'):
        to_remove.append(dv)
for dv in to_remove:
    ve.data_validations.dataValidation.remove(dv)
print(f"  Removed {len(to_remove)} old validations")


# ════════════════════════════════════════════════════════════════
# STEP 4: Write new travel section
# ════════════════════════════════════════════════════════════════
print("Step 4: Writing travel section...")

# Row references
R_HDR = 65
R_LABEL = 66
R_ORIGIN = 67
R_DEST = 68
R_TTYPE = 69
R_FTYPE = 70
R_LASTMIN = 71
R_STAFF = 72
R_NIGHTS = 73
R_HOTELBUD = 74
R_VTYPE = 75
R_VSERVICE = 76
R_VHOURS = 77
R_VCUSTOM = 78
R_TRAVEL = 80
R_HOTEL = 81
R_PERDIEM = 82
R_VEHICLE = 83
R_TRIPTOTAL = 84
R_TOTAL = 86

# ── Header row ──
ve.cell(row=R_HDR, column=1).value = "QC TRAVEL EXPENSES"
for c in range(1, 5):
    ve.cell(row=R_HDR, column=c).fill = FILL_HEADER
    ve.cell(row=R_HDR, column=c).font = FONT_SECTION
for i, label in enumerate(["Trip 1", "Trip 2", "Trip 3"]):
    cell = ve.cell(row=R_HDR, column=2 + i)
    cell.value = label
    cell.alignment = AC

# ── Input rows ──
input_fields = [
    (R_LABEL,    "Trip Label"),
    (R_ORIGIN,   "Origin"),
    (R_DEST,     "Destination"),
    (R_TTYPE,    "Travel Type"),
    (R_FTYPE,    "Flight Type"),
    (R_LASTMIN,  "Last Minute Buffer"),
    (R_STAFF,    "Staff Traveling"),
    (R_NIGHTS,   "Nights"),
    (R_HOTELBUD, "Hotel Budget"),
    (R_VTYPE,    "Vehicle Type"),
    (R_VSERVICE, "Vehicle Service"),
    (R_VHOURS,   "Vehicle Hours"),
    (R_VCUSTOM,  "Custom Vehicle Cost"),
]

for row, label in input_fields:
    ve.cell(row=row, column=1).value = label
    ve.cell(row=row, column=1).font = FONT_LABEL
    ve.cell(row=row, column=1).fill = FILL_ROW_BG
    ve.cell(row=row, column=1).alignment = AL
    # Style input cells B, C, D
    for col in range(2, 5):
        cell = ve.cell(row=row, column=col)
        cell.fill = FILL_INPUT
        cell.font = FONT_INPUT
        cell.alignment = AC
    # Custom Vehicle Cost gets dollar format
    if row == R_VCUSTOM:
        for col in range(2, 5):
            ve.cell(row=row, column=col).number_format = '$#,##0'

# ── Calculated rows ──
calc_fields = [
    (R_TRAVEL,    "Travel Cost",    False),
    (R_HOTEL,     "Hotel Cost",     False),
    (R_PERDIEM,   "Per Diem Cost",  False),
    (R_VEHICLE,   "Vehicle Cost",   False),
    (R_TRIPTOTAL, "Trip Total",     True),
]

for row, label, is_total in calc_fields:
    ve.cell(row=row, column=1).value = label
    ve.cell(row=row, column=1).font = FONT_BOLD if is_total else FONT_LABEL
    ve.cell(row=row, column=1).fill = FILL_ROW_BG
    ve.cell(row=row, column=1).alignment = AL
    for col in range(2, 5):
        cell = ve.cell(row=row, column=col)
        cell.fill = FILL_CALC
        cell.font = FONT_BOLD_CALC if is_total else FONT_CALC
        cell.alignment = AC
        cell.number_format = '$#,##0'
    if is_total:
        for col in range(1, 5):
            ve.cell(row=row, column=col).border = THIN_BORDER

# ── Total Travel Expenses row ──
ve.cell(row=R_TOTAL, column=1).value = "Total Travel Expenses"
ve.cell(row=R_TOTAL, column=1).font = FONT_BOLD
ve.cell(row=R_TOTAL, column=1).fill = FILL_SUBHEAD
ve.cell(row=R_TOTAL, column=1).alignment = AL
for c in [2, 3]:
    ve.cell(row=R_TOTAL, column=c).fill = FILL_SUBHEAD
ve.cell(row=R_TOTAL, column=4).fill = FILL_SUBHEAD
ve.cell(row=R_TOTAL, column=4).font = FONT_BOLD_CALC
ve.cell(row=R_TOTAL, column=4).alignment = AC
ve.cell(row=R_TOTAL, column=4).number_format = '$#,##0'


# ════════════════════════════════════════════════════════════════
# STEP 5: Write Trip 1 formulas (column B)
# ════════════════════════════════════════════════════════════════
print("Step 5: Writing Trip 1 formulas...")

CL = 'B'  # Trip 1 column

dest = f'{CL}{R_DEST}'
origin = f'{CL}{R_ORIGIN}'
ttype = f'{CL}{R_TTYPE}'
ftype = f'{CL}{R_FTYPE}'
lastmin = f'{CL}{R_LASTMIN}'
staff = f'{CL}{R_STAFF}'
nights = f'{CL}{R_NIGHTS}'
hotelbud = f'{CL}{R_HOTELBUD}'
vtype = f'{CL}{R_VTYPE}'
vservice = f'{CL}{R_VSERVICE}'
vhours = f'{CL}{R_VHOURS}'
vcustom = f'{CL}{R_VCUSTOM}'

# Travel Cost
ve.cell(row=R_TRAVEL, column=2).value = (
    f'=IF(OR({dest}="",{staff}=""),0,'
    f'IF({ttype}="Drive",'
    f'IFERROR(VLOOKUP({origin}&" to "&{dest},DriveRoutes,2,FALSE),0),'
    f'IF({ttype}="Train",'
    f'IFERROR(VLOOKUP({origin}&" to "&{dest},TrainRoutes,2,FALSE),0)*{staff},'
    f'IF({ttype}="Flight",'
    f'IFERROR(VLOOKUP({ftype},FlightTypes,2,FALSE),0)*{staff}'
    f'+IF({lastmin}="Yes",150*{staff},0),'
    f'0))))'
)

# Hotel Cost
ve.cell(row=R_HOTEL, column=2).value = (
    f'=IF(OR({dest}="",{staff}="",{nights}=""),0,'
    f'{nights}*IFERROR(IF({hotelbud}="High",'
    f'VLOOKUP({dest},HotelRates,3,FALSE),'
    f'VLOOKUP({dest},HotelRates,2,FALSE)),0)*{staff})'
)

# Per Diem Cost
# Standard: Full=CS!B94, Half=CS!C94 | NYC: Full=CS!B95, Half=CS!C95
ve.cell(row=R_PERDIEM, column=2).value = (
    f'=IF(OR({dest}="",{staff}=""),0,'
    f'IF({dest}="NYC",'
    f'IF({nights}=0,\'Client Setup\'!C95,'
    f'IF({nights}=1,\'Client Setup\'!C95+\'Client Setup\'!B95,'
    f'2*\'Client Setup\'!C95+({nights}-1)*\'Client Setup\'!B95)),'
    f'IF({nights}=0,\'Client Setup\'!C94,'
    f'IF({nights}=1,\'Client Setup\'!C94+\'Client Setup\'!B94,'
    f'2*\'Client Setup\'!C94+({nights}-1)*\'Client Setup\'!B94)))'
    f'*{staff})'
)

# Vehicle Cost
# VehicleRates cols: 1=Market,2=SedanHr,3=SedanAirport,4=SUVHr,5=SUVAirport,6=SprinterHr,7=SprinterAirport
ve.cell(row=R_VEHICLE, column=2).value = (
    f'=IF(AND(ISNUMBER({vcustom}),{vcustom}>0),{vcustom},'
    f'IF(OR({vtype}="",{vtype}="None"),0,'
    f'IF({vservice}="Airport Transfer",'
    f'IFERROR(VLOOKUP({dest},VehicleRates,'
    f'IF({vtype}="Sedan",3,IF({vtype}="SUV",5,7)),FALSE),0),'
    f'IF({vservice}="Hourly",'
    f'IFERROR(VLOOKUP({dest},VehicleRates,'
    f'IF({vtype}="Sedan",2,IF({vtype}="SUV",4,6)),FALSE),0)*{vhours},'
    f'0))))'
)

# Trip Total
ve.cell(row=R_TRIPTOTAL, column=2).value = (
    f'={CL}{R_TRAVEL}+{CL}{R_HOTEL}+{CL}{R_PERDIEM}+{CL}{R_VEHICLE}'
)

# Total Travel Expenses (Phase 2: Trip 1 only)
ve.cell(row=R_TOTAL, column=4).value = f'=B{R_TRIPTOTAL}'


# ════════════════════════════════════════════════════════════════
# STEP 6: Add data validations (all 3 trip columns)
# ════════════════════════════════════════════════════════════════
print("Step 6: Adding data validations...")

validations = [
    (R_ORIGIN,   "OriginList",    True),
    (R_DEST,     "DestinationList", True),
    (R_TTYPE,    '"Drive,Train,Flight"', False),
    (R_FTYPE,    '"Short Haul,Medium Haul,Major Market to NYC"', False),
    (R_LASTMIN,  '"No,Yes"', False),
    (R_STAFF,    '"1,2,3,4,5"', False),
    (R_NIGHTS,   '"0,1,2,3,4,5"', False),
    (R_HOTELBUD, '"Low,High"', False),
    (R_VTYPE,    '"None,Sedan,SUV,Sprinter"', False),
    (R_VSERVICE, '"Hourly,Airport Transfer"', False),
    (R_VHOURS,   '"1,2,3,4,5,6,7,8,10,12"', False),
]

for row, formula, is_named in validations:
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.sqref = f"B{row} C{row} D{row}"
    ve.add_data_validation(dv)

print(f"  Added {len(validations)} data validations")


# ════════════════════════════════════════════════════════════════
# STEP 7: Fix P&M True Net Profit formula
# ════════════════════════════════════════════════════════════════
print("Step 7: Fixing P&M reference...")

# P&M is now at rows 88-99 (shifted from 82-93)
# Row 97 = True Net Profit (was row 91)
# Old formula: =D86-D90-D81 → shifted to =D92-D96-D81
# Need: =D92-D96-D86 (D86 = new Total Travel Expenses row)
R_PM_QC_REV = 92
R_PM_OPEX = 96
R_PM_NET = 97

old_formula = ve.cell(row=R_PM_NET, column=4).value
ve.cell(row=R_PM_NET, column=4).value = f'=D{R_PM_QC_REV}-D{R_PM_OPEX}-D{R_TOTAL}'
print(f"  Row {R_PM_NET}: was '{old_formula}' → '=D{R_PM_QC_REV}-D{R_PM_OPEX}-D{R_TOTAL}'")


# ════════════════════════════════════════════════════════════════
# STEP 8: Save
# ════════════════════════════════════════════════════════════════
print("\nSaving...")
wb.save(SRC)
print(f"Saved to {SRC}")


# ════════════════════════════════════════════════════════════════
# STEP 9: Test simulation
# ════════════════════════════════════════════════════════════════
print("\n" + "=" * 60)
print("TEST SIMULATION")
print("=" * 60)

# Rate table data (from Client Setup)
DRIVE = {
    "DC to Richmond VA": 205, "Charlotte to Raleigh NC": 230,
    "Charlotte to Asheville NC": 190, "Charlotte to Charleston SC": 300,
    "Charlotte to Atlanta": 360, "DC to Virginia Secondary": 215,
    "DC to Philadelphia": 190, "DC to NYC": 325, "Charlotte to DC": 270,
}
TRAIN = {"DC to NYC": (180, 300), "Philadelphia to NYC": (120, 220), "DC to Philadelphia": (120, 200)}
FLIGHT = {"Short Haul": (300, 450), "Medium Haul": (400, 600), "Major Market to NYC": (450, 650)}
HOTEL = {
    "NYC": (450, 650), "DC": (350, 550), "Charleston SC": (350, 550),
    "Atlanta": (275, 450), "Charlotte": (250, 400), "Asheville NC": (300, 500),
    "Raleigh NC": (250, 400), "Philadelphia": (350, 550), "Virginia Secondary": (200, 350),
}
PERDIEM = {"Standard": (68, 34), "NYC": (92, 46)}  # (full, half)
VEHICLE = {
    "NYC":      {"Sedan": (85, 250), "SUV": (125, 350), "Sprinter": (175, 500)},
    "DC":       {"Sedan": (75, 200), "SUV": (100, 275), "Sprinter": (150, 400)},
    "Philadelphia": {"Sedan": (70, 175), "SUV": (95, 250), "Sprinter": (140, 375)},
    "Charlotte":    {"Sedan": (60, 150), "SUV": (85, 200), "Sprinter": (125, 325)},
    "Raleigh NC":   {"Sedan": (60, 150), "SUV": (85, 200), "Sprinter": (125, 325)},
    "Asheville NC": {"Sedan": (65, 175), "SUV": (90, 225), "Sprinter": (130, 350)},
    "Charleston SC":{"Sedan": (60, 150), "SUV": (85, 200), "Sprinter": (125, 325)},
    "Atlanta":      {"Sedan": (65, 175), "SUV": (95, 250), "Sprinter": (140, 375)},
    "Virginia Secondary": {"Sedan": (55, 125), "SUV": (75, 175), "Sprinter": (110, 300)},
    "Other":        {"Sedan": (65, 175), "SUV": (90, 225), "Sprinter": (135, 350)},
}


def calc_travel(origin, dest, ttype, ftype, lastmin, staff):
    if ttype == "Drive":
        route = f"{origin} to {dest}"
        return DRIVE.get(route, 0)
    elif ttype == "Train":
        route = f"{origin} to {dest}"
        return TRAIN.get(route, (0, 0))[0] * staff  # Low default
    elif ttype == "Flight":
        base = FLIGHT.get(ftype, (0, 0))[0] * staff  # Low default
        buffer = 150 * staff if lastmin == "Yes" else 0
        return base + buffer
    return 0


def calc_hotel(dest, nights, budget, staff):
    if dest not in HOTEL or nights == 0:
        return 0
    idx = 1 if budget == "High" else 0
    return nights * HOTEL[dest][idx] * staff


def calc_perdiem(dest, nights, staff):
    rates = PERDIEM["NYC"] if dest == "NYC" else PERDIEM["Standard"]
    full, half = rates
    if nights == 0:
        return half * staff
    elif nights == 1:
        return (half + full) * staff
    else:
        return (2 * half + (nights - 1) * full) * staff


def calc_vehicle(dest, vtype, vservice, vhours, vcustom):
    if vcustom and vcustom > 0:
        return vcustom
    if not vtype or vtype == "None":
        return 0
    if dest not in VEHICLE or vtype not in VEHICLE.get(dest, {}):
        return 0
    hourly, airport = VEHICLE[dest][vtype]
    if vservice == "Airport Transfer":
        return airport
    elif vservice == "Hourly":
        return hourly * (vhours or 0)
    return 0


def run_test(name, origin, dest, ttype, ftype, lastmin, staff, nights,
             budget, vtype, vservice, vhours, vcustom,
             exp_travel, exp_hotel, exp_perdiem, exp_vehicle, exp_total):
    travel = calc_travel(origin, dest, ttype, ftype, lastmin, staff)
    hotel = calc_hotel(dest, nights, budget, staff)
    perdiem = calc_perdiem(dest, nights, staff)
    vehicle = calc_vehicle(dest, vtype, vservice, vhours, vcustom)
    total = travel + hotel + perdiem + vehicle

    results = [
        ("Travel",  travel,  exp_travel),
        ("Hotel",   hotel,   exp_hotel),
        ("Per Diem", perdiem, exp_perdiem),
        ("Vehicle", vehicle, exp_vehicle),
        ("Total",   total,   exp_total),
    ]

    all_pass = True
    print(f"\n{name}:")
    for label, actual, expected in results:
        ok = actual == expected
        if not ok:
            all_pass = False
        status = "PASS" if ok else "FAIL"
        print(f"  {status}: {label} = ${actual:,} (expected ${expected:,})")
    return all_pass


t1 = run_test(
    "Test 1 — Train DC to NYC, 1 night",
    origin="DC", dest="NYC", ttype="Train", ftype="", lastmin="No",
    staff=2, nights=1, budget="High",
    vtype="SUV", vservice="Airport Transfer", vhours=0, vcustom=None,
    exp_travel=360, exp_hotel=1300, exp_perdiem=276, exp_vehicle=350, exp_total=2286
)

t2 = run_test(
    "Test 2 — Drive Charlotte to Raleigh, day trip",
    origin="Charlotte", dest="Raleigh NC", ttype="Drive", ftype="", lastmin="No",
    staff=1, nights=0, budget="Low",
    vtype="Sedan", vservice="Hourly", vhours=3, vcustom=None,
    exp_travel=230, exp_hotel=0, exp_perdiem=34, exp_vehicle=180, exp_total=444
)

# ── Formula verification ──
print("\n" + "-" * 40)
print("Formula cell verification:")
wb2 = openpyxl.load_workbook(SRC)
ve2 = wb2['Venue Estimate']
for row, label in [(R_TRAVEL, "Travel"), (R_HOTEL, "Hotel"),
                    (R_PERDIEM, "Per Diem"), (R_VEHICLE, "Vehicle"),
                    (R_TRIPTOTAL, "Trip Total"), (R_TOTAL, "Total Travel")]:
    col = 2 if row != R_TOTAL else 4
    val = ve2.cell(row=row, column=col).value
    print(f"  Row {row} ({label}): {val[:80]}..." if len(str(val)) > 80 else f"  Row {row} ({label}): {val}")

# P&M check
pm_net = ve2.cell(row=R_PM_NET, column=4).value
print(f"  Row {R_PM_NET} (True Net Profit): {pm_net}")

print(f"\n{'=' * 60}")
print(f"OVERALL: {'ALL TESTS PASSED' if (t1 and t2) else 'FAILURES — review above'}")
print(f"{'=' * 60}")
