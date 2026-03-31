"""
Fix, protect, validate, and save the QC Estimate Template.
Steps: 1) Formula Fixes  2) Protection  3) Auto-Validation
Source: QC_Estimate_Template_2026_v2.xlsx (has travel section matching cell refs)
Output: QC_Estimate_Template_2026_v2_FINAL.xlsx
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Protection, numbers
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from copy import copy

SRC = "2026-pricing-strategy/docs/templates/QC_Estimate_Template_2026_v2.xlsx"
DST = "2026-pricing-strategy/docs/templates/QC_Estimate_Template_2026_v2_FINAL.xlsx"

wb = openpyxl.load_workbook(SRC)
cs = wb["Client Setup"]
venue = wb["Venue Estimate"]
decor = wb["Decor Estimate"]
sample = wb["SAMPLE Decor Estimate"]

# ═══════════════════════════════════════════════════════════════════════════
# STEP 1: FORMULA FIXES
# ═══════════════════════════════════════════════════════════════════════════

print("Step 1: Formula Fixes")

# --- Fix A: Service Charge / Gratuity / Admin Fee (Venue D51:E53) ---
print("  Fix A: Service Charge / Gratuity / Admin Fee")
venue["D51"] = '=D43*IF(OR(C17="None",C17="",C17=0),0,IF(ISNUMBER(C17),C17,IFERROR(VALUE(SUBSTITUTE(C17,"%",""))/100,0)))'
venue["E51"] = '=E43*IF(OR(C17="None",C17="",C17=0),0,IF(ISNUMBER(C17),C17,IFERROR(VALUE(SUBSTITUTE(C17,"%",""))/100,0)))'
venue["D52"] = '=D43*IF(OR(C18="None",C18="",C18=0),0,IF(ISNUMBER(C18),C18,IFERROR(VALUE(SUBSTITUTE(C18,"%",""))/100,0)))'
venue["E52"] = '=E43*IF(OR(C18="None",C18="",C18=0),0,IF(ISNUMBER(C18),C18,IFERROR(VALUE(SUBSTITUTE(C18,"%",""))/100,0)))'
venue["D53"] = '=D43*IF(OR(C19="None",C19="",C19=0),0,IF(ISNUMBER(C19),C19,IFERROR(VALUE(SUBSTITUTE(C19,"%",""))/100,0)))'
venue["E53"] = '=E43*IF(OR(C19="None",C19="",C19=0),0,IF(ISNUMBER(C19),C19,IFERROR(VALUE(SUBSTITUTE(C19,"%",""))/100,0)))'

# --- Fix B: F&B Minimum Check (Venue B62) — compare D43 not E43 ---
print("  Fix B: F&B Minimum Check")
venue["B62"] = '=IF(B7=0,"No F&B Minimum Set",IF(D43>=B7,"✓ F&B MINIMUM MET","✗ NOT MET — Shortfall: "&TEXT(B7-D43,"$#,##0")))'

# --- Fix C: Decor Non-Taxable Rental Fee Categories (rows 70-74) ---
print("  Fix C: Decor I70:I74 category fix")
for sheet in [decor, sample]:
    for row in range(70, 75):
        sheet[f"I{row}"] = "Delivery & Logistics"

# --- Fix D: GDP Commission Rate ---
print("  Fix D: GDP Commission Rate")
# Client Setup: make GDP rate editable
cs["A37"] = "GDP Commission:"
cs["B37"] = "Yes"  # Toggle stays (user test data also sets B37=Yes)
cs["C37"] = 0.065
cs["C37"].number_format = "0.0%"
cs["D37"] = "Toggle Yes/No; rate in C37"

# Venue Estimate D84 (GDP row): reference editable rate from Client Setup
# D84 is the actual GDP Commission row in v2
venue["D84"] = '=IF(\'Client Setup\'!B37="Yes",(E43+E46+E48+E49+E51+E52+E53)*\'Client Setup\'!C37,0)'

# --- Fix E: Equipment & Staffing Tax (Venue D47:E47) ---
print("  Fix E: Equipment & Staffing Tax row-by-row")
venue["D47"] = "=IFERROR(D28*F28+D29*F29+D30*F30+D31*F31,0)"
venue["E47"] = "=IFERROR(E28*F28+E29*F29+E30*F30+E31*F31,0)"

# --- Fix F: Dropdown Validations (Venue C17:C19) ---
print("  Fix F: Dropdown Validations for fee overrides")
# Remove any existing data validations on C14:C16 (old rows) or C17:C19
# We'll add new ones
dv_svc = DataValidation(type="list", formula1='"20%,21.5%,None"', allow_blank=True)
dv_svc.sqref = "C17"
dv_grat = DataValidation(type="list", formula1='"20%,None"', allow_blank=True)
dv_grat.sqref = "C18"
dv_admin = DataValidation(type="list", formula1='"5%,None"', allow_blank=True)
dv_admin.sqref = "C19"

# Remove old validations that conflict
old_dvs = list(venue.data_validations.dataValidation)
for dv in old_dvs:
    sqref_str = str(dv.sqref)
    if sqref_str in ("C14", "C15", "C16", "C17", "C18", "C19"):
        venue.data_validations.dataValidation.remove(dv)

venue.add_data_validation(dv_svc)
venue.add_data_validation(dv_grat)
venue.add_data_validation(dv_admin)

# --- Fix G: Cell Comments on Production Fee ---
print("  Fix G: Production Fee comments")
prod_fee_comment_venue = Comment(
    "Mirrors E55 intentionally. Commission (payout) and CC fee are netted out "
    "separately in the profit analysis below.",
    "Quill Creative"
)
venue["D55"].comment = prod_fee_comment_venue

# AV section is embedded in Decor sheets (E82 area)
# In the Decor sheets, check what's at E82 and E53
# Decor E82 = production fee area in AV section
for sheet in [decor, sample]:
    av_comment = Comment(
        "Mirrors F53 intentionally. Production fee is calculated on client-facing totals.",
        "Quill Creative"
    )
    sheet["E82"].comment = av_comment

# --- Fix H: Populate Test Data ---
print("  Fix H: Populating test data")

# Client Setup test data
cs["B5"] = "TEST CLIENT"
cs["B6"] = "2026-06-15"
cs["B7"] = 100
cs["B8"] = "Plated"
cs["B9"] = "Full Bar"
cs["B10"] = "6-10"
cs["B11"] = "Acme Corp"
cs["B12"] = "Summer Gala 2026"
cs["B13"] = "The Watergate Hotel"
cs["B16"] = "DC"
cs["B35"] = 0.035
cs["B35"].number_format = "0.0%"
cs["B36"] = "5%"
# B37 already set above to "Yes"
cs["B40"] = "20%"
cs["B41"] = "20%"
cs["B42"] = "5%"

# Venue Estimate test data
venue["B5"] = "TEST RESTAURANT"
venue["B6"] = "Main Dining Room"
venue["B7"] = 15000
venue["B7"].number_format = '#,##0'
venue["B8"] = "Yes"

# F&B items (guest count comes from Client Setup B7 = 100)
venue["C24"] = 100
venue["C25"] = 30
venue["C26"] = 10

# Taxable Equipment & Staffing
venue["B28"] = 2
venue["C28"] = 300
venue["B29"] = 1
venue["C29"] = 500
venue["B30"] = 1
venue["C30"] = 400
venue["B31"] = 1
venue["C31"] = 200

# Venue Fees
venue["B33"] = 1
venue["C33"] = 2000
venue["B34"] = 1
venue["C34"] = 500
venue["B35"] = 0
venue["C35"] = 0

# QC Staffing
venue["B38"] = 2
venue["C38"] = 500
venue["B39"] = 1
venue["C39"] = 250
venue["B40"] = 1
venue["C40"] = 150

# Travel data — mapped to v2 travel section structure
# v2: Row 66=headers, 67=Trip Label, 68=Destination, 69=Travel Type(formula),
#     70=Staff, 71=Nights, 72=Vehicle, 73=Custom Cost
venue["B66"] = "Event Day"
venue["C66"] = "Site Visit"
venue["B67"] = "DC"
venue["C67"] = "Charlotte"
venue["B68"] = "DC"        # Destination (from DestinationList)
venue["C68"] = "Charlotte"  # Destination (from DestinationList)
# Row 69 = Travel Type formula (auto-populated from destination)
venue["B70"] = 2   # Staff Traveling
venue["C70"] = 1
venue["B71"] = 1   # Nights
venue["C71"] = 0
venue["B72"] = "None"  # On-Site Vehicle
venue["C72"] = "None"

# Decor Estimate test data
decor["B9"] = "TEST RESTAURANT — Main Dining Room"

# Florals (Taxable)
decor["A16"] = "Table Centerpieces"
decor["B16"] = "Test Florist"
decor["C16"] = 10
decor["D16"] = 75
decor["A17"] = "Cocktail Arrangements"
decor["B17"] = "Test Florist"
decor["C17"] = 5
decor["D17"] = 50
decor["A18"] = "Bar Garland"
decor["B18"] = "Test Florist"
decor["C18"] = 1
decor["D18"] = 200

# Non-Taxable Floral Fees
decor["A27"] = "Floral Delivery & Setup"
decor["B27"] = "Test Florist"
decor["D27"] = 300
decor["A28"] = "Floral Strike"
decor["B28"] = "Test Florist"
decor["D28"] = 100

# Seating
decor["A35"] = "Gold Chiavari Chair"
decor["B35"] = "Test Rentals"
decor["C35"] = 100
decor["D35"] = 12
decor["A36"] = "Chair Cushion"
decor["B36"] = "Test Rentals"
decor["C36"] = 100
decor["D36"] = 3

# Lounge
decor["A44"] = "Club Chair"
decor["B44"] = "Test Rentals"
decor["C44"] = 4
decor["D44"] = 150
decor["A45"] = "Sectional Sofa"
decor["B45"] = "Test Rentals"
decor["C45"] = 2
decor["D45"] = 350

# Tables
decor["A53"] = "Side Table"
decor["B53"] = "Test Rentals"
decor["C53"] = 4
decor["D53"] = 65
decor["A54"] = "Coffee Table"
decor["B54"] = "Test Rentals"
decor["C54"] = 2
decor["D54"] = 120

# Decor & Accessories
decor["A62"] = "Votives & Candles"
decor["B62"] = "Amazon"
decor["C62"] = 20
decor["D62"] = 8

# Non-Taxable Rental Fees
decor["A69"] = "Rental Delivery & Setup"
decor["B69"] = "Test Rentals"
decor["D69"] = 450
decor["A70"] = "Rental Strike"
decor["B70"] = "Test Rentals"
decor["D70"] = 200


# ═══════════════════════════════════════════════════════════════════════════
# STEP 2: FORMULA PROTECTION
# ═══════════════════════════════════════════════════════════════════════════

print("\nStep 2: Formula Protection")

LOCKED = Protection(locked=True)
UNLOCKED = Protection(locked=False)


def lock_all_then_unlock(ws, unlock_ranges):
    """Lock every cell, then unlock specified input ranges."""
    # Lock all cells in the used range
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.protection = LOCKED

    # Unlock specified ranges
    for rng in unlock_ranges:
        result = ws[rng]
        # Single cell (no colon in range)
        if not isinstance(result, tuple):
            result.protection = UNLOCKED
        else:
            for row in result:
                if hasattr(row, '__iter__'):
                    for cell in row:
                        cell.protection = UNLOCKED
                else:
                    row.protection = UNLOCKED


def protect_sheet(ws):
    """Apply sheet protection with formatting and sort/filter allowed."""
    ws.protection.sheet = True
    ws.protection.set_password("")
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.insertRows = False
    ws.protection.sort = False
    ws.protection.autoFilter = False


# --- Client Setup ---
print("  Protecting Client Setup")
cs_unlock = [
    "B5:B13", "B16",
    "B23:B33",
    "B35", "B36", "B37", "C37",
    "B40:B42",
]
lock_all_then_unlock(cs, cs_unlock)
protect_sheet(cs)

# --- Venue Estimate ---
print("  Protecting Venue Estimate")
venue_unlock = [
    "B5:B6", "B7", "B8",
    "C17:C19",
    "C24:C26",
    "A28:C31", "G28:G31",
    "A33:C35", "G33:G35",
    "A38:C40", "G38:G40",
    "I28:I35",
    "B66:D78",
]
lock_all_then_unlock(venue, venue_unlock)
protect_sheet(venue)

# --- Decor Estimate ---
print("  Protecting Decor Estimate")
decor_unlock = [
    "B9",
    "A16:D25", "H16:H25", "I16:I25",
    "A27:A29", "B27:B29", "D27:D29", "H27:H29", "I27:I29",
    "A35:D42", "H35:H42", "I35:I42",
    "A44:D51", "H44:H51", "I44:I51",
    "A53:D60", "H53:H60", "I53:I60",
    "A62:D67", "H62:H67", "I62:I67",
    "A69:A74", "B69:B74", "D69:D74", "H69:H74", "I69:I74",
    "M29:M40",
]
lock_all_then_unlock(decor, decor_unlock)
protect_sheet(decor)

# --- SAMPLE Decor Estimate ---
print("  Protecting SAMPLE Decor Estimate")
lock_all_then_unlock(sample, decor_unlock)  # Same unlock ranges
protect_sheet(sample)


# ═══════════════════════════════════════════════════════════════════════════
# STEP 3: AUTO-VALIDATION (Conditional Formatting)
# ═══════════════════════════════════════════════════════════════════════════

print("\nStep 3: Auto-Validation (Conditional Formatting)")

RED_FILL = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFF0", end_color="FFFFF0", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFE8CC", end_color="FFE8CC", fill_type="solid")

# --- RED: Missing category ---
print("  Red: Missing category validations")

# Venue: G28:G31, G33:G35, G38:G40
for rng in ["G28:G31", "G33:G35", "G38:G40"]:
    venue.conditional_formatting.add(
        rng,
        FormulaRule(
            formula=[f'AND({rng.split(":")[0]}<>"",ISBLANK({rng.split(":")[0]}))'],
            fill=RED_FILL
        )
    )
    # Actually use a simpler approach: highlight if the cell in col G is blank
    # but the corresponding row has data (check col B or C)
    start_cell = rng.split(":")[0]
    row_start = int(start_cell[1:])
    row_end = int(rng.split(":")[1][1:])
    venue.conditional_formatting.add(
        rng,
        FormulaRule(
            formula=[f'AND(OR($B{row_start}<>"",$C{row_start}<>""),{start_cell}="")'],
            fill=RED_FILL
        )
    )

# For category columns, a simpler approach: blank cell = red when row has data
# Venue G columns
for ranges_pair in [("G28:G31", "B28"), ("G33:G35", "B33"), ("G38:G40", "B38")]:
    rng, ref_col = ranges_pair
    first_row = int(rng.split(":")[0][1:])
    venue.conditional_formatting.add(
        rng,
        FormulaRule(
            formula=[f'AND(OR($B{first_row}<>"",$C{first_row}<>""),{rng.split(":")[0]}="")'],
            fill=RED_FILL
        )
    )

# Decor + SAMPLE: I columns for category
decor_cat_ranges = [
    "I16:I25", "I35:I42", "I44:I51", "I53:I60", "I62:I67"
]
for sheet in [decor, sample]:
    for rng in decor_cat_ranges:
        first_cell = rng.split(":")[0]
        first_row = int(first_cell[1:])
        sheet.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f'AND(OR($A{first_row}<>"",$C{first_row}<>""),{first_cell}="")'],
                fill=RED_FILL
            )
        )

# --- YELLOW: Incomplete row ---
print("  Yellow: Incomplete row validations")

# Venue: B28:C31, B33:C35, B38:C40
venue_yellow_ranges = ["B28:C31", "B33:C35", "B38:C40"]
for rng in venue_yellow_ranges:
    first_cell = rng.split(":")[0]
    first_row = int(first_cell[1:])
    # Yellow if one of B/C is filled but the other is blank (incomplete row)
    venue.conditional_formatting.add(
        rng,
        FormulaRule(
            formula=[f'AND(OR($B{first_row}<>"",$C{first_row}<>""),OR($B{first_row}="",$C{first_row}=""))'],
            fill=YELLOW_FILL
        )
    )

# Decor + SAMPLE: C:D columns for qty/price
decor_yellow_ranges = [
    "C16:D25", "C35:D42", "C44:D51", "C53:D60", "C62:D67"
]
# Also D37:D46 equivalent for AV section
decor_yellow_av = ["D37:D46"]

for sheet in [decor, sample]:
    for rng in decor_yellow_ranges:
        first_cell = rng.split(":")[0]
        first_row = int(first_cell[1:])
        sheet.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f'AND(OR($C{first_row}<>"",$D{first_row}<>""),OR($C{first_row}="",$D{first_row}=""))'],
                fill=YELLOW_FILL
            )
        )

# --- ORANGE: Fee mismatch (Venue only) ---
print("  Orange: Fee mismatch validations")

# C17: not matching "20%","21.5%","None",0.2,0.215
venue.conditional_formatting.add(
    "C17",
    FormulaRule(
        formula=['AND(C17<>"",C17<>"20%",C17<>"21.5%",C17<>"None",C17<>0.2,C17<>0.215)'],
        fill=ORANGE_FILL
    )
)

# C18: not matching "20%","None",0.2
venue.conditional_formatting.add(
    "C18",
    FormulaRule(
        formula=['AND(C18<>"",C18<>"20%",C18<>"None",C18<>0.2)'],
        fill=ORANGE_FILL
    )
)

# C19: not matching "5%","None",0.05
venue.conditional_formatting.add(
    "C19",
    FormulaRule(
        formula=['AND(C19<>"",C19<>"5%",C19<>"None",C19<>0.05)'],
        fill=ORANGE_FILL
    )
)


# ═══════════════════════════════════════════════════════════════════════════
# FINAL VERIFICATION & SAVE
# ═══════════════════════════════════════════════════════════════════════════

print("\nFinal Verification:")

# Check formulas are in place
checks = {
    "Venue D51": venue["D51"].value,
    "Venue E51": venue["E51"].value,
    "Venue B62": venue["B62"].value,
    "Venue D47": venue["D47"].value,
    "Venue D84 (GDP)": venue["D84"].value,
    "CS A37": cs["A37"].value,
    "CS C37": cs["C37"].value,
    "Decor I70": decor["I70"].value,
    "Decor I74": decor["I74"].value,
    "SAMPLE I70": sample["I70"].value,
}

for label, val in checks.items():
    status = "✓" if val is not None else "✗ EMPTY"
    print(f"  {label}: {status} -> {repr(val)[:80]}")

# Check protection is active
for name, ws in [("Client Setup", cs), ("Venue", venue), ("Decor", decor), ("SAMPLE", sample)]:
    print(f"  {name} protection active: {ws.protection.sheet}")

# Verify test data in place
print(f"\n  Venue B5 (test restaurant): {venue['B5'].value}")
print(f"  Venue B7 (F&B min): {venue['B7'].value}")
print(f"  Venue C24 (per person food): {venue['C24'].value}")
print(f"  CS B7 (guest count): {cs['B7'].value}")
print(f"  Decor B9 (venue): {decor['B9'].value}")

print(f"\n  Expected after Excel recalc:")
print(f"    Venue D43 = SUM(D24:D26) = 100*100 + 100*30 + 100*10 = 14,000")
print(f"    Venue D51 = 14000 * 0.20 = 2,800  (Service Charge)")
print(f"    Venue D52 = 14000 * 0.20 = 2,800  (Gratuity)")
print(f"    Venue D53 = 14000 * 0.05 = 700    (Admin Fee)")
print(f"    Venue D47 = 600*0.06+500*0.06+400*0.06+200*0.06 = 102  (Equip Tax)")
print(f"    Venue B62 = '✗ NOT MET — Shortfall: $1,000'")

wb.save(DST)
print(f"\nSaved: {DST}")
print("Done! Open in Excel to verify calculated values.")
