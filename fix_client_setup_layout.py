"""
Fix Client Setup layout: restore column C helper text, fix formula references.
"""
import os
from openpyxl import load_workbook

os.chdir(os.path.join(os.path.dirname(__file__), '2026-pricing-strategy', 'docs', 'templates'))
wb = load_workbook('QC_Estimate_Template_2026_v2_FINAL.xlsx')

# ────────────────────────────────────────────────────────────────────
# 1. RESTORE CLIENT SETUP CELL VALUES on both tabs
# ────────────────────────────────────────────────────────────────────
for suffix in ['Sample', 'Blank']:
    ws = wb[f'Client Setup - {suffix}']

    # Row 35 — CC Processing Fee
    # B35: keep 0.035
    ws['C35'].value = 'Applied to subtotal'
    ws['C35'].number_format = 'General'
    ws['D35'].value = None

    # Row 36 — Commission
    ws['A36'].value = 'Client Commission %:'
    ws['B36'].value = 0.0
    ws['B36'].number_format = '0%'
    ws['C36'].value = 'Select commission rate (if applicable)'
    ws['C36'].number_format = 'General'
    ws['D36'].value = None

    # Row 37 — GDP
    ws['A37'].value = 'GDP Commission (6.5%):'
    # B37: keep "Yes"
    ws['C37'].value = 'Yes = 6.5% applied'
    ws['C37'].number_format = 'General'
    ws['D37'].value = None

    # Row 40 — Service Charge (text strings, not numbers)
    ws['B40'].value = '20%'
    ws['B40'].number_format = 'General'
    ws['C40'].value = '20%, 21.5%, or None'
    ws['C40'].number_format = 'General'
    ws['D40'].value = None

    # Row 41 — Gratuity
    ws['B41'].value = '20%'
    ws['B41'].number_format = 'General'
    ws['C41'].value = '20% or None'
    ws['C41'].number_format = 'General'
    ws['D41'].value = None

    # Row 42 — Admin Fee
    ws['B42'].value = '5%'
    ws['B42'].number_format = 'General'
    ws['C42'].value = '5% or None'
    ws['C42'].number_format = 'General'
    ws['D42'].value = None

    # Markup rates
    ws['B23'].value = 0.55

print('[1] Client Setup cell values restored on both tabs.')

# ────────────────────────────────────────────────────────────────────
# 2. FIX FORMULA REFERENCES across all sheets
# ────────────────────────────────────────────────────────────────────

# --- Venue Estimate sheets ---
for suffix, cs in [('Sample', 'Sample'), ('Blank', 'Blank')]:
    ws = wb[f'Venue Estimate - {suffix}']

    # E55: CC Fee — C35 → B35
    ws['E55'].value = f"=IFERROR(E54*'Client Setup - {cs}'!B35,0)"

    # E56: Commission — C36 → B36
    ws['E56'].value = f"=IFERROR((E43+E46+E48+E49+E51+E52+E53)*'Client Setup - {cs}'!B36,0)"

    # C83: Commission rate display — C36 → B36
    ws['C83'].value = f"='Client Setup - {cs}'!B36"

    # C84: GDP rate — hardcode 0.065 since C37 is no longer a rate cell
    ws['C84'].value = f"=IF('Client Setup - {cs}'!B37=\"Yes\",0.065,0)"

    # D84: GDP amount — hardcode 0.065
    ws['D84'].value = f"=IF('Client Setup - {cs}'!B37=\"Yes\",(E43+E46+E48+E49+E51+E52+E53)*0.065,0)"

# --- Decor Estimate sheets ---
for suffix, cs in [('Sample', 'Sample'), ('Blank', 'Blank')]:
    ws = wb[f'Decor Estimate - {suffix}']

    # F102: CC Fee — C35 → B35
    ws['F102'].value = f"=IFERROR((F99+F100+F101)*'Client Setup - {cs}'!B35,0)"

    # F103: Commission — C36 → B36
    ws['F103'].value = f"=IFERROR((F99+F100+F101)*'Client Setup - {cs}'!B36,0)"

    # C128: Commission rate display — C36 → B36
    ws['C128'].value = f"='Client Setup - {cs}'!B36"

    # C129: GDP rate — hardcode 0.065
    ws['C129'].value = f"=IF('Client Setup - {cs}'!B37=\"Yes\",0.065,0)"

    # D129: GDP amount — hardcode 0.065
    ws['D129'].value = f"=IF('Client Setup - {cs}'!B37=\"Yes\",F104*0.065,0)"

print('[2] Formula references updated (C35→B35, C36→B36, C37→0.065).')

# ────────────────────────────────────────────────────────────────────
# 3. SAVE & VERIFY
# ────────────────────────────────────────────────────────────────────
wb.save('QC_Estimate_Template_2026_v2_FINAL.xlsx')
print('\nSaved. Running verification...\n')

wb2 = load_workbook('QC_Estimate_Template_2026_v2_FINAL.xlsx')
ok = True

# Verify Client Setup helper text
for suffix in ['Sample', 'Blank']:
    ws = wb2[f'Client Setup - {suffix}']
    checks = {
        'C35': 'Applied to subtotal',
        'C36': 'Select commission rate (if applicable)',
        'C37': 'Yes = 6.5% applied',
        'C40': '20%, 21.5%, or None',
        'C41': '20% or None',
        'C42': '5% or None',
    }
    for ref, expected in checks.items():
        actual = ws[ref].value
        if actual != expected:
            print(f'  FAIL {suffix} {ref}: expected {expected!r}, got {actual!r}')
            ok = False

    # B40:B42 text strings
    for ref, expected in [('B40', '20%'), ('B41', '20%'), ('B42', '5%')]:
        actual = ws[ref].value
        if actual != expected or not isinstance(actual, str):
            print(f'  FAIL {suffix} {ref}: expected str {expected!r}, got {actual!r} (type={type(actual).__name__})')
            ok = False

    # D column cleared
    for ref in ['D35', 'D36', 'D37', 'D40', 'D41', 'D42']:
        if ws[ref].value is not None:
            print(f'  FAIL {suffix} {ref}: expected None, got {ws[ref].value!r}')
            ok = False

    # B23 markup
    if ws['B23'].value != 0.55:
        print(f'  FAIL {suffix} B23: expected 0.55, got {ws["B23"].value!r}')
        ok = False

# Verify formula references don't contain C35/C36 on fee rows
formula_checks = [
    ('Venue Estimate - Sample', 'E55'),
    ('Venue Estimate - Sample', 'E56'),
    ('Venue Estimate - Blank', 'E55'),
    ('Venue Estimate - Blank', 'E56'),
    ('Decor Estimate - Sample', 'F102'),
    ('Decor Estimate - Sample', 'F103'),
    ('Decor Estimate - Blank', 'F102'),
    ('Decor Estimate - Blank', 'F103'),
]
for sheet, ref in formula_checks:
    val = wb2[sheet][ref].value
    if 'C35' in str(val) or 'C36' in str(val):
        print(f'  FAIL {sheet}!{ref} still references C35/C36: {val}')
        ok = False

# Verify GDP formulas use hardcoded 0.065
gdp_checks = [
    ('Venue Estimate - Sample', 'C84'),
    ('Venue Estimate - Sample', 'D84'),
    ('Venue Estimate - Blank', 'C84'),
    ('Venue Estimate - Blank', 'D84'),
    ('Decor Estimate - Sample', 'C129'),
    ('Decor Estimate - Sample', 'D129'),
    ('Decor Estimate - Blank', 'C129'),
    ('Decor Estimate - Blank', 'D129'),
]
for sheet, ref in gdp_checks:
    val = wb2[sheet][ref].value
    if 'C37' in str(val):
        print(f'  FAIL {sheet}!{ref} still references C37: {val}')
        ok = False

# Verify no formula errors (check that formulas are still present)
for sheet, ref in formula_checks + gdp_checks:
    val = wb2[sheet][ref].value
    if not isinstance(val, str) or not val.startswith('='):
        print(f'  FAIL {sheet}!{ref}: not a formula: {val!r}')
        ok = False

if ok:
    print('All verifications passed.')
else:
    print('\nWARNING: Some checks failed!')
