"""
Restructure the FINAL workbook into 6 tabs.
Remove AV sheet (none exists). Create Sample/Blank pairs.
Update all cross-sheet references. Clear blank tab inputs.
"""

import openpyxl
from openpyxl.worksheet.protection import SheetProtection

SRC = "2026-pricing-strategy/docs/templates/QC_Estimate_Template_2026_v2_FINAL.xlsx"

wb = openpyxl.load_workbook(SRC)
print(f"Loaded: {wb.sheetnames}")

# ═══════════════════════════════════════════════════════════════════════════
# STEP 1: CREATE COPIES FOR BLANK VERSIONS (before renaming anything)
# ═══════════════════════════════════════════════════════════════════════════

print("\n1. Creating copies for blank tabs...")
cs_blank = wb.copy_worksheet(wb["Client Setup"])
venue_blank = wb.copy_worksheet(wb["Venue Estimate"])

print(f"   Sheets after copy: {wb.sheetnames}")

# ═══════════════════════════════════════════════════════════════════════════
# STEP 2: RENAME ALL SHEETS
# ═══════════════════════════════════════════════════════════════════════════

print("\n2. Renaming sheets...")

# Map: old name -> new name
renames = {
    "Client Setup": "Client Setup - Sample",
    "Venue Estimate": "Venue Estimate - Sample",
    "SAMPLE Decor Estimate": "Decor Estimate - Sample",
    "Decor Estimate": "Decor Estimate - Blank",
}

for old_name, new_name in renames.items():
    wb[old_name].title = new_name
    print(f"   '{old_name}' → '{new_name}'")

cs_blank.title = "Client Setup - Blank"
venue_blank.title = "Venue Estimate - Blank"
print(f"   Copy → 'Client Setup - Blank'")
print(f"   Copy → 'Venue Estimate - Blank'")

print(f"   Sheets now: {wb.sheetnames}")

# ═══════════════════════════════════════════════════════════════════════════
# STEP 3: REORDER TABS
# ═══════════════════════════════════════════════════════════════════════════

print("\n3. Reordering tabs...")
desired = [
    "Client Setup - Sample",
    "Client Setup - Blank",
    "Venue Estimate - Sample",
    "Venue Estimate - Blank",
    "Decor Estimate - Sample",
    "Decor Estimate - Blank",
]

for i, name in enumerate(desired):
    current_idx = wb.sheetnames.index(name)
    wb.move_sheet(name, offset=i - current_idx)

print(f"   Final order: {wb.sheetnames}")

# ═══════════════════════════════════════════════════════════════════════════
# STEP 4: UPDATE NAMED RANGES
# ═══════════════════════════════════════════════════════════════════════════

print("\n4. Updating named ranges...")
# Point all named ranges to 'Client Setup - Blank' (reference data is identical)
for name, defn in wb.defined_names.items():
    old_text = defn.attr_text
    if "'Client Setup'" in old_text:
        new_text = old_text.replace("'Client Setup'", "'Client Setup - Blank'")
        defn.attr_text = new_text
        print(f"   {name}: {old_text} → {new_text}")

# ═══════════════════════════════════════════════════════════════════════════
# STEP 5: UPDATE ALL CROSS-SHEET FORMULA REFERENCES
# ═══════════════════════════════════════════════════════════════════════════

print("\n5. Updating cross-sheet references in formulas...")

def update_sheet_refs(ws, old_ref, new_ref):
    """Replace 'old_ref'! with 'new_ref'! in all formula cells."""
    old_pattern = f"'{old_ref}'!"
    new_pattern = f"'{new_ref}'!"
    count = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                if old_pattern in cell.value:
                    cell.value = cell.value.replace(old_pattern, new_pattern)
                    count += 1
    return count

# Sample tabs → reference 'Client Setup - Sample'
for sheet_name in ["Venue Estimate - Sample", "Decor Estimate - Sample"]:
    ws = wb[sheet_name]
    n = update_sheet_refs(ws, "Client Setup", "Client Setup - Sample")
    print(f"   {sheet_name}: updated {n} formula refs → 'Client Setup - Sample'")

# Blank tabs → reference 'Client Setup - Blank'
for sheet_name in ["Venue Estimate - Blank", "Decor Estimate - Blank"]:
    ws = wb[sheet_name]
    n = update_sheet_refs(ws, "Client Setup", "Client Setup - Blank")
    print(f"   {sheet_name}: updated {n} formula refs → 'Client Setup - Blank'")

# ═══════════════════════════════════════════════════════════════════════════
# STEP 6: CLEAR BLANK TABS
# ═══════════════════════════════════════════════════════════════════════════

print("\n6. Clearing input data on blank tabs...")

def clear_unlocked_values(ws, skip_cells=None):
    """Clear all unlocked cells that are NOT formulas."""
    skip = set(skip_cells or [])
    count = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            coord = cell.coordinate
            if coord in skip:
                continue
            if not cell.protection.locked:
                if cell.value is not None and not (isinstance(cell.value, str) and cell.value.startswith('=')):
                    cell.value = None
                    count += 1
    return count

# --- Client Setup - Blank ---
cs_b = wb["Client Setup - Blank"]
# Only clear the specific cells the user listed
cs_clear_ranges = ["B5:B13", "B16", "B35", "B36", "B37", "C37"]
cleared = 0
for rng in cs_clear_ranges:
    cells = cs_b[rng]
    if not isinstance(cells, tuple):
        cells = ((cells,),)
    for row in cells:
        if not hasattr(row, '__iter__'):
            row = (row,)
        for cell in row:
            if cell.value is not None:
                cell.value = None
                cleared += 1

# Reset fee defaults
cs_b["B40"] = "20%"
cs_b["B41"] = "20%"
cs_b["B42"] = "5%"
print(f"   Client Setup - Blank: cleared {cleared} cells, reset B40:B42 to defaults")

# --- Venue Estimate - Blank ---
venue_b = wb["Venue Estimate - Blank"]
n = clear_unlocked_values(venue_b)
print(f"   Venue Estimate - Blank: cleared {n} unlocked value cells")

# --- Decor Estimate - Blank ---
decor_b = wb["Decor Estimate - Blank"]
n = clear_unlocked_values(decor_b)
# Re-apply structural category defaults on fee rows
for r in range(27, 30):  # Floral fee rows
    decor_b[f"I{r}"] = "Delivery & Logistics"
for r in range(69, 75):  # Rental fee rows
    decor_b[f"I{r}"] = "Delivery & Logistics"
# AV fee rows keep "Staffing & Labor"
for r in range(91, 96):
    decor_b[f"I{r}"] = "Staffing & Labor"
print(f"   Decor Estimate - Blank: cleared {n} unlocked value cells, restored fee categories")

# ═══════════════════════════════════════════════════════════════════════════
# STEP 7: RE-APPLY SHEET PROTECTION (copy_worksheet may lose it)
# ═══════════════════════════════════════════════════════════════════════════

print("\n7. Verifying/re-applying sheet protection...")
for name in wb.sheetnames:
    ws = wb[name]
    if not ws.protection.sheet:
        ws.protection.sheet = True
        ws.protection.set_password("")
        ws.protection.formatCells = False
        ws.protection.formatColumns = False
        ws.protection.formatRows = False
        ws.protection.insertRows = False
        ws.protection.sort = False
        ws.protection.autoFilter = False
        print(f"   {name}: re-applied protection")
    else:
        print(f"   {name}: protection OK")

# ═══════════════════════════════════════════════════════════════════════════
# STEP 8: VERIFICATION
# ═══════════════════════════════════════════════════════════════════════════

print("\n8. Verification...")

# Check tab structure
print(f"\n   Tab order: {wb.sheetnames}")
assert wb.sheetnames == desired, f"Tab order mismatch! Expected {desired}"
print("   ✓ Tab order correct")

# Check Sample tabs have test data
vs = wb["Venue Estimate - Sample"]
assert vs["B5"].value == "TEST RESTAURANT", f"B5={vs['B5'].value}"
assert vs["B7"].value == 15000, f"B7={vs['B7'].value}"
assert vs["C24"].value == 100, f"C24={vs['C24'].value}"
print("   ✓ Venue Sample has test data")

css = wb["Client Setup - Sample"]
assert css["B5"].value == "TEST CLIENT"
assert css["B7"].value == 100
assert css["B16"].value == "DC"
print("   ✓ Client Setup Sample has test data")

ds = wb["Decor Estimate - Sample"]
assert ds["B9"].value is not None and "101 Constitution" in str(ds["B9"].value)
assert ds["A16"].value is not None  # Should have sample data
print(f"   ✓ Decor Sample has data (B9={ds['B9'].value})")

# Check Blank tabs are clean
vb = wb["Venue Estimate - Blank"]
assert vb["B5"].value is None, f"Venue Blank B5 should be empty, got {vb['B5'].value}"
assert vb["C24"].value is None, f"Venue Blank C24 should be empty, got {vb['C24'].value}"
assert vb["B7"].value is None, f"Venue Blank B7 should be empty, got {vb['B7'].value}"
print("   ✓ Venue Blank inputs cleared")

csb = wb["Client Setup - Blank"]
assert csb["B5"].value is None, f"CS Blank B5 should be empty, got {csb['B5'].value}"
assert csb["B16"].value is None
assert csb["B40"].value == "20%"
assert csb["B41"].value == "20%"
assert csb["B42"].value == "5%"
print("   ✓ Client Setup Blank cleared with defaults")

db = wb["Decor Estimate - Blank"]
assert db["A16"].value is None, f"Decor Blank A16 should be empty, got {db['A16'].value}"
assert db["I69"].value == "Delivery & Logistics"
assert db["I70"].value == "Delivery & Logistics"
print("   ✓ Decor Blank cleared, fee categories intact")

# Check cross-sheet references on Sample tabs
print("\n   Checking cross-sheet refs on Sample tabs...")
sample_ref = "'Client Setup - Sample'!"
for cell_addr in ["B9", "B10", "B11"]:
    val = vs[cell_addr].value
    if isinstance(val, str) and val.startswith('='):
        assert sample_ref in val, f"Venue Sample {cell_addr} has wrong ref: {val}"
print(f"   ✓ Venue Sample refs point to 'Client Setup - Sample'")

# Check a few on Decor Sample
for cell_addr in ["B5", "B6", "B7"]:
    val = ds[cell_addr].value
    if isinstance(val, str) and val.startswith('='):
        assert sample_ref in val, f"Decor Sample {cell_addr} has wrong ref: {val}"
print(f"   ✓ Decor Sample refs point to 'Client Setup - Sample'")

# Check cross-sheet refs on Blank tabs
blank_ref = "'Client Setup - Blank'!"
for cell_addr in ["C17", "C18", "C19"]:
    val = vb[cell_addr].value
    if isinstance(val, str) and val.startswith('='):
        assert blank_ref in val, f"Venue Blank {cell_addr} has wrong ref: {val}"
print(f"   ✓ Venue Blank C17:C19 point to 'Client Setup - Blank'")

for cell_addr in ["B5", "B6", "B7"]:
    val = db[cell_addr].value
    if isinstance(val, str) and val.startswith('='):
        assert blank_ref in val, f"Decor Blank {cell_addr} has wrong ref: {val}"
print(f"   ✓ Decor Blank refs point to 'Client Setup - Blank'")

# Check formulas intact on Sample tabs
print("\n   Checking key formulas on Venue Sample...")
formula_checks = {
    "D24": "=B24*C24",
    "D43": "=SUM(D24:D26)",
    "D47": "=IFERROR(D28*F28+D29*F29+D30*F30+D31*F31,0)",
    "J23": "=SUM(J15:J21)",
}
for cell_addr, expected_fragment in formula_checks.items():
    val = vs[cell_addr].value
    assert val == expected_fragment, f"Venue Sample {cell_addr}: expected '{expected_fragment}', got '{val}'"
    print(f"   ✓ {cell_addr} = {val}")

# Check D51 has robust formula
d51 = vs["D51"].value
assert "SUBSTITUTE" in d51, f"Venue Sample D51 missing robust formula: {d51}"
print(f"   ✓ D51 has robust fee parsing formula")

# Check B62 compares D43
b62 = vs["B62"].value
assert "D43>=B7" in b62 or "D43>=" in b62, f"Venue Sample B62 wrong: {b62}"
print(f"   ✓ B62 compares D43 against B7")

# Check formulas on Venue Blank are intact (not cleared)
print("\n   Checking formulas intact on Venue Blank...")
for cell_addr in ["D24", "E24", "D43", "D47", "D51", "B62", "J23"]:
    val = vb[cell_addr].value
    assert isinstance(val, str) and val.startswith('='), f"Venue Blank {cell_addr} formula missing: {val}"
print("   ✓ All key formulas intact on Venue Blank")

# Check protection
print("\n   Checking protection on all tabs...")
for name in wb.sheetnames:
    ws = wb[name]
    assert ws.protection.sheet, f"{name} protection not active!"
    print(f"   ✓ {name}: protected")

# Check named ranges
print("\n   Named ranges:")
for name, defn in wb.defined_names.items():
    print(f"   {name}: {defn.attr_text}")

# ═══════════════════════════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════════════════════════

wb.save(SRC)
print(f"\nSaved: {SRC}")
print("Done!")
