#!/usr/bin/env python3
"""
Phase 2: Apply consistent thin borders to all table cells across all tabs.
Preserves existing medium-bottom borders on total/subtotal rows.
"""

import openpyxl
from openpyxl.styles import Border, Side

SRC = '/Users/alex/2026-pricing-strategy/2026-pricing-strategy/docs/templates/QC_Estimate_Template_2026_v2.xlsx'
wb = openpyxl.load_workbook(SRC)

THIN = Side(style='thin', color='FFD4C5B0')
THIN_BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def add_thin_borders(ws, row_start, row_end, col_start, col_end, skip_empty=False):
    """Add thin borders to all cells in range. Optionally skip empty cells."""
    count = 0
    for row in range(row_start, row_end + 1):
        for col in range(col_start, col_end + 1):
            cell = ws.cell(row=row, column=col)
            if skip_empty and cell.value is None:
                continue
            # Preserve medium bottom if present
            existing = cell.border
            bottom = existing.bottom if existing.bottom.style == 'medium' else THIN
            cell.border = Border(top=THIN, bottom=bottom, left=THIN, right=THIN)
            count += 1
    return count


cs = wb['Client Setup']
ve = wb['Venue Estimate']
de = wb['Decor Estimate']
sde = wb['SAMPLE Decor Estimate']

total_fixed = 0

# ════════════════════════════════════════════════════════════════
# CLIENT SETUP
# ════════════════════════════════════════════════════════════════
print("Client Setup:")

# Rows 11-13: Company, Program, Client Hotel (A-B)
n = add_thin_borders(cs, 11, 13, 1, 2)
print(f"  Rows 11-13 (Company/Program/Hotel): {n} cells")
total_fixed += n

# Rows 23-34: Markup table (A-D, col D has notes)
n = add_thin_borders(cs, 23, 34, 1, 4)
print(f"  Rows 23-34 (Markup table): {n} cells")
total_fixed += n

# Rows 52-53: Main header + blank
n = add_thin_borders(cs, 52, 52, 1, 7)
total_fixed += n

# Transportation section (rows 54-78, cols A-C max)
n = add_thin_borders(cs, 54, 78, 1, 3, skip_empty=True)
print(f"  Rows 54-78 (Transportation rates): {n} cells")
total_fixed += n

# Hotel Rates (rows 80-90, cols A-C)
n = add_thin_borders(cs, 80, 90, 1, 3)
print(f"  Rows 80-90 (Hotel rates): {n} cells")
total_fixed += n

# Per Diem Rates (rows 92-101, cols A-D)
n = add_thin_borders(cs, 92, 101, 1, 4, skip_empty=True)
print(f"  Rows 92-101 (Per Diem rates): {n} cells")
total_fixed += n

# Vehicle Rates (rows 103-114, cols A-G)
n = add_thin_borders(cs, 103, 114, 1, 7)
print(f"  Rows 103-114 (Vehicle rates): {n} cells")
total_fixed += n

# Quick Budget Reference (rows 116-120, cols A-B)
n = add_thin_borders(cs, 116, 120, 1, 2)
print(f"  Rows 116-120 (Budget reference): {n} cells")
total_fixed += n

# Origin/Destination lists in column I (rows 52-71)
n = add_thin_borders(cs, 52, 71, 9, 9, skip_empty=True)
print(f"  Column I lists: {n} cells")
total_fixed += n


# ════════════════════════════════════════════════════════════════
# VENUE ESTIMATE
# ════════════════════════════════════════════════════════════════
print("\nVenue Estimate:")

# Row 65: Travel header
n = add_thin_borders(ve, 65, 65, 1, 4)
print(f"  Row 65 (Travel header): {n} cells")
total_fixed += n

# Rows 66-78: Travel inputs (already have borders per audit, but ensure complete)
n = add_thin_borders(ve, 66, 78, 1, 4)
print(f"  Rows 66-78 (Travel inputs): {n} cells")
total_fixed += n

# Rows 80-84: Calculated rows + Trip Total
n = add_thin_borders(ve, 80, 84, 1, 4)
print(f"  Rows 80-84 (Travel calcs + Trip Total): {n} cells")
total_fixed += n

# Row 86: Total Travel Expenses
n = add_thin_borders(ve, 86, 86, 1, 4)
print(f"  Row 86 (Total Travel): {n} cells")
total_fixed += n

# Rows 94-99: Margin Health through True Net Health
n = add_thin_borders(ve, 94, 99, 1, 5, skip_empty=True)
print(f"  Rows 94-99 (P&M lower section): {n} cells")
total_fixed += n


# ════════════════════════════════════════════════════════════════
# DECOR ESTIMATE + SAMPLE DECOR ESTIMATE
# ════════════════════════════════════════════════════════════════

for ws, name in [(de, "Decor Estimate"), (sde, "SAMPLE Decor Estimate")]:
    print(f"\n{name}:")

    # Column I category tags (rows 14-95)
    n = add_thin_borders(ws, 14, 95, 9, 9, skip_empty=True)
    print(f"  Column I (Category tags): {n} cells")
    total_fixed += n

    # Row 110: Travel header
    n = add_thin_borders(ws, 110, 110, 1, 4)
    print(f"  Row 110 (Travel header): {n} cells")
    total_fixed += n

    # Rows 111-123: Travel inputs
    n = add_thin_borders(ws, 111, 123, 1, 4)
    print(f"  Rows 111-123 (Travel inputs): {n} cells")
    total_fixed += n

    # Rows 125-129: Calculated rows + Trip Total
    n = add_thin_borders(ws, 125, 129, 1, 4)
    print(f"  Rows 125-129 (Travel calcs + Trip Total): {n} cells")
    total_fixed += n

    # Row 131: Total Travel Expenses
    n = add_thin_borders(ws, 131, 131, 1, 4)
    print(f"  Row 131 (Total Travel): {n} cells")
    total_fixed += n

    # Rows 133-144: P&M section
    n = add_thin_borders(ws, 133, 144, 1, 5, skip_empty=True)
    print(f"  Rows 133-144 (P&M section): {n} cells")
    total_fixed += n


# ════════════════════════════════════════════════════════════════
# SAVE & VERIFY
# ════════════════════════════════════════════════════════════════
print(f"\nSaving... ({total_fixed} cells updated)")
wb.save(SRC)

# Verify
print("\n" + "=" * 60)
print("VERIFICATION")
print("=" * 60)

wb2 = openpyxl.load_workbook(SRC)
for sn in wb2.sheetnames:
    ws = wb2[sn]
    no_border = 0
    partial = 0
    full = 0
    for row in range(1, min(ws.max_row + 1, 200)):
        for col in range(1, min(ws.max_column + 1, 20)):
            cell = ws.cell(row=row, column=col)
            if cell.value is None:
                continue
            sides = [cell.border.top.style, cell.border.bottom.style,
                     cell.border.left.style, cell.border.right.style]
            has = sum(1 for s in sides if s is not None)
            if has == 0:
                no_border += 1
            elif has < 4:
                partial += 1
            else:
                full += 1
    total = no_border + partial + full
    print(f"  {sn}: {full}/{total} full ({100*full/total:.0f}%), {partial} partial, {no_border} no border")

print(f"\n{'=' * 60}")
print("BORDERS COMPLETE")
print(f"{'=' * 60}")
