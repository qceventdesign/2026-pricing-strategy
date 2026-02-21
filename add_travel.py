#!/usr/bin/env python3
"""
Add Travel Expense Calculator to QC_Estimate_Template_2026_v2.xlsx

1. Client Setup: Travel Rate Table (rows 52-63) + Vehicle Rate Table (rows 65-72)
2. Venue/Decor/SAMPLE Decor: Insert 18-row travel section before Profit & Margin
3. Rewrite all Profit & Margin formulas after row shift
4. True Net Profit now = QC Revenue - OpEx - Travel Expenses
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter

DST = '/Users/alex/2026-pricing-strategy/2026-pricing-strategy/docs/templates/QC_Estimate_Template_2026_v2.xlsx'
wb = openpyxl.load_workbook(DST)

cs = wb['Client Setup']
ve = wb['Venue Estimate']
de = wb['Decor Estimate']
sde = wb['SAMPLE Decor Estimate']

# ── Styles (matching fix_v2.py) ──
FILL_INPUT   = PatternFill('solid', fgColor='FFFFF8F0')   # peach — user inputs
FILL_CALC    = PatternFill('solid', fgColor='FFF5F0EB')   # beige — calculated
FILL_ROW_BG  = PatternFill('solid', fgColor='FFFAF6F3')   # cream — label rows
FILL_HEADER  = PatternFill('solid', fgColor='FFECDFCE')   # section headers
FILL_SUBHEAD = PatternFill('solid', fgColor='FFE8D9C8')   # sub-headers / accents

FONT_SECTION = Font(name='Playfair Display', size=11, bold=True, color='FF846E60')
FONT_COLHEAD = Font(name='Playfair Display', size=10, bold=True, color='FFC19C81')
FONT_LABEL   = Font(name='Playfair Display', size=10, color='FF464543')
FONT_INPUT   = Font(name='Playfair Display', size=10, bold=True, color='FF846E60')
FONT_BOLD    = Font(name='Playfair Display', size=10, bold=True, color='FF464543')
FONT_CALC    = Font(name='Playfair Display', size=10, color='FF846E60')

AC = Alignment(horizontal='center', vertical='center')
AL = Alignment(horizontal='left', vertical='center')
AR = Alignment(horizontal='right', vertical='center')

THIN_BORDER = Border(
    bottom=Side(style='thin', color='FFD4C5B0')
)

changes = []


def log(sheet, cell_ref, desc):
    changes.append(f"  {sheet}!{cell_ref}: {desc}")


# ════════════════════════════════════════════════════════════════
# STEP 1: Client Setup — Travel Rate Table (rows 52-63)
# ════════════════════════════════════════════════════════════════
print("Step 1: Travel Rate Table on Client Setup...")

# Header row 52
cs.cell(row=52, column=1).value = "QC TRAVEL RATE TABLE"
for c in range(1, 7):
    cs.cell(row=52, column=c).fill = FILL_HEADER
    cs.cell(row=52, column=c).font = FONT_SECTION
log("Client Setup", "A52", "QC TRAVEL RATE TABLE header")

# Column headers row 53
col_headers = ["Destination", "Default Travel Type", "Drive Mileage (per trip)",
               "Flight Cost (per person)", "Hotel/Night (per person)", "Per Diem/Day (per person)"]
for i, h in enumerate(col_headers):
    cell = cs.cell(row=53, column=1 + i)
    cell.value = h
    cell.fill = FILL_SUBHEAD
    cell.font = FONT_COLHEAD
    cell.alignment = AC
log("Client Setup", "A53:F53", "Travel rate column headers")

# Travel rate data rows 54-63
TRAVEL_DATA = [
    ("NYC",           "Flight", 300, 500, 500, 100),
    ("DC",            "Drive",  200, 500, 250,  75),
    ("Philadelphia",  "Drive",  200, 500, 250,  75),
    ("Charlotte",     "Drive",  100, 400, 200,  75),
    ("Raleigh NC",    "Drive",  100, 400, 200,  75),
    ("Asheville NC",  "Drive",  100, 400, 200,  75),
    ("Charleston SC", "Drive",  100, 400, 200,  75),
    ("Atlanta",       "Drive",  150, 400, 200,  75),
    ("Richmond VA",   "Drive",  150, 400, 200,  75),
    ("Other",         "Drive",  150, 500, 200,  75),
]

for i, (dest, ttype, mileage, flight, hotel, perdiem) in enumerate(TRAVEL_DATA):
    row = 54 + i
    cs.cell(row=row, column=1).value = dest
    cs.cell(row=row, column=1).font = FONT_LABEL
    cs.cell(row=row, column=1).fill = FILL_ROW_BG
    cs.cell(row=row, column=1).alignment = AL

    cs.cell(row=row, column=2).value = ttype
    cs.cell(row=row, column=2).font = FONT_INPUT
    cs.cell(row=row, column=2).fill = FILL_INPUT
    cs.cell(row=row, column=2).alignment = AC

    for j, val in enumerate([mileage, flight, hotel, perdiem], start=3):
        cell = cs.cell(row=row, column=j)
        cell.value = val
        cell.font = FONT_INPUT
        cell.fill = FILL_INPUT
        cell.alignment = AC
        cell.number_format = '$#,##0'

log("Client Setup", "A54:F63", f"10 destination rows with travel rates")

# Named ranges
dn_travel = DefinedName('TravelRates', attr_text="'Client Setup'!$A$54:$F$63")
wb.defined_names.add(dn_travel)
dn_dest = DefinedName('DestinationList', attr_text="'Client Setup'!$A$54:$A$63")
wb.defined_names.add(dn_dest)
log("Client Setup", "Named Ranges", "TravelRates (A54:F63), DestinationList (A54:A63)")


# ════════════════════════════════════════════════════════════════
# STEP 2: Client Setup — Vehicle Rate Table (rows 65-72)
# ════════════════════════════════════════════════════════════════
print("Step 2: Vehicle Rate Table on Client Setup...")

# Header row 65
cs.cell(row=65, column=1).value = "VEHICLE RATE TABLE"
for c in range(1, 3):
    cs.cell(row=65, column=c).fill = FILL_HEADER
    cs.cell(row=65, column=c).font = FONT_SECTION
log("Client Setup", "A65", "VEHICLE RATE TABLE header")

# Column headers row 66
for i, h in enumerate(["Vehicle", "Cost"]):
    cell = cs.cell(row=66, column=1 + i)
    cell.value = h
    cell.fill = FILL_SUBHEAD
    cell.font = FONT_COLHEAD
    cell.alignment = AC
log("Client Setup", "A66:B66", "Vehicle rate column headers")

# Vehicle data rows 67-72
VEHICLE_DATA = [
    ("None",                    0),
    ("SUV ($450)",            450),
    ("Sprinter ($850)",       850),
    ("Mini Bus ($1,200)",    1200),
    ("Motor Coach ($1,800)", 1800),
    ("Custom",                  0),
]

for i, (vehicle, cost) in enumerate(VEHICLE_DATA):
    row = 67 + i
    cs.cell(row=row, column=1).value = vehicle
    cs.cell(row=row, column=1).font = FONT_LABEL
    cs.cell(row=row, column=1).fill = FILL_ROW_BG
    cs.cell(row=row, column=1).alignment = AL

    cs.cell(row=row, column=2).value = cost
    cs.cell(row=row, column=2).font = FONT_INPUT
    cs.cell(row=row, column=2).fill = FILL_INPUT
    cs.cell(row=row, column=2).alignment = AC
    cs.cell(row=row, column=2).number_format = '$#,##0'

log("Client Setup", "A67:B72", "6 vehicle rows (None through Custom)")

# Named ranges
dn_vrates = DefinedName('VehicleRates', attr_text="'Client Setup'!$A$67:$B$72")
wb.defined_names.add(dn_vrates)
dn_vlist = DefinedName('VehicleList', attr_text="'Client Setup'!$A$67:$A$72")
wb.defined_names.add(dn_vlist)
log("Client Setup", "Named Ranges", "VehicleRates (A67:B72), VehicleList (A67:A72)")


# ════════════════════════════════════════════════════════════════
# STEP 3: Insert 18 rows on each estimate tab
# ════════════════════════════════════════════════════════════════
print("Step 3: Inserting rows...")

ve.insert_rows(64, amount=18)
log("Venue Estimate", "row 64", "Inserted 18 rows (Profit & Margin now rows 82-93)")

de.insert_rows(109, amount=18)
log("Decor Estimate", "row 109", "Inserted 18 rows (Profit & Margin now rows 127-138)")

sde.insert_rows(109, amount=18)
log("SAMPLE Decor Estimate", "row 109", "Inserted 18 rows (Profit & Margin now rows 127-138)")


# ════════════════════════════════════════════════════════════════
# STEP 4: Write travel section on each estimate tab
# ════════════════════════════════════════════════════════════════
print("Step 4: Writing travel sections...")


def write_travel_section(ws, sheet_label, base_row):
    """
    Write the QC Travel Expenses section starting at base_row.
    Uses 18 rows: base_row to base_row+17.
    """
    r = base_row  # 64 for Venue, 109 for Decor/SAMPLE

    # ── Row r+0: blank separator ──
    # (already blank from insert)

    # ── Row r+1: Section header ──
    ws.cell(row=r + 1, column=1).value = "QC TRAVEL EXPENSES"
    for c in range(1, 5):
        ws.cell(row=r + 1, column=c).fill = FILL_HEADER
        ws.cell(row=r + 1, column=c).font = FONT_SECTION

    # ── Row r+2: Column headers ──
    for i, label in enumerate(["", "Trip 1", "Trip 2", "Trip 3"]):
        cell = ws.cell(row=r + 2, column=1 + i)
        cell.value = label
        cell.fill = FILL_SUBHEAD
        cell.font = FONT_COLHEAD
        cell.alignment = AC

    # ── Input rows r+3 through r+9 ──
    input_labels = [
        "Trip Label",
        "Destination",
        "Travel Type",
        "Staff Traveling",
        "Nights",
        "On-Site Vehicle",
        "Custom Vehicle Cost",
    ]

    for i, label in enumerate(input_labels):
        row = r + 3 + i
        # Label in column A
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = FONT_LABEL
        ws.cell(row=row, column=1).fill = FILL_ROW_BG
        ws.cell(row=row, column=1).alignment = AL

        # Style input cells in B, C, D
        for col in range(2, 5):
            cell = ws.cell(row=row, column=col)
            cell.fill = FILL_INPUT
            cell.font = FONT_INPUT
            cell.alignment = AC

    # Custom Vehicle Cost — number format
    for col in range(2, 5):
        ws.cell(row=r + 9, column=col).number_format = '$#,##0'

    # ── Travel Type: auto-suggest formulas ──
    for col_idx, col_letter in [(2, 'B'), (3, 'C'), (4, 'D')]:
        dest_row = r + 4
        formula = f'=IF({col_letter}{dest_row}<>"",VLOOKUP({col_letter}{dest_row},TravelRates,2,FALSE),"")'
        ws.cell(row=r + 5, column=col_idx).value = formula
        ws.cell(row=r + 5, column=col_idx).fill = FILL_CALC
        ws.cell(row=r + 5, column=col_idx).font = FONT_CALC

    # ── Data Validations ──
    # Destination dropdown (r+4)
    dv_dest = DataValidation(type="list", formula1="DestinationList", allow_blank=True)
    dv_dest.sqref = f"B{r + 4} C{r + 4} D{r + 4}"
    ws.add_data_validation(dv_dest)

    # Travel Type dropdown (r+5) — override for auto-suggest
    dv_ttype = DataValidation(type="list", formula1='"Drive,Flight"', allow_blank=True)
    dv_ttype.sqref = f"B{r + 5} C{r + 5} D{r + 5}"
    ws.add_data_validation(dv_ttype)

    # Staff Traveling dropdown (r+6)
    dv_staff = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    dv_staff.sqref = f"B{r + 6} C{r + 6} D{r + 6}"
    ws.add_data_validation(dv_staff)

    # Nights dropdown (r+7)
    dv_nights = DataValidation(type="list", formula1='"0,1,2,3,4,5"', allow_blank=True)
    dv_nights.sqref = f"B{r + 7} C{r + 7} D{r + 7}"
    ws.add_data_validation(dv_nights)

    # On-Site Vehicle dropdown (r+8)
    dv_vehicle = DataValidation(type="list", formula1="VehicleList", allow_blank=True)
    dv_vehicle.sqref = f"B{r + 8} C{r + 8} D{r + 8}"
    ws.add_data_validation(dv_vehicle)

    # ── Row r+10: blank separator ──
    # (already blank)

    # ── Calculated rows r+11 through r+15 ──
    calc_labels = [
        "Travel Cost",
        "Hotel Cost",
        "Per Diem Cost",
        "Vehicle Cost",
        "Trip Total",
    ]

    for i, label in enumerate(calc_labels):
        row = r + 11 + i
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = FONT_BOLD if label == "Trip Total" else FONT_LABEL
        ws.cell(row=row, column=1).fill = FILL_ROW_BG
        ws.cell(row=row, column=1).alignment = AL

        # Style calculated cells
        for col in range(2, 5):
            cell = ws.cell(row=row, column=col)
            cell.fill = FILL_CALC
            cell.font = FONT_CALC
            cell.alignment = AC
            cell.number_format = '$#,##0'

    # Write formulas for each trip column
    for col_idx, cl in [(2, 'B'), (3, 'C'), (4, 'D')]:
        dest = f'{cl}{r + 4}'      # Destination
        ttype = f'{cl}{r + 5}'     # Travel Type
        staff = f'{cl}{r + 6}'     # Staff Traveling
        nights = f'{cl}{r + 7}'    # Nights
        vehicle = f'{cl}{r + 8}'   # On-Site Vehicle
        custom = f'{cl}{r + 9}'    # Custom Vehicle Cost

        # Travel Cost (r+11)
        # Flight: flight_cost × staff. Drive: mileage flat.
        ws.cell(row=r + 11, column=col_idx).value = (
            f'=IF(OR({dest}="",{staff}=""),0,'
            f'IF({ttype}="Flight",'
            f'VLOOKUP({dest},TravelRates,4,FALSE)*{staff},'
            f'IF({ttype}="Drive",'
            f'VLOOKUP({dest},TravelRates,3,FALSE),'
            f'0)))'
        )

        # Hotel Cost (r+12)
        # hotel_rate × nights × staff
        ws.cell(row=r + 12, column=col_idx).value = (
            f'=IF(OR({dest}="",{staff}="",{nights}=""),0,'
            f'VLOOKUP({dest},TravelRates,5,FALSE)*{nights}*{staff})'
        )

        # Per Diem Cost (r+13)
        # perdiem_rate × (nights+1) × staff
        ws.cell(row=r + 13, column=col_idx).value = (
            f'=IF(OR({dest}="",{staff}=""),0,'
            f'VLOOKUP({dest},TravelRates,6,FALSE)*({nights}+1)*{staff})'
        )

        # Vehicle Cost (r+14)
        ws.cell(row=r + 14, column=col_idx).value = (
            f'=IF(OR({vehicle}="",{vehicle}="None"),0,'
            f'IF({vehicle}="Custom",'
            f'IF(ISNUMBER({custom}),{custom},0),'
            f'IFERROR(VLOOKUP({vehicle},VehicleRates,2,FALSE),0)))'
        )

        # Trip Total (r+15)
        ws.cell(row=r + 15, column=col_idx).value = (
            f'={cl}{r + 11}+{cl}{r + 12}+{cl}{r + 13}+{cl}{r + 14}'
        )

    # Bold border on Trip Total row
    for col in range(1, 5):
        ws.cell(row=r + 15, column=col).border = THIN_BORDER
        if col >= 2:
            ws.cell(row=r + 15, column=col).font = Font(
                name='Playfair Display', size=10, bold=True, color='FF846E60'
            )

    # ── Row r+16: blank ──

    # ── Row r+17: Total Travel Expenses ──
    ws.cell(row=r + 17, column=1).value = "Total Travel Expenses"
    ws.cell(row=r + 17, column=1).font = FONT_BOLD
    ws.cell(row=r + 17, column=1).fill = FILL_SUBHEAD
    ws.cell(row=r + 17, column=1).alignment = AL

    # Put total in column D for consistency with Profit & Margin convention
    ws.cell(row=r + 17, column=4).value = f'=B{r + 15}+C{r + 15}+D{r + 15}'
    ws.cell(row=r + 17, column=4).fill = FILL_SUBHEAD
    ws.cell(row=r + 17, column=4).font = Font(
        name='Playfair Display', size=10, bold=True, color='FF846E60'
    )
    ws.cell(row=r + 17, column=4).alignment = AC
    ws.cell(row=r + 17, column=4).number_format = '$#,##0'

    # Style remaining cells in the total row
    for c in [2, 3]:
        ws.cell(row=r + 17, column=c).fill = FILL_SUBHEAD

    log(sheet_label, f"rows {r + 1}-{r + 17}",
        "QC Travel Expenses section (3 trips, 7 inputs + 5 calcs + total)")


# Write travel sections on all three tabs
write_travel_section(ve, "Venue Estimate", 64)
write_travel_section(de, "Decor Estimate", 109)
write_travel_section(sde, "SAMPLE Decor Estimate", 109)


# ════════════════════════════════════════════════════════════════
# STEP 5: Rewrite Profit & Margin formulas after row shift
# ════════════════════════════════════════════════════════════════
print("Step 5: Rewriting Profit & Margin formulas...")


def style_profit_row(ws, row, is_header=False, is_health=False, is_pct=False,
                     is_dollar=True, is_input=False):
    """Apply consistent styling to a Profit & Margin row."""
    if is_header:
        for c in range(1, 6):
            ws.cell(row=row, column=c).fill = FILL_HEADER
            ws.cell(row=row, column=c).font = FONT_SECTION
        return

    ws.cell(row=row, column=1).font = FONT_BOLD if is_health else FONT_LABEL
    ws.cell(row=row, column=1).fill = FILL_ROW_BG
    ws.cell(row=row, column=1).alignment = AL

    if is_input:
        ws.cell(row=row, column=3).fill = FILL_INPUT
        ws.cell(row=row, column=3).font = FONT_INPUT
        ws.cell(row=row, column=3).alignment = AC

    ws.cell(row=row, column=4).fill = FILL_CALC
    ws.cell(row=row, column=4).font = FONT_CALC
    ws.cell(row=row, column=4).alignment = AC

    if is_pct:
        ws.cell(row=row, column=4).number_format = '0.0%'
    elif is_dollar:
        ws.cell(row=row, column=4).number_format = '$#,##0'
    # Health indicator is plain text — no number format


# ── Venue Estimate: Profit & Margin (now rows 82-93) ──
# Travel total is at D81

# Row 82: Header
ve.cell(row=82, column=1).value = "PROFIT & MARGIN ANALYSIS"
style_profit_row(ve, 82, is_header=True)

# Row 83: Third-Party Commission
ve.cell(row=83, column=1).value = "Third-Party Commission"
ve.cell(row=83, column=3).value = "='Client Setup'!C36"
ve.cell(row=83, column=3).number_format = '0%'
ve.cell(row=83, column=4).value = "=(E43+E46+E48+E49+E51+E52+E53)*C83"
style_profit_row(ve, 83)

# Row 84: Global DMC Fee
ve.cell(row=84, column=1).value = "Global DMC Fee (6.5%)"
ve.cell(row=84, column=3).value = '=IF(\'Client Setup\'!C37="Yes",0.065,0)'
ve.cell(row=84, column=3).number_format = '0.0%'
ve.cell(row=84, column=4).value = '=IF(\'Client Setup\'!C37="Yes",(E43+E46+E48+E49+E51+E52+E53)*0.065,0)'
style_profit_row(ve, 84)

# Row 85: Total Vendor Costs
ve.cell(row=85, column=1).value = "Total Vendor Costs (incl. CC fee)"
ve.cell(row=85, column=4).value = "=D54+D55"
style_profit_row(ve, 85)

# Row 86: QC Revenue
ve.cell(row=86, column=1).value = "QC Revenue"
ve.cell(row=86, column=4).value = "=IFERROR(E57-D83-D84-D85,0)"
style_profit_row(ve, 86)
ve.cell(row=86, column=1).font = FONT_BOLD

# Row 87: QC Margin %
ve.cell(row=87, column=1).value = "QC Margin %"
ve.cell(row=87, column=4).value = "=IFERROR(IF(E57>0,D86/E57,0),0)"
style_profit_row(ve, 87, is_pct=True)

# Row 88: Margin Health
ve.cell(row=88, column=1).value = "Margin Health"
ve.cell(row=88, column=4).value = (
    '=IF(D87>=0.45,"\u2713 STRONG",'
    'IF(D87>=0.35,"\u2192 ON TARGET",'
    'IF(D87>=0.25,"\u26A0 REVIEW",'
    '"\u2717 BELOW FLOOR")))'
)
style_profit_row(ve, 88, is_health=True, is_dollar=False)

# Row 89: blank

# Row 90: Estimated Team Hours
ve.cell(row=90, column=1).value = "Estimated Team Hours"
ve.cell(row=90, column=3).value = None  # keep as input
ve.cell(row=90, column=4).value = "=IF(ISNUMBER(C90),C90*90,0)"
ve.cell(row=90, column=5).value = "Hours \u00D7 $90/hr"
ve.cell(row=90, column=5).font = FONT_LABEL
style_profit_row(ve, 90, is_input=True)

# Row 91: True Net Profit — NOW INCLUDES TRAVEL
ve.cell(row=91, column=1).value = "True Net Profit"
ve.cell(row=91, column=4).value = "=D86-D90-D81"
style_profit_row(ve, 91)
ve.cell(row=91, column=1).font = FONT_BOLD

# Row 92: True Net Margin %
ve.cell(row=92, column=1).value = "True Net Margin %"
ve.cell(row=92, column=4).value = "=IFERROR(IF(E57>0,D91/E57,0),0)"
style_profit_row(ve, 92, is_pct=True)

# Row 93: True Net Health
ve.cell(row=93, column=1).value = "True Net Health"
ve.cell(row=93, column=4).value = (
    '=IF(D92>=0.15,"\u2713 STRONG",'
    'IF(D92>=0.07,"\u2192 ON TARGET",'
    'IF(D92>=0,"\u26A0 THIN",'
    '"\u2717 LOSING MONEY")))'
)
style_profit_row(ve, 93, is_health=True, is_dollar=False)

log("Venue Estimate", "rows 82-93",
    "Rewrote all Profit & Margin formulas; D91 now = D86-D90-D81 (includes travel)")


# ── Decor Estimate: Profit & Margin (now rows 127-138) ──
# Travel total is at D126

def rewrite_decor_profit(ws, sheet_label):
    """Rewrite Profit & Margin for Decor-type sheets (rows 127-138)."""

    # Row 127: Header
    ws.cell(row=127, column=1).value = "PROFIT & MARGIN"
    style_profit_row(ws, 127, is_header=True)

    # Row 128: Third-Party Commission
    ws.cell(row=128, column=1).value = "Third-Party Commission"
    ws.cell(row=128, column=3).value = "='Client Setup'!C36"
    ws.cell(row=128, column=3).number_format = '0%'
    ws.cell(row=128, column=4).value = "=F104*C128"
    style_profit_row(ws, 128)

    # Row 129: Global DMC Fee
    ws.cell(row=129, column=1).value = "Global DMC Fee (6.5%)"
    ws.cell(row=129, column=3).value = '=IF(\'Client Setup\'!C37="Yes",0.065,0)'
    ws.cell(row=129, column=3).number_format = '0.0%'
    ws.cell(row=129, column=4).value = '=IF(\'Client Setup\'!C37="Yes",F104*0.065,0)'
    style_profit_row(ws, 129)

    # Row 130: Total Vendor Costs
    ws.cell(row=130, column=1).value = "Total Vendor Costs (incl. CC fee)"
    ws.cell(row=130, column=4).value = "=E99+E100+E101+E102"
    style_profit_row(ws, 130)

    # Row 131: QC Revenue
    ws.cell(row=131, column=1).value = "QC Revenue"
    ws.cell(row=131, column=4).value = "=IFERROR(F104-D128-D129-D130,0)"
    style_profit_row(ws, 131)
    ws.cell(row=131, column=1).font = FONT_BOLD

    # Row 132: QC Margin %
    ws.cell(row=132, column=1).value = "QC Margin %"
    ws.cell(row=132, column=4).value = "=IFERROR(IF(F104>0,D131/F104,0),0)"
    style_profit_row(ws, 132, is_pct=True)

    # Row 133: Margin Health
    ws.cell(row=133, column=1).value = "Margin Health"
    ws.cell(row=133, column=4).value = (
        '=IF(D132>=0.45,"\u2713 STRONG",'
        'IF(D132>=0.35,"\u2192 ON TARGET",'
        'IF(D132>=0.25,"\u26A0 REVIEW",'
        '"\u2717 BELOW FLOOR")))'
    )
    style_profit_row(ws, 133, is_health=True, is_dollar=False)

    # Row 134: blank

    # Row 135: Estimated Team Hours
    ws.cell(row=135, column=1).value = "Estimated Team Hours"
    ws.cell(row=135, column=3).value = None  # input
    ws.cell(row=135, column=4).value = "=IF(ISNUMBER(C135),C135*90,0)"
    ws.cell(row=135, column=5).value = "Hours \u00D7 $90/hr"
    ws.cell(row=135, column=5).font = FONT_LABEL
    style_profit_row(ws, 135, is_input=True)

    # Row 136: True Net Profit — NOW INCLUDES TRAVEL
    ws.cell(row=136, column=1).value = "True Net Profit"
    ws.cell(row=136, column=4).value = "=D131-D135-D126"
    style_profit_row(ws, 136)
    ws.cell(row=136, column=1).font = FONT_BOLD

    # Row 137: True Net Margin %
    ws.cell(row=137, column=1).value = "True Net Margin %"
    ws.cell(row=137, column=4).value = "=IFERROR(IF(F104>0,D136/F104,0),0)"
    style_profit_row(ws, 137, is_pct=True)

    # Row 138: True Net Health
    ws.cell(row=138, column=1).value = "True Net Health"
    ws.cell(row=138, column=4).value = (
        '=IF(D137>=0.15,"\u2713 STRONG",'
        'IF(D137>=0.07,"\u2192 ON TARGET",'
        'IF(D137>=0,"\u26A0 THIN",'
        '"\u2717 LOSING MONEY")))'
    )
    style_profit_row(ws, 138, is_health=True, is_dollar=False)

    log(sheet_label, "rows 127-138",
        "Rewrote all Profit & Margin formulas; D136 now = D131-D135-D126 (includes travel)")


rewrite_decor_profit(de, "Decor Estimate")
rewrite_decor_profit(sde, "SAMPLE Decor Estimate")


# ════════════════════════════════════════════════════════════════
# STEP 6: Final audit
# ════════════════════════════════════════════════════════════════
print("\nStep 6: Final audit...")

error_patterns = ["#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!"]
errors = []
stale_refs = []

for sn in wb.sheetnames:
    ws = wb[sn]
    for row in range(1, min(ws.max_row + 1, 200)):
        for col in range(1, min(ws.max_column + 1, 20)):
            v = str(ws.cell(row=row, column=col).value) if ws.cell(row=row, column=col).value else ''
            for ep in error_patterns:
                if ep in v:
                    errors.append(f"  {sn}!{ws.cell(row=row, column=col).coordinate}: {v[:60]}")
            if "'Client Setup'!C19" in v and "VLOOKUP" not in v:
                stale_refs.append(f"  {sn}!{ws.cell(row=row, column=col).coordinate}: stale C19")

if errors:
    print(f"  ERRORS FOUND: {len(errors)}")
    for e in errors:
        print(e)
else:
    print("  No formula error strings found.")

if stale_refs:
    print(f"  STALE REFS: {len(stale_refs)}")
    for s in stale_refs:
        print(s)
else:
    print("  No stale C19 references.")

# ════════════════════════════════════════════════════════════════
# SAVE (before audit output to prevent data loss on audit errors)
# ════════════════════════════════════════════════════════════════
wb.save(DST)
print(f"\nSaved to {DST}")

# Verify named ranges
print("\n  Named ranges:")
for dn in wb.defined_names.values():
    print(f"    {dn.name} = {dn.attr_text}")

# Verify data validations per sheet
for sn in ['Venue Estimate', 'Decor Estimate', 'SAMPLE Decor Estimate']:
    ws = wb[sn]
    print(f"  {sn}: {len(ws.data_validations.dataValidation)} data validations total")

# ════════════════════════════════════════════════════════════════
# CHANGE SUMMARY
# ════════════════════════════════════════════════════════════════
print(f"\n{'=' * 60}")
print(f"CHANGE SUMMARY ({len(changes)} changes)")
print(f"{'=' * 60}")
for c in changes:
    print(c)
