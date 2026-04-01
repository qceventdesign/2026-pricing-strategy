"""
Fix 8 issues in QC_Estimate_Template_2026_v2_FINAL.xlsx.
Does NOT move cells or change formulas.
"""
import os
from copy import copy
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

os.chdir(os.path.join(os.path.dirname(__file__), '2026-pricing-strategy', 'docs', 'templates'))
LOGO = os.path.join(os.path.dirname(__file__), 'Quill_Creative_Branding-03.png')

wb = load_workbook('QC_Estimate_Template_2026_v2_FINAL.xlsx')

# ── Brand palette ──
MOCHA   = PatternFill('solid', fgColor='846E60')
CHARCOAL_FILL = PatternFill('solid', fgColor='464543')
LINEN   = PatternFill('solid', fgColor='FAF6F3')
CREAM   = PatternFill('solid', fgColor='ECDFCE')
WHITE_FONT = Font(color='FFFFFF')
WHITE_FONT_BOLD = Font(color='FFFFFF', bold=True)
WHITE_FONT_BOLD_12 = Font(color='FFFFFF', bold=True, size=12)
WHITE_FONT_BOLD_UPPER = Font(color='FFFFFF', bold=True)  # text uppercased separately

# ────────────────────────────────────────────────────────────────────
# 1. INSERT LOGOS on all 6 sheets
# ────────────────────────────────────────────────────────────────────
for ws in wb.worksheets:
    img = XlImage(LOGO)
    # Scale to ~60px tall, maintain aspect ratio
    orig_w, orig_h = img.width, img.height
    target_h = 60
    scale = target_h / orig_h
    img.width = int(orig_w * scale)
    img.height = target_h
    img.anchor = 'A1'
    ws.add_image(img)
print('[1] Logos inserted on all 6 sheets.')

# ────────────────────────────────────────────────────────────────────
# 2. ALTERNATING ROW TINTS on data entry rows
# ────────────────────────────────────────────────────────────────────
def apply_alternating_tints(ws, row_groups, col_start='A', col_end='G'):
    """Apply Linen (odd) / Cream (even) tints to data rows."""
    col_s = ord(col_start)
    col_e = ord(col_end)
    row_index = 0
    for group in row_groups:
        for r in group:
            fill = LINEN if row_index % 2 == 0 else CREAM
            for c in range(col_s, col_e + 1):
                cell = ws[f'{chr(c)}{r}']
                cell.fill = fill
            row_index += 1

# Venue data row groups
venue_rows = [
    list(range(24, 27)),   # F&B: 24-26
    list(range(28, 32)),   # Equipment: 28-31
    list(range(33, 36)),   # Venue fees: 33-35
    list(range(38, 41)),   # Staffing: 38-40
]

for name in ['Venue Estimate - Sample', 'Venue Estimate - Blank']:
    apply_alternating_tints(wb[name], venue_rows, 'A', 'G')

# Decor data row groups
decor_rows = [
    list(range(16, 26)),   # Florals: 16-25
    list(range(35, 43)),   # Seating: 35-42
    list(range(44, 52)),   # Lounge: 44-51
    list(range(53, 61)),   # Tables: 53-60
    list(range(62, 68)),   # Accessories: 62-67
]

for name in ['Decor Estimate - Sample', 'Decor Estimate - Blank']:
    apply_alternating_tints(wb[name], decor_rows, 'A', 'H')

print('[2] Alternating row tints applied.')

# ────────────────────────────────────────────────────────────────────
# 3. CLIENT-READY TABLE STYLING
# ────────────────────────────────────────────────────────────────────

def style_cell(cell, fill=None, font=None, uppercase=False):
    """Apply fill/font to cell, optionally uppercase the value."""
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if uppercase and isinstance(cell.value, str):
        cell.value = cell.value.upper()

# -- Venue sheets: I14:K24 --
for suffix in ['Sample', 'Blank']:
    ws = wb[f'Venue Estimate - {suffix}']

    # Header row I14:K14 — add labels + Mocha fill, white bold uppercase
    headers = {'I14': 'CATEGORY', 'J14': 'AMOUNT', 'K14': 'PER PERSON'}
    for ref, label in headers.items():
        cell = ws[ref]
        cell.value = label
        cell.fill = MOCHA
        cell.font = WHITE_FONT_BOLD
        cell.alignment = Alignment(horizontal='center')

    # Data rows I15:K21 — Linen fill
    for r in range(15, 22):
        for col in 'IJK':
            ws[f'{col}{r}'].fill = LINEN

    # Total row I23:K23 — Charcoal fill, white bold 12pt
    for col in 'IJK':
        cell = ws[f'{col}23']
        cell.fill = CHARCOAL_FILL
        cell.font = WHITE_FONT_BOLD_12

# -- Decor sheets: J/K client-ready columns --
# Header: J14:K14 (already has text 'Item'/'Amount')
# Subtotals: J29, J68, J86
# Total: J90
for suffix in ['Sample', 'Blank']:
    ws = wb[f'Decor Estimate - {suffix}']

    # Header row J14:K14
    for col in 'JK':
        cell = ws[f'{col}14']
        cell.fill = MOCHA
        cell.font = WHITE_FONT_BOLD
        if isinstance(cell.value, str):
            cell.value = cell.value.upper()

    # Data rows J15:K28, J32:K67, J71:K85 — Linen fill
    data_ranges = list(range(15, 29)) + list(range(30, 68)) + list(range(69, 86)) + [88]
    for r in data_ranges:
        for col in 'JK':
            cell = ws[f'{col}{r}']
            if cell.value is not None:
                cell.fill = LINEN

    # Subtotal rows — Cream fill, Charcoal text, bold
    for r in [29, 68, 86]:
        for col in 'JK':
            cell = ws[f'{col}{r}']
            cell.fill = CREAM
            cell.font = Font(color='464543', bold=True)

    # Total row J90:K90 — Charcoal fill, white bold 12pt
    for col in 'JK':
        cell = ws[f'{col}90']
        cell.fill = CHARCOAL_FILL
        cell.font = WHITE_FONT_BOLD_12

print('[3] Client-ready table styling applied.')

# ────────────────────────────────────────────────────────────────────
# 4. FONT — skip (Playfair Display is correct)
# ────────────────────────────────────────────────────────────────────
print('[4] Skipped — Playfair Display is correct.')

# ────────────────────────────────────────────────────────────────────
# 5. PRINT SETUP
# ────────────────────────────────────────────────────────────────────
for ws in wb.worksheets:
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.oddHeader.center.text = 'Quill Creative Event Design — Confidential'
    ws.oddFooter.center.text = 'Page &P of &N'

print('[5] Print setup configured on all 6 sheets.')

# ────────────────────────────────────────────────────────────────────
# 6. CLIENT SETUP - BLANK DEFAULTS
# ────────────────────────────────────────────────────────────────────
ws_cb = wb['Client Setup - Blank']
ws_cb['B35'].value = 0.035
ws_cb['B35'].number_format = '0.0%'
ws_cb['B36'].value = 0
ws_cb['B36'].number_format = '0%'
ws_cb['B37'].value = 'Yes'
ws_cb['C37'].value = 0.065
ws_cb['C37'].number_format = '0.0%'

print('[6] Client Setup - Blank defaults set.')

# ────────────────────────────────────────────────────────────────────
# 7. COLUMN H WIDTH on Venue and Decor sheets
# ────────────────────────────────────────────────────────────────────
for name in ['Venue Estimate - Sample', 'Venue Estimate - Blank',
             'Decor Estimate - Sample', 'Decor Estimate - Blank']:
    wb[name].column_dimensions['H'].width = 32

print('[7] Column H width set to 32.')

# ────────────────────────────────────────────────────────────────────
# 8. TOTAL ROW FONT COLOR → pure White
# ────────────────────────────────────────────────────────────────────
# Venue sheets row 57
for name in ['Venue Estimate - Sample', 'Venue Estimate - Blank']:
    ws = wb[name]
    for col in 'ABCDEFG':
        cell = ws[f'{col}57']
        # Preserve existing font properties, just change color
        old = cell.font
        cell.font = Font(
            name=old.name, size=old.size, bold=old.bold, italic=old.italic,
            underline=old.underline, strike=old.strike, color='FFFFFF',
        )

# Decor sheets row 104
for name in ['Decor Estimate - Sample', 'Decor Estimate - Blank']:
    ws = wb[name]
    for col in 'ABCDEFGHIJK':
        cell = ws[f'{col}104']
        if cell.value is not None or cell.font.color:
            old = cell.font
            cell.font = Font(
                name=old.name, size=old.size, bold=old.bold, italic=old.italic,
                underline=old.underline, strike=old.strike, color='FFFFFF',
            )

print('[8] Total row font color set to white.')

# ────────────────────────────────────────────────────────────────────
# SAVE + VERIFY
# ────────────────────────────────────────────────────────────────────
wb.save('QC_Estimate_Template_2026_v2_FINAL.xlsx')
print('\nSaved. Verifying formulas...')

# Reload and spot-check that formulas are intact
wb2 = load_workbook('QC_Estimate_Template_2026_v2_FINAL.xlsx')
checks = [
    ('Venue Estimate - Sample', 'D57', '=D54+D55+D56'),
    ('Venue Estimate - Sample', 'J23', '=SUM(J15:J21)'),
    ('Venue Estimate - Blank',  'E57', '=E54+E55+E56'),
    ('Decor Estimate - Sample', 'E104', '=SUM(E99:E103)'),
    ('Decor Estimate - Sample', 'K90', '=K29+K68+K86+K88+ROUNDUP(F102/10,0)*10+ROUNDUP(F103/10,0)*10'),
    ('Decor Estimate - Blank',  'F104', '=SUM(F99:F103)'),
]
all_ok = True
for sheet, ref, expected in checks:
    actual = wb2[sheet][ref].value
    if actual != expected:
        print(f'  MISMATCH {sheet}!{ref}: expected {expected!r}, got {actual!r}')
        all_ok = False

if all_ok:
    print('All formula spot-checks passed.')
else:
    print('WARNING: Some formulas may have changed!')
