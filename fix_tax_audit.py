#!/usr/bin/env python3
"""
Phase 1 Fix: Tax formulas in B17-B19 point to wrong range.
B17-B19 reference $A$75:$D$96 (travel data) instead of $A$124:$D$145 (tax table).
All estimate tabs reference B17-B19, causing 100+ #N/A errors.
Fix: Update B17-B19 to correct range, remove redundant C17-C19.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

SRC = '/Users/alex/2026-pricing-strategy/2026-pricing-strategy/docs/templates/QC_Estimate_Template_2026_v2.xlsx'
wb = openpyxl.load_workbook(SRC)
cs = wb['Client Setup']

# ════════════════════════════════════════════════════════════════
# FIX: Update B17-B19 to reference correct tax table
# ════════════════════════════════════════════════════════════════
print("Fixing B17-B19 tax VLOOKUP formulas...")

old_b17 = cs.cell(row=17, column=2).value
old_b18 = cs.cell(row=18, column=2).value
old_b19 = cs.cell(row=19, column=2).value

cs.cell(row=17, column=2).value = '=IF(B16="","",VLOOKUP(B16,$A$124:$D$145,2,FALSE))'
cs.cell(row=18, column=2).value = '=IF(B16="","",VLOOKUP(B16,$A$124:$D$145,3,FALSE))'
cs.cell(row=19, column=2).value = '=IF(B16="","",VLOOKUP(B16,$A$124:$D$145,4,FALSE))'

print(f"  B17: was '{old_b17}'")
print(f"       now '=IF(B16=\"\",\"\",VLOOKUP(B16,$A$124:$D$145,2,FALSE))'")
print(f"  B18: was '{old_b18}'")
print(f"       now '=IF(B16=\"\",\"\",VLOOKUP(B16,$A$124:$D$145,3,FALSE))'")
print(f"  B19: was '{old_b19}'")
print(f"       now '=IF(B16=\"\",\"\",VLOOKUP(B16,$A$124:$D$145,4,FALSE))'")

# Remove redundant C17-C19 (Phase 1 wrote duplicates there)
print("\nClearing redundant C17-C19...")
for r in [17, 18, 19]:
    old_val = cs.cell(row=r, column=3).value
    cs.cell(row=r, column=3).value = None
    print(f"  C{r}: cleared (was '{old_val}')")

# ════════════════════════════════════════════════════════════════
# SAVE
# ════════════════════════════════════════════════════════════════
print("\nSaving...")
wb.save(SRC)

# ════════════════════════════════════════════════════════════════
# FULL AUDIT: Scan all sheets for formula errors
# ════════════════════════════════════════════════════════════════
print("\n" + "=" * 60)
print("FULL FORMULA AUDIT")
print("=" * 60)

wb2 = openpyxl.load_workbook(SRC)
error_patterns = ["#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!"]
total_errors = 0

for sn in wb2.sheetnames:
    ws = wb2[sn]
    sheet_errors = []
    for row in range(1, min(ws.max_row + 1, 200)):
        for col in range(1, min(ws.max_column + 1, 20)):
            v = str(ws.cell(row=row, column=col).value) if ws.cell(row=row, column=col).value else ''
            for ep in error_patterns:
                if ep in v:
                    sheet_errors.append(f"    {ws.cell(row=row, column=col).coordinate}: {v[:80]}")
    if sheet_errors:
        print(f"\n  {sn}: {len(sheet_errors)} errors")
        for e in sheet_errors:
            print(e)
        total_errors += len(sheet_errors)
    else:
        print(f"  {sn}: No formula errors")

# Verify B17-B19 now reference correct range
cs2 = wb2['Client Setup']
print(f"\nVerification:")
for r in [17, 18, 19]:
    f = cs2.cell(row=r, column=2).value
    ok = '$A$124:$D$145' in str(f)
    print(f"  B{r}: {f} — {'PASS' if ok else 'FAIL'}")

# Verify tax table has data
print(f"\nTax table spot check (row 124):")
for col in range(1, 5):
    print(f"  {chr(64+col)}124 = {cs2.cell(row=124, column=col).value}")

# Verify named ranges
print(f"\nNamed ranges:")
for dn in sorted(wb2.defined_names.values(), key=lambda x: x.name):
    print(f"  {dn.name} = {dn.attr_text}")

print(f"\n{'=' * 60}")
print(f"AUDIT RESULT: {'CLEAN — no formula errors' if total_errors == 0 else f'{total_errors} ERRORS REMAINING'}")
print(f"{'=' * 60}")
