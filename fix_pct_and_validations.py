from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
import os

os.chdir(os.path.join(os.path.dirname(__file__), '2026-pricing-strategy', 'docs', 'templates'))

wb = load_workbook('QC_Estimate_Template_2026_v2_FINAL.xlsx')


def fmt_pct(ws, cells):
    """Convert string percentages to numbers and apply percent format."""
    for ref in cells:
        cell = ws[ref]
        if isinstance(cell.value, str) and cell.value.endswith('%'):
            cell.value = float(cell.value.strip('%')) / 100
        cell.number_format = '0%'


def copy_all_validations(src_ws, dst_ws):
    """Copy all data validations from source to destination sheet."""
    for dv in src_ws.data_validations.dataValidation:
        new_dv = DataValidation(
            type=dv.type,
            formula1=dv.formula1,
            formula2=dv.formula2,
            allow_blank=dv.allow_blank,
            showErrorMessage=dv.showErrorMessage,
            showInputMessage=dv.showInputMessage,
            errorTitle=dv.errorTitle,
            error=dv.error,
            promptTitle=dv.promptTitle,
            prompt=dv.prompt,
        )
        new_dv.sqref = dv.sqref
        dst_ws.add_data_validation(new_dv)


# --- Client Setup ---
ws_cs = wb['Client Setup - Sample']
ws_cb = wb['Client Setup - Blank']

pct_cells_client = ['B40', 'B41', 'B42', 'C40', 'C41', 'C42']
fmt_pct(ws_cs, pct_cells_client)
fmt_pct(ws_cb, pct_cells_client)

copy_all_validations(ws_cs, ws_cb)

# --- Venue Estimate ---
ws_vs = wb['Venue Estimate - Sample']
ws_vb = wb['Venue Estimate - Blank']

pct_cells_venue = ['C17', 'C18', 'C19']
fmt_pct(ws_vs, pct_cells_venue)
fmt_pct(ws_vb, pct_cells_venue)

copy_all_validations(ws_vs, ws_vb)

wb.save('QC_Estimate_Template_2026_v2_FINAL.xlsx')
print('Done.')
