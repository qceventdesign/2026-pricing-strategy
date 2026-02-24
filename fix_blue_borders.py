#!/usr/bin/env python3
"""
Task 1: Clean up stray blue fills (non-dropdown cells with blue).
Task 2: Fix inconsistent borders across all tabs.
"""

import openpyxl
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import range_boundaries, coordinate_to_tuple, get_column_letter
from openpyxl.cell.cell import MergedCell
from collections import Counter

SRC = '/Users/alex/2026-pricing-strategy/2026-pricing-strategy/docs/templates/QC_Estimate_Template_2026_v2.xlsx'
wb = openpyxl.load_workbook(SRC)


def get_dropdown_cells(ws):
    """Return set of (row, col) for all dropdown-validated cells."""
    cells = set()
    for dv in ws.data_validations.dataValidation:
        if dv.type != 'list':
            continue
        for sq_range in str(dv.sqref).split():
            if ':' in sq_range:
                min_col, min_row, max_col, max_row = range_boundaries(sq_range)
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        cells.add((r, c))
            else:
                r, c = coordinate_to_tuple(sq_range)
                cells.add((r, c))
    return cells


def is_mc(ws, row, col):
    return isinstance(ws.cell(row=row, column=col), MergedCell)


def has_full_borders(cell):
    if isinstance(cell, MergedCell):
        return False
    b = cell.border
    return all(getattr(b, s).style is not None for s in ['top', 'bottom', 'left', 'right'])


def get_border_color(cell):
    """Get the most common border color from a fully-bordered cell."""
    if isinstance(cell, MergedCell):
        return None
    colors = []
    for s in ['top', 'bottom', 'left', 'right']:
        side = getattr(cell.border, s)
        if side.style and side.color and side.color.rgb:
            colors.append(str(side.color.rgb))
    return Counter(colors).most_common(1)[0][0] if colors else None


# ═══════════════════════════════════════════════════════
# TASK 1: BLUE FILL CLEANUP
# ═══════════════════════════════════════════════════════
print("=" * 60)
print("TASK 1: BLUE FILL CLEANUP")
print("=" * 60)

total_cleaned = 0
for sn in wb.sheetnames:
    ws = wb[sn]
    dd_cells = get_dropdown_cells(ws)
    cleaned = 0
    for row in range(1, min(ws.max_row + 1, 200)):
        for col in range(1, min(ws.max_column + 1, 20)):
            if is_mc(ws, row, col):
                continue
            cell = ws.cell(row=row, column=col)
            if not cell.fill or not cell.fill.fgColor:
                continue
            rgb = str(cell.fill.fgColor.rgb) if cell.fill.fgColor.rgb else ''
            if 'D6EAF8' not in rgb:
                continue
            if (row, col) in dd_cells:
                continue  # Valid dropdown — keep blue

            # Stray blue cell — find surrounding section fill
            replacement = PatternFill()  # Default: no fill
            for dr, dc in [(-1, 0), (1, 0), (0, -1), (0, 1), (-2, 0), (2, 0)]:
                r2, c2 = row + dr, col + dc
                if r2 < 1 or c2 < 1 or r2 > ws.max_row or c2 > ws.max_column:
                    continue
                if is_mc(ws, r2, c2):
                    continue
                nb = ws.cell(row=r2, column=c2)
                if not nb.fill or not nb.fill.patternType:
                    continue
                nb_rgb = str(nb.fill.fgColor.rgb) if nb.fill.fgColor else ''
                if any(x in nb_rgb for x in ['D6EAF8', '464543', '846E60', 'C19C81', '00000000']):
                    continue
                # Copy fill properties into a new PatternFill
                replacement = PatternFill(
                    patternType=nb.fill.patternType,
                    fgColor=str(nb.fill.fgColor.rgb) if nb.fill.fgColor else None,
                    bgColor=str(nb.fill.bgColor.rgb) if nb.fill.bgColor else None,
                )
                break

            cell.fill = replacement
            cleaned += 1
            coord = f"{get_column_letter(col)}{row}"
            val = cell.value
            val_str = f" (val: {str(val)[:40]})" if val else " (empty)"
            print(f"  {sn}: {coord}{val_str} — blue removed")

    print(f"  {sn}: {cleaned} stray blue cells cleaned")
    total_cleaned += cleaned

print(f"\nTotal cleaned: {total_cleaned}")


# ═══════════════════════════════════════════════════════
# TASK 2: BORDER FIXES
# ═══════════════════════════════════════════════════════
print()
print("=" * 60)
print("TASK 2: BORDER FIXES")
print("=" * 60)

grand_total = 0

for sn in wb.sheetnames:
    ws = wb[sn]

    # Find default border color for this tab (most common)
    color_counts = Counter()
    for row in range(1, min(ws.max_row + 1, 200)):
        for col in range(1, min(ws.max_column + 1, 20)):
            if is_mc(ws, row, col):
                continue
            cell = ws.cell(row=row, column=col)
            if has_full_borders(cell):
                c = get_border_color(cell)
                if c:
                    color_counts[c] += 1
    default_color = color_counts.most_common(1)[0][0] if color_counts else 'FFD4C5B0'

    fixed_rows = {}
    tab_fixed = 0

    def row_needs_borders(ws, row):
        """Check if a row should have borders: either ≥50% of its cells
        already have them, OR nearby rows (within ±5) have bordered cells
        in similar columns (indicating this row is part of a table)."""
        content_cols = []
        full_count = 0
        for col in range(1, min(ws.max_column + 1, 20)):
            if is_mc(ws, row, col):
                continue
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                content_cols.append(col)
                if has_full_borders(cell):
                    full_count += 1
        if not content_cols:
            return False, []
        # Pass 1: ≥50% already bordered
        if full_count >= len(content_cols) * 0.5:
            return True, content_cols
        # Pass 2: check if nearby rows have bordered cells in overlapping columns
        for dist in range(1, 6):
            for dr in [dist, -dist]:
                r2 = row + dr
                if r2 < 1 or r2 > min(ws.max_row, 200):
                    continue
                for col in content_cols:
                    if is_mc(ws, r2, col):
                        continue
                    ref = ws.cell(row=r2, column=col)
                    if has_full_borders(ref):
                        return True, content_cols
        return False, content_cols

    for row in range(1, min(ws.max_row + 1, 200)):
        needs, content_cols = row_needs_borders(ws, row)
        if not needs:
            continue

        for col in content_cols:
            cell = ws.cell(row=row, column=col)
            if has_full_borders(cell):
                continue

            # Find reference border color from nearest neighbor
            ref_color = None
            for dist in range(1, 6):
                for dr, dc in [(0, -1), (0, 1), (-1, 0), (1, 0)]:
                    r2, c2 = row + dr * dist, col + dc * dist
                    if r2 < 1 or c2 < 1 or r2 > ws.max_row or c2 > ws.max_column:
                        continue
                    if is_mc(ws, r2, c2):
                        continue
                    ref = ws.cell(row=r2, column=c2)
                    if has_full_borders(ref):
                        ref_color = get_border_color(ref)
                        break
                if ref_color:
                    break
            if not ref_color:
                ref_color = default_color

            side = Side(style='thin', color=ref_color)
            existing = cell.border
            top = existing.top if existing.top.style == 'medium' else side
            bottom = existing.bottom if existing.bottom.style == 'medium' else side
            left = existing.left if existing.left.style == 'medium' else side
            right = existing.right if existing.right.style == 'medium' else side
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)

            fixed_rows[row] = fixed_rows.get(row, 0) + 1
            tab_fixed += 1

    print(f"\n{sn}:")
    if tab_fixed == 0:
        print(f"  No border fixes needed")
    else:
        rows = sorted(fixed_rows.keys())
        ranges = []
        start = end = rows[0]
        for r in rows[1:]:
            if r <= end + 2:
                end = r
            else:
                ranges.append((start, end))
                start = end = r
        ranges.append((start, end))

        for s, e in ranges:
            count = sum(fixed_rows.get(r, 0) for r in range(s, e + 1))
            if s == e:
                print(f"  Row {s}: {count} cells")
            else:
                print(f"  Rows {s}-{e}: {count} cells")
        print(f"  Total: {tab_fixed} cells fixed")
    grand_total += tab_fixed

print(f"\nGrand total border fixes: {grand_total}")


# ═══════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════
print("\nSaving...")
wb.save(SRC)


# ═══════════════════════════════════════════════════════
# VERIFICATION
# ═══════════════════════════════════════════════════════
print()
print("=" * 60)
print("VERIFICATION")
print("=" * 60)

wb2 = openpyxl.load_workbook(SRC)
for sn in wb2.sheetnames:
    ws = wb2[sn]
    dd_cells = get_dropdown_cells(ws)
    stray_blue = 0
    for row in range(1, min(ws.max_row + 1, 200)):
        for col in range(1, min(ws.max_column + 1, 20)):
            if is_mc(ws, row, col):
                continue
            cell = ws.cell(row=row, column=col)
            if cell.fill and cell.fill.fgColor and 'D6EAF8' in str(cell.fill.fgColor.rgb):
                if (row, col) not in dd_cells:
                    stray_blue += 1

    no_border = partial = full = 0
    for row in range(1, min(ws.max_row + 1, 200)):
        for col in range(1, min(ws.max_column + 1, 20)):
            if is_mc(ws, row, col):
                continue
            cell = ws.cell(row=row, column=col)
            if cell.value is None:
                continue
            sides = [cell.border.top.style, cell.border.bottom.style,
                     cell.border.left.style, cell.border.right.style]
            has = sum(1 for s in sides if s is not None)
            if has == 0:
                no_border += 1
            elif has < 4:
                partial += 1
            else:
                full += 1
    total = no_border + partial + full
    pct = 100 * full / total if total > 0 else 0

    print(f"  {sn}: stray_blue={stray_blue}, "
          f"borders: {full}/{total} full ({pct:.0f}%), "
          f"{partial} partial, {no_border} none")

print("\nDONE")
